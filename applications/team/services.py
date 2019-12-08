# coding=utf-8

from django.db import transaction
import json
import logging

from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import quote_sheetname

from django.db.models import Q
from openpyxl.styles import Font, Alignment

from applications.activity.models import ExpertActivity, Rule
from applications.activity.share import is_activity_owner
from applications.common.services import is_superuser, area_by_pov, area_name
from applications.expert.models import Expert
from applications.user.models import Area
from models import Team, TeamExpert
from applications.work.models import Work
from applications.score.models import Score, FinalScore
from utils.const_err import *
from utils.const_def import *
from utils.file_fun import gen_path
from utils.public_fun import paging_by_page, seq2list
from utils.utils_except import BusinessException

logger = logging.getLogger(__name__)


def list_team(activity, expert=None, keyword=None, judger=None, id_list=None):
    qs = Team.objects.filter(activity=activity)
    if id_list:
        qs = qs.filter(id__in=id_list)
    if keyword:
        qs = qs.filter(name__contains=keyword)
    if expert:
        expert_team = TeamExpert.objects.filter(
                expert=expert, team__activity=activity).values_list('team__id', flat=True)
        qs = qs.filter(id__in=expert_team)
    if judger:
        with_judger_team = TeamExpert.objects.filter(
                expert__account__name__contains=judger, team__activity=activity).values_list('team__id', flat=True)
        qs = qs.filter(id__in=with_judger_team)
    return qs


def list_team_by_super(account, activity, keyword, judger, page, rows, id_list=None):
    rule = Rule.objects.filter(activity=activity).first()
    if not rule:
        raise BusinessException(ERR_RULE_NOT_EXIST)
    if not is_activity_owner(activity, account):
        raise BusinessException(ERR_USER_AUTH)
    raw = list_team(activity, keyword=keyword, judger=judger, id_list=id_list)

    paged_data, result = paging_by_page(raw, rows, page)
    rule = Rule.objects.filter(activity=activity).first()
    if not rule:
        raise BusinessException(ERR_RULE_NOT_EXIST)
    max = str(rule.parse_rule().expert_count())
    for each in paged_data:
        d = detail_team(each, max)
        result['items'].append(d)
    result['id_all'] = ','.join([str(each.id) for each in raw])
    result['judger_max'] = str(max)
    return result


def export_team_by_super(account, activity, user, role, team_id_list):
    rule = Rule.objects.filter(activity=activity).first()
    if not rule:
        raise BusinessException(ERR_RULE_NOT_EXIST)
    max = rule.parse_rule().expert_count()

    wb = Workbook()
    sheet = wb.active

    expert_info_map = {}

    sheet.cell(column=1, row=1, value=u'组名')
    sheet.cell(column=2, row=1, value=u'组内作品数')
    for i in xrange(max):
        col = 3+i
        sn = 1+i
        sheet.cell(column=col, row=1, value=u'评审专家%s'%(sn))
        expert_info_map[sn] = col

    teams = list_team(activity, keyword=None, judger=None, id_list=team_id_list)
    for r, each in enumerate(teams):
        cur_row = 2+r
        d = detail_team(each, max)
        judger_dict = {int(judger['sn']):judger['expert_name'] for judger in d['judger_list']}

        sheet.cell(column=1, row=cur_row, value=d['team_name'])
        sheet.cell(column=2, row=cur_row, value=d['team_work_count'])
        for i in xrange(max):
            sn = i+1
            col = expert_info_map[sn]
            sheet.cell(column=col, row=cur_row, value=judger_dict[sn])
    return wb


def list_team_by_judger(account, activity, as_leader, page, rows, id_list=None):
    e_a = ExpertActivity.objects.filter(
            expert__del_flag=FALSE_INT, activity=activity, expert__account=account).first()
    if not e_a:
        raise BusinessException(ERR_USER_AUTH)
    expert = e_a.expert
    qs = list_team(activity, expert=expert)
    if as_leader and str(as_leader) == TRUE_STR:
        teams_as_leader = TeamExpert.objects.filter(
                expert=expert, team__activity=activity, is_leader=TRUE_INT).values_list('team__id', flat=True)
        qs = qs.filter(id__in=teams_as_leader)
    else:
        teams_as_member = TeamExpert.objects.filter(
                expert=expert, team__activity=activity).values_list('team__id', flat=True)
        qs = qs.filter(id__in=teams_as_member)

    paged_data, result = paging_by_page(qs, rows, page)
    for each in paged_data:
        result['items'].append({
            'team_id': str(each.id),
            'team_name': each.name,
        })
    return result


def detail_team(team, max):
    experts_in_team = TeamExpert.objects.filter(team=team, expert__del_flag=FALSE_INT).order_by('sn')
    count = Work.objects.filter(team=team).count()
    result = {
        'team_id': str(team.id),
        'team_name': team.name,
        'team_work_count': str(count),
        'judger_max': max,
        'judger_list': list(),
    }
    expert_max = Rule.objects.filter(activity=team.activity).first().parse_rule().expert_count()
    for i in xrange(expert_max):
        sn = i+1
        for et in experts_in_team:
            if et.sn == sn:
                result['judger_list'].append({
                    'expert_id': str(et.expert.id),
                    'expert_name': et.expert.account.name,
                    'sn': str(et.sn),
                    'is_leader': str(et.is_leader),
                    'expert_position': et.expert.position,
                    'expert_area_id': str(et.expert.area_id),
                    'expert_area_name_full': area_name(et.expert.area_id),
                })
                break
        else:
            # 某位置没有专家，补一个空
            result['judger_list'].append({
                'expert_id': '',
                'expert_name': '',
                'sn': str(sn),
                'is_leader': '',
                'expert_position': '',
                'expert_area_id': '',
                'expert_area_name_full': '',
            })

    return result


def available_add_expert_in_team(account, team, expert_in_same_team, keyword, page, rows):
    # # load expert from expert_lib
    # from applications.expert.services import list_expert_user
    # expert_in_lib = [each['id'] for each in list_expert_user(account, user.id, None, None, None)]

    expert_in_acti = ExpertActivity.objects.filter(activity=team.activity, expert__del_flag=FALSE_INT)
    if keyword:
        expert_in_acti = expert_in_acti.filter(Q(expert__account__username__contains=keyword) | Q(expert__account__name__contains=keyword))
    expert_in_acti = expert_in_acti.values_list('expert__id', flat=True)
    expert_exist_in_team = TeamExpert.objects.filter(team__activity=team.activity).exclude(team=team).values_list('expert__id', flat=True)
    result_id_list = set(expert_in_acti) - set(expert_exist_in_team) - set(expert_in_same_team)
    qs = Expert.objects.filter(id__in=result_id_list, del_flag=FALSE_INT)
    paged_data, result = paging_by_page(qs, rows, page)
    for each in qs:
        result['items'].append({
            'expert_id': str(each.id),
            'expert_name': each.account.name,
            'expert_mobile': each.account.username,
            'expert_position': each.position,
            'expert_area_id': str(each.area.id) if each.area else '',
            'expert_area_name_full': area_name(each.area.id) if each.area else '',
        })
    return result


@transaction.atomic()
def update_expert_in_team(account, user, role, team, new_expert_json):
    if not is_activity_owner(team.activity, account):
        raise BusinessException(ERR_USER_AUTH)

    """
        [{"sn":"2", "expert_id": "1167"}]
    """
    new_expert_list = json.loads(new_expert_json)
    new_expert_id_list = [int(each['expert_id']) for each in new_expert_list]
    new_expert_id_dict = {each['expert_id']: each['sn'] for each in new_expert_list}
    t_e_old = TeamExpert.objects.filter(team=team).all()
    handled = list()
    for old in t_e_old:
        if old.expert.id in new_expert_id_list:
            new_sn = new_expert_id_dict[str(old.expert.id)]
            old.sn = new_sn
            if int(new_sn) == 1:
                old.is_leader = TRUE_INT
            else:
                old.is_leader = FALSE_INT
            handled.append(old.expert.id)
            old.save()
        else:
            Score.objects.filter(del_flag=FALSE_INT, team_expert=old).update(del_flag=TRUE_INT)
            FinalScore.objects.filter(del_flag=FALSE_INT, team_expert=old).update(del_flag=TRUE_INT)
            old.del_flag=TRUE_INT
            old.save()
    new_add = set(new_expert_id_list) - set(handled)
    for each in new_add:
        is_leader = TRUE_INT if int(new_expert_id_dict[str(each)]) == 1 else FALSE_INT
        TeamExpert.objects.create(team=team, expert_id=int(each), sn=new_expert_id_dict[str(each)], is_leader=is_leader)

    # 增加或者减少组内专家，更新最终得分数据（针对平均分制）
    # TODO

    return TRUE_INT


def list_expert_area_in_team(account, team):
    owner_user = is_activity_owner(team.activity, account)
    if not owner_user:
        raise BusinessException(ERR_USER_AUTH)
    result = list()
    area_id_list_raw = set(TeamExpert.objects.filter(team=team).values_list('expert__area__id', flat=True))
    area_list_raw = Area.objects.filter(del_flag=FALSE_INT, id__in=area_id_list_raw)
    for each in area_list_raw:
        area = area_by_pov(each, owner_user.area)
        if area:
            result.append({
                'area_id': str(area.id), 'area_name': area.area_name
            })
    return result


def edit_team(account, activity, team, new_name):
    if not is_activity_owner(activity, account):
        raise BusinessException(ERR_USER_AUTH)
    if team:  # MODIFY
        if Team.objects.filter(activity=activity, name=new_name).exclude(id=team.id).exists():
            raise BusinessException(ERR_TEAM_NAME_CONFLICT)
        team.name = new_name
        team.save()
    else:  # CREATE
        Team.objects.create(activity=activity, name=new_name)
    return TRUE_INT


@transaction.atomic
def delete_team(account, activity, team_list):
    if not is_activity_owner(activity, account):
        raise BusinessException(ERR_USER_AUTH)
    succ_count = 0
    for each in team_list:
        if Work.objects.filter(activity=activity, team=each).exists():
            logger.warn('team %s contains works, so can not delete' % each.id)
            continue
        succ_count += 1
        each.del_flag=TRUE_INT
        each.save()
        TeamExpert.objects.filter(team=each).update(del_flag=TRUE_INT)
    msg1 = (u'%d个分组删除成功' % succ_count) if succ_count > 0 else ''
    fail_count = len(team_list) - succ_count
    msg2 = (u'%d个分组删除失败或不支持删除' % fail_count) if fail_count > 0 else ''
    return msg1 + msg2


@transaction.atomic
def add_team_work(account, activity, team, work_id_list):
    # TODO check auth
    #
    Work.objects.filter(id__in=work_id_list, activity=activity).update(team=team, status=WORK_STATUS_REVIEWING[0])
    count = Work.objects.filter(activity=activity, team=team).count()
    team.work_count=count
    team.save()
    return TRUE_INT


@transaction.atomic
def remove_team_work(account, activity, team, work_id_list):
    # TODO check auth
    #
    Work.objects.filter(id__in=work_id_list, activity=activity, team=team). \
                update(team_id=None, status=WORK_STATUS_HAVE_EXAMINED[0])
    Score.objects.filter(work_id__in=work_id_list).update(del_flag=TRUE_INT)
    FinalScore.objects.filter(work_id__in=work_id_list).update(del_flag=TRUE_INT)
    return TRUE_INT


