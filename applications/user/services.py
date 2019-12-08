#!/usr/bin python
# coding=utf-8
import logging
import re
import json
import openpyxl
from django.contrib import auth
from django.db import transaction
from django.db.models import Q
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter, quote_sheetname
from openpyxl.worksheet.datavalidation import DataValidation

from applications.common.services import get_account_activity_role, get_areadetail_by_id, get_child_by_areaid, get_rank_child_area_name, \
    get_regiondetail_by_id, update_area_fullname, area_bind_region
from applications.expert.models import Expert
from applications.expert.services import list_expert_user, add_expert_user
from applications.team.models import TeamExpert
from applications.upload_resumable.err_code import ERR_SUCCESS
from applications.user.sms_netease import send_message, verify_messagecode, del_messagecode
from applications.weixinmp.models import WeixinScanConfirm, WeixinAccount
from models import *
from applications.activity.models import Activity, Role, ExpertActivity
from utils.const_def import *
from utils.const_err import *
from utils.file_fun import get_image_url, gen_path
from utils.public_fun import get_pages, get_lastid_index
from utils.utils_except import BusinessException
from utils.utils_type import is_mobile

logger = logging.getLogger(__name__)


def detail_account(user):
    user_list_info = User.objects.filter(del_flag=FLAG_NO, account=user). \
        values("id", "area_id", "area__area_level", "area__area_name", "name", "sex").distinct()
    user_list_info = list(user_list_info)
    user_list = []
    for item in user_list_info:
        dict_tmp = {
            "user_id": item['id'],
            "area_id": item['area_id'],
            "area_level": item['area__area_level'],
            "name": item['name'],
            'sex': item['sex'],
            "area_name": item['area__area_name']
        }
        user_list.append(dict_tmp)

    # 查询是否显示我的参赛标签
    accountactivity = Activity.objects.filter(del_flag=FALSE_INT, user__account=user)
    is_show_mycreateactivity = '1' if accountactivity else '0'

    # 是否显示创建过赛事按钮
    accountright = AccountRight.objects.filter(del_flag=FALSE_INT, account=user)
    is_show_createactivity = '1' if accountright else '0'

    # 如果需要用户确认信息，需要告知前端确认谁的信息，如果是用户，只用确认姓名和性别，如果是专家，需要确认姓名性别地区职务
    area_id = ''
    position = ''
    nation = ''
    province = ''
    city = ''
    county = ''
    institution = ''
    manage_direct = ''
    weixin_bind_status = '0'
    expert = Expert.objects.filter(del_flag=FALSE_INT, account=user).first()
    if expert:
        area_id = expert.area_id
        position = expert.position
        areadetail = get_areadetail_by_id(area_id)
        nation = areadetail['nation']
        province = areadetail['province']
        city = areadetail['city']
        county = areadetail['country']
        institution = areadetail['institution']
        manage_direct = areadetail['manage_direct']

    # data_confirm_mod = ''
    # if not user.is_data_confirm:
    #     data_confirm_mod = 'EXPERT' if expert else 'USER'

    userauth = bin(user.auth)[2:]
    userauth_namelist = get_userauth_namelist(user.auth)

    # 运维人员返回为空，专家返回‘EXPERT’，用户返回‘USER’
    if userauth_namelist:
        data_confirm_mod = ''
    elif expert:
        data_confirm_mod = 'EXPERT'
    else:
        data_confirm_mod = 'USER'

    return {
        "id": user.id,
        "name": user.name,
        "username": user.username,
        "mobile": user.mobile,
        "sex": user.sex,
        "email": user.email,
        "is_self_reg": user.is_self_reg,  # 是否自主注册
        "region_id": user.region_id if user.region_id else '',  # 归属区域id
        "region_name": user.region.region_name if user.region else '',  # 归属区域名称
        "region_fullname": user.region.region_fullname if user.region else '',
        "area_id": area_id,
        "nation": nation,
        "province": province,
        "city": city,
        "county": county,
        "institution": institution,
        "manage_direct": manage_direct,
        "position": position,
        "image_url": get_image_url(user.image.url) if user.image else "",
        "user_list": user_list,
        "activity_mask": user.activity_mask,
        "is_data_confirm": user.is_data_confirm,
        "data_confirm_mod": data_confirm_mod,
        "is_show_mycreateactivity": is_show_mycreateactivity,
        "is_show_createactivity": is_show_createactivity,
        "auth": userauth,  # 转为二进制后，每一位分别代表一个菜单权限。
        "userauth_namelist": userauth_namelist,
        "is_bind_weixin": '1' if is_bind_weixin(user) else '0',
    }


def is_bind_weixin(user):
    weixinaccount = WeixinAccount.objects.filter(del_flag=FALSE_INT, account=user)
    if weixinaccount:
        return True
    return False


def get_userauth_namelist(userauth):
    userauth_namelist = []
    if userauth & EDITOR_RIGHT_ACTIVITY[0] == EDITOR_RIGHT_ACTIVITY[0]:
        userauth_namelist.append(EDITOR_RIGHT_ACTIVITY[1])
    if userauth & EDITOR_RIGHT_NEWS[0] == EDITOR_RIGHT_NEWS[0]:
        userauth_namelist.append(EDITOR_RIGHT_NEWS[1])
    if userauth & EDITOR_RIGHT_USERS[0] == EDITOR_RIGHT_USERS[0]:
        userauth_namelist.append(EDITOR_RIGHT_USERS[1])
    if userauth & EDITOR_RIGHT_TEMPLATES[0] == EDITOR_RIGHT_TEMPLATES[0]:
        userauth_namelist.append(EDITOR_RIGHT_TEMPLATES[1])
    return userauth_namelist


def list_user(user, user_id, area_name, name, is_show_store='1', activity_id=None, only_can_add=None):
    user_obj = User.objects.filter(del_flag=FLAG_NO, id=user_id).first()
    if user != user_obj.account:
        return dict(c=ERR_USER_AUTH[0], m=ERR_USER_AUTH[1])
    user_list = User.objects.filter(del_flag=FLAG_NO, area__parent=user_obj.area, parent_account=user)
    if only_can_add == '1':
        user_in_activity = Role.objects.filter(del_flag=FALSE_INT, activity_id=activity_id).values_list('user_id', flat=True)
        user_in_activity = list(user_in_activity)
        user_list = user_list.exclude(id__in=user_in_activity)
    if is_show_store:
        user_list = user_list.filter(is_show_store=is_show_store)
    if area_name:
        if area_name in ['国直属', '省直属', '市直属']:
            user_list = user_list.filter(area__manage_direct=TRUE_INT)
        else:
            user_list = user_list.filter(area__area_name=area_name)
    if name:
        # user_list = user_list.filter(name__contains=name)
        user_list = user_list.filter(Q(name__contains=name) | Q(account__mobile__contains=name) | Q(account__username__contains=name))
    user_info_list = user_list.values("id", "name", "sex", "account__username", "account__mobile", "area__area_name",
                                      "account_id", "area__manage_direct", "area_id", "area__area_fullname").order_by("-update_time")
    user_info_list = list(user_info_list)
    data = []
    for item in user_info_list:
        dict_tmp = {"id": item["id"], "name": item['name'], "sex": item['sex'], "username": item['account__username'],
                    'mobile': item['account__mobile'], "area_name": item['area__area_name'],
                    "area_fullname": item['area__area_fullname'],
                    "account_id": item['account_id']}
        if item["area__manage_direct"]:
            if user_obj.area.area_level == LEVEL_PROVINCE:
                dict_tmp["area_name"] = u"省直属-" + dict_tmp['area_name']
            elif user_obj.area.area_level == LEVEL_CITY:
                dict_tmp["area_name"] = u"市直属-" + dict_tmp['area_name']
            elif user_obj.area.area_level == LEVEL_NATION:
                dict_tmp["area_name"] = u"国直属-" + dict_tmp['area_name']
            else:
                pass
        area_detail = get_areadetail_by_id(item['area_id'])
        dict_tmp['area_id'] = area_detail['area_id']
        dict_tmp['nation'] = area_detail['nation']
        dict_tmp['province'] = area_detail['province']
        dict_tmp['city'] = area_detail['city']
        dict_tmp['country'] = area_detail['country']
        dict_tmp['institution'] = area_detail['institution']
        dict_tmp['manage_direct'] = area_detail['manage_direct']
        data.append(dict_tmp)
    return dict(c=SUCCESS[0], m=SUCCESS[1], d=data)


def login(request, username="", password=""):
    username = username.strip()
    password = password.strip()
    user = auth.authenticate(username=username, password=password)
    if user is not None:
        auth.login(request, user)
        return True
    return False


def api_confirmcode_login(request, confirm_code):
    # 检查确认码是否登陆生成的
    weixinscanconfirm = WeixinScanConfirm.objects.filter(code=confirm_code, del_flag=FALSE_INT, busitype=WEIXIN_SCAN_TYPE_LOGIN).first()
    if not weixinscanconfirm:
        raise BusinessException(ERR_QRCODE_TIMEOUT)

    # 检查是否微信扫过码
    if not weixinscanconfirm.openid or not weixinscanconfirm.openid_fh:
        raise BusinessException(ERR_WEIXIN_NOT_SCAN)

    # 检查扫码的微信是否绑定过用户
    weixinaccount = WeixinAccount.objects.filter(openid=weixinscanconfirm.openid, del_flag=FALSE_INT).first()
    if not weixinaccount:
        raise BusinessException(ERR_WEIXIN_IS_NOT_BIND)

    # 用户登陆
    weixinaccount.account.backend = settings.AUTHENTICATION_BACKENDS[0]
    auth.login(request, weixinaccount.account)

    # 将确认码失效，否则可以反复登陆，有安全风险
    weixinscanconfirm.del_flag = TRUE_INT
    weixinscanconfirm.status = WEIXIN_CONFIRM_STATUS_CONFIRM
    weixinscanconfirm.save()
    return True


def __check_username_valid(username):
    username = username.strip()
    if re.match('[gl]', username, re.I) and not re.search(r'\W', username) and len(username) > 12 and len(username) < 21:
        return {"c": SUCCESS[0], "m": SUCCESS[1], "d": []}
    else:
        if len(username) != 11 or not username.isdigit():
            return {"c": ERR_USER_USERNAME_ERROR[0], "m": ERR_USER_USERNAME_ERROR[1], "d": []}
    return {"c": SUCCESS[0], "m": SUCCESS[1], "d": []}


def check_username(user, username):
    err_code = __check_username_valid(username)
    if err_code['c'] != SUCCESS[0]:
        return err_code

    if Account.objects.filter(username=username, del_flag=0).exists():
        return {"c": SUCCESS[0], "m": SUCCESS[1], "d": ["True"]}
    else:
        return {"c": SUCCESS[0], "m": SUCCESS[1], "d": ["False"]}


def add_user(user, username, name, sex, area_id, direct_area_id, institution):
    result = dict()
    # user_now = User.objects.filter(del_flag=FLAG_NO, id=current_user_id).first()
    # if user != user_now.account:
    #     return dict(c=ERR_USER_AUTH[0], m=ERR_USER_AUTH[1])

    # 注意institution机构名称、direct_area_id直属地区id，比如湖北省直属则需要direct_area_id传入湖北省id、area_id非直属并非机构传此参数三者传值方式
    # 1、添加各级直属机构专家/用户，需要传入direct_area_id、institution
    # 2、添加省、市、区县专家/用户，只需要传area_id
    # 3、添加区县下机构的专家/用户，可以只传入direct_area_id、institution,也可只传入area_id
    # 4、添加机构下的普通用户，需要传入direct_area_id，
    if direct_area_id and int(direct_area_id):
        direct_area = Area.objects.get(id=direct_area_id, del_flag=FALSE_INT)
        if direct_area.area_level == LEVEL_INSTITUTION or direct_area.manage_direct == TRUE_INT:
            # 机构添加普通用户/专家
            area_zone, _ = Area.objects.get_or_create(del_flag=FLAG_NO, area_name=USER_TYPE_NORMAL, parent_id=direct_area.id,
                                                      manage_direct=FLAG_NO, area_level=LEVEL_NONE)
        elif direct_area.area_level == LEVEL_COUNTY:
            # 添加区县下的机构用户/专家
            area_zone, _ = Area.objects.get_or_create(del_flag=FLAG_NO, area_name=institution, parent_id=direct_area.id,
                                                      manage_direct=FLAG_NO, area_level=direct_area.area_level >> 1)
            area_bind_region(area_zone)
        else:
            # 地区加直属机构用户/专家
            area_zone, _ = Area.objects.get_or_create(del_flag=FLAG_NO, area_name=institution, parent_id=direct_area.id,
                                                      manage_direct=FLAG_YES, area_level=direct_area.area_level >> 1)
            area_bind_region(area_zone)
    elif area_id and int(area_id):
        cur_area = Area.objects.get(id=area_id, del_flag=FALSE_INT)
        area_zone = cur_area
        # if cur_area.area_level >= LEVEL_INSTITUTION:
        #     area_zone = cur_area
        # else:
        #     raise BusinessException(ERR_USER_AREA)
    else:
        raise BusinessException(ERR_REQUEST_PARAMETER_ERROR)

    if not area_zone:
        return dict(c=REQUEST_PARAM_ERROR[0], m=REQUEST_PARAM_ERROR[1])

    update_area_fullname(area_zone.id)

    # 检查用户名格式是否正确
    resp_code = __check_username_valid(username)
    if resp_code['c'] != SUCCESS[0]:
        return resp_code

    # 添加用户
    account = Account.objects.filter(del_flag=FLAG_NO, username=username).first()
    if not account:
        account = Account.objects.create_user(username, password=settings.DEFAULT_PASSWORD)
        account.name = name if name else ""
        if sex:  account.sex = sex
        if re.match('[gl]', username, re.I):
            account.code = username
        elif len(username) == 11 and username.isdigit():
            account.mobile = username
        account.save()
        useradd = User.objects.create(account=account, name=name, sex=sex, area=area_zone, parent_account=user, is_show_store=1)
        result['user_id'] = useradd.id
    else:
        # if User.objects.filter(del_flag=FLAG_NO, account=account, area__area_level=area_zone.area_level).exists():
        #     raise Exception(u"同等级用户已存在，创建冲突。")
        useradd = User.objects.filter(del_flag=FLAG_NO, account=account, area=area_zone, parent_account=user).first()
        if useradd:
            logger.info(u"同等级用户已存在，不需要再创建用户。")
            useradd.is_show_store = 1
            useradd.save()
        else:
            useradd = User.objects.create(account=account, name=name, sex=sex, area=area_zone, parent_account=user, is_show_store=1)
        result['user_id'] = useradd.id

    # 如果用户尚未确认资料，则以新资料为准。
    if not account.is_data_confirm:
        account.name = name
        account.sex = sex
        account.save()

        useradd.name = name
        useradd.sex = sex
        useradd.save()

    return dict(c=SUCCESS[0], m=SUCCESS[1], d=result)


@transaction.atomic
def mod_user(user, user_id, username, name, sex, area_id, direct_area_id, institution):
    result = ''
    # 查询当前登录的用户
    # user_now = User.objects.filter(del_flag=FLAG_NO, id=current_user_id).first()
    # if user != user_now.account:
    #     return dict(c=ERR_USER_AUTH[0], m=ERR_USER_AUTH[1])

    # 注意institution机构名称、direct_area_id直属地区id，比如湖北省直属则需要direct_area_id传入湖北省id、area_id非直属并非机构传此参数三者传值方式
    # 1、添加各级直属机构专家/用户，需要传入direct_area_id、institution
    # 2、添加省、市、区县专家/用户，只需要传area_id
    # 3、添加区县下机构的专家/用户，可以只传入direct_area_id、institution,也可只传入area_id
    # 4、添加机构下的普通用户，需要传入direct_area_id，
    if direct_area_id and int(direct_area_id):
        direct_area = Area.objects.get(id=direct_area_id, del_flag=FALSE_INT)
        if direct_area.area_level == LEVEL_INSTITUTION or direct_area.manage_direct == TRUE_INT:
            # 机构添加普通用户
            area_zone, _ = Area.objects.get_or_create(del_flag=FLAG_NO, area_name=USER_TYPE_NORMAL, parent_id=direct_area.id,
                                                      manage_direct=FLAG_NO, area_level=LEVEL_NONE)
        elif direct_area.area_level == LEVEL_COUNTY:
            # 添加区县下的机构用户
            area_zone, _ = Area.objects.get_or_create(del_flag=FLAG_NO, area_name=institution, parent_id=direct_area.id,
                                                      manage_direct=FLAG_NO, area_level=direct_area.area_level >> 1)
            area_bind_region(area_zone)
        else:
            # 地区加直属机构
            area_zone, _ = Area.objects.get_or_create(del_flag=FLAG_NO, area_name=institution, parent_id=direct_area.id,
                                                      manage_direct=FLAG_YES, area_level=direct_area.area_level >> 1)
            area_bind_region(area_zone)
    elif area_id and int(area_id):
        cur_area = Area.objects.get(id=area_id, del_flag=FALSE_INT)
        area_zone = cur_area
        # if cur_area.area_level >= LEVEL_INSTITUTION:
        #     area_zone = cur_area
        # else:
        #     raise BusinessException(ERR_USER_AREA)
    else:
        raise BusinessException(ERR_REQUEST_PARAMETER_ERROR)

    if not area_zone:
        # 修改的用户地区不是当前用户的区域直接下级
        raise BusinessException(ERR_AREA_ERROR)

    update_area_fullname(area_zone.id)

    # 检查用户名格式是否正确
    if username:
        resp_code = __check_username_valid(username)
        if resp_code['c'] != SUCCESS[0]:
            return resp_code

    # 更新USER信息
    moduser = User.objects.get(id=user_id, del_flag=FALSE_INT)

    # if moduser.account.is_data_confirm:
    #     raise BusinessException(ERR_ACCOUNTDATA_IS_CONFIRM)

    if not moduser.account.is_data_confirm or moduser.account == user:
        if name:
            moduser.name = name
        if sex:
            moduser.sex = sex
        result = u'仅更新地区信息，未更新用户资料，原因是用户已经确认资料信息，仅用户自己可以修改资料！'
    moduser.area = area_zone
    moduser.save()

    # 检查用户名是否已存在,除了自己
    if username:
        account = Account.objects.filter(del_flag=FLAG_NO, username=username).exclude(id=moduser.account_id).first()
        if account:
            raise BusinessException(ERR_USERNAME_DUPLICATE)

    # 更新account信息
    modaccount = Account.objects.get(id=moduser.account_id, del_flag=FALSE_INT)
    if not moduser.account.is_data_confirm:
        if name:
            modaccount.name = name
        if sex:
            modaccount.sex = sex
    if username:
        if re.match('[gl]', username, re.I):
            modaccount.code = username
        elif len(username) == 11 and username.isdigit():
            modaccount.mobile = username
    modaccount.save()

    return dict(c=SUCCESS[0], m=SUCCESS[1], d=result)


@transaction.atomic
def reset_others_password(user, account_id="", admin_password="", new_password=settings.DEFAULT_PASSWORD):
    admin_password = admin_password.strip()
    new_password = new_password.strip()
    if not new_password:
        new_password = settings.DEFAULT_PASSWORD

    if not user.check_password(admin_password):
        return {"c": -1, "m": u"管理员密码不正确", "d": []}

    if len(new_password) < 6:
        return {"c": -1, "m": u"新密码长度太短", "d": []}

    account_id = int(account_id)
    account = Account.objects.filter(id=account_id, del_flag=FLAG_NO).first()
    if not account:
        return {"c": ERR_USER_NOT_EXIST[0], "m": ERR_USER_NOT_EXIST[1], "d": []}

    account.set_password(new_password)
    account.encoded_pwd = xor_crypt_string(data=new_password, encode=True)
    account.save()
    return {"c": SUCCESS[0], "m": SUCCESS[1], "d": []}


@transaction.atomic
def reset_own_password(user, old_password="", new_password=""):
    old_password = old_password.strip()
    new_password = new_password.strip()

    if not user.check_password(old_password):
        return {"c": ERR_OLD_PASSWORD_ERROR[0], "m": ERR_OLD_PASSWORD_ERROR[1], "d": []}

    if len(new_password) < 6:
        return {"c": -1, "m": u"新密码长度太短", "d": []}

    user.set_password(new_password)
    user.encoded_pwd = xor_crypt_string(data=new_password, encode=True)
    user.save()
    return {"c": SUCCESS[0], "m": SUCCESS[1], "d": []}


def delete_user(user, cur_user_id, user_id_list, del_all_user=None, area_name="", name=""):
    cur_user = User.objects.filter(del_flag=FLAG_NO, id=cur_user_id).first()
    if cur_user.account != user:
        return dict(c=ERR_USER_AUTH[0], m=ERR_USER_AUTH[1])

    if user_id_list:
        del_user_list = map(lambda x: int(x), json.loads(user_id_list))
    elif del_all_user:
        del_userdata_list = list_user(user, cur_user_id, area_name, name)['d']
        del_user_list = list()
        for each_userdata in del_userdata_list:
            del_user_list.append(each_userdata['id'])
    else:
        # user_id_list 和del_all_user不能同时为空
        raise BusinessException(REQUEST_PARAM_ERROR)

    # del_id_enable = User.objects.filter(del_flag=FLAG_NO, area__parent=cur_user.area, id__in=del_user_list)
    # del_id_enable = del_id_enable.values_list("id", flat=True)
    # del_id_enable = list(del_id_enable)
    # del_id_disable = list(set(del_user_list) - set(del_id_enable))

    # # 检查用户是否在活动中。如果用户已经被加入到活动中则不允许删除！
    # user_roles = Role.objects.filter(user_id__in=del_id_enable, del_flag=FALSE_INT).values_list("user__account__username", flat=True)
    # if user_roles:
    #     return dict(c=ERR_CANT_DEL_USER_INACT[0], m=ERR_CANT_DEL_USER_INACT[1] + ','.join(set(user_roles)))
    #
    # User.objects.filter(del_flag=FLAG_NO, id__in=del_id_enable).update(del_flag=FLAG_YES)
    # if del_id_disable:
    #     if del_id_enable:
    #         msg = u"user_id_list {0} deleted. while user_id_list {1} can not delete ".\
    #             format(del_id_enable, del_id_disable)
    #         return dict(c=ERR_PARTLY_DELETE_SUCCESS[0], m=ERR_PARTLY_DELETE_SUCCESS[1], d=[msg])
    #     else:
    #         return dict(c=ERR_USER_NOT_EXIST[0], m=ERR_USER_NOT_EXIST[1])
    # else:
    #     return dict(c=SUCCESS[0], m=SUCCESS[1], d=[])

    # v3版本只是在用户库是改变显示状态，不做逻辑删除
    result = User.objects.filter(del_flag=FLAG_NO, area__parent=cur_user.area, id__in=del_user_list, parent_account=user) \
        .update(is_show_store=0)
    return dict(c=SUCCESS[0], m=SUCCESS[1], d=result)


def list_sub_area(user, cur_user_id, area_id, manage_direct, area_name, is_school):
    if cur_user_id:
        # 關聯用户时，由用户发起的地域条件过滤
        cur_user = User.objects.filter(del_flag=FLAG_NO, id=cur_user_id).first()
        if cur_user.account != user:
            return dict(c=ERR_USER_AUTH[0], m=ERR_USER_AUTH[1])

        if not area_id:
            area = cur_user.area
        else:
            area = Area.objects.filter(del_flag=FLAG_NO, id=area_id).first()
            if not area:
                return dict(c=ERR_REQUEST_PARAMETER_ERROR[0], m=ERR_REQUEST_PARAMETER_ERROR[1])
    else:
        # 直接由地域查询的子地域请求
        if not area_id:
            nation_area = Area.objects.filter(parent_id__isnull=True, del_flag=FALSE_INT).first()
            area = nation_area
        else:
            area = Area.objects.filter(del_flag=FLAG_NO, id=area_id).first()

    if not area:
        return dict(c=ERR_REQUEST_PARAMETER_ERROR[0], m=ERR_REQUEST_PARAMETER_ERROR[1])

    if area.manage_direct:
        return dict(c=SUCCESS[0], m=SUCCESS[1], d=[])
    sub_areas = Area.objects.filter(del_flag=FLAG_NO, parent=area)

    if manage_direct:
        sub_areas = sub_areas.filter(manage_direct=manage_direct)
    else:
        sub_areas = sub_areas.filter(manage_direct=FLAG_NO)
    if area_name:
        sub_areas = sub_areas.filter(area_name__contains=area_name)
    if is_school:
        sub_areas = sub_areas.filter(region__isnull=False if is_school == '1' else True)

    sub_areas = sub_areas.values("id", "area_code", "area_level", "area_name", "parent__area_name", "manage_direct", "parent_id")
    sub_areas = list(sub_areas)
    data = []
    for item in sub_areas:
        tmp_dict = {
            "id": item['id'],
            "code": item["area_code"],
            "level": item["area_level"],
            "name": item["area_name"],
            "parent_name": item['parent__area_name'],
            "manage_direct": item['manage_direct'],
            "parent_id": item['parent_id'],
        }
        data.append(tmp_dict)

    # 查询全部时，直属全部归为一条记录。
    if not manage_direct:
        if area.area_level == LEVEL_PROVINCE:
            data.append({"id": "", "code": "", "level": LEVEL_CITY, "name": u"省直属", "parent_name": area.area_name,
                         "manage_direct": 1, "parent_id": area.area_name})
        elif area.area_level == LEVEL_CITY:
            data.append({"id": "", "code": "", "level": LEVEL_COUNTY, "name": u"市直属", "parent_name": area.area_name,
                         "manage_direct": 1})
        elif area.area_level == LEVEL_NATION:
            data.append({"id": "", "code": "", "level": LEVEL_NATION, "name": u"国直属", "parent_name": area.area_name,
                         "manage_direct": 1})
    return dict(c=SUCCESS[0], m=SUCCESS[1], d=data)


@transaction.atomic
def config_user(user, account_id, activity_mask, area_id):
    if not user.is_admin:
        return dict(c=ERR_USER_AUTH[0], m=ERR_USER_AUTH[1])
    account = Account.objects.filter(del_flag=FLAG_NO, id=account_id).first()
    if not account:
        return dict(c=ERR_REQUEST_PARAMETER_ERROR[0], m=ERR_REQUEST_PARAMETER_ERROR[1])
    activity_mask = int(activity_mask)
    area = Area.objects.filter(del_flag=FLAG_NO, id=area_id).first()
    if not area:
        return dict(c=ERR_REQUEST_PARAMETER_ERROR[0], m=ERR_REQUEST_PARAMETER_ERROR[1])
    user_obj_query = User.objects.filter(del_flag=FLAG_NO, account=account, area__area_level=area.area_level)
    if not user_obj_query:
        mask = account.activity_mask | activity_mask
        if not mask & area.area_level:
            return dict(c=ERR_REQUEST_PARAMETER_ERROR[0], m=ERR_REQUEST_PARAMETER_ERROR[1])
        else:
            User.objects.create(account=account, area=area, name=account.name, sex=account.sex)
    else:
        raise Exception(u"同等级用户已经创建。")
    account.activity_mask |= activity_mask
    account.save()
    return dict(c=SUCCESS[0], m=SUCCESS[1], d=[])


@transaction.atomic
def add_account(user, username, name, sex, activity_mask, area_id, password=settings.DEFAULT_PASSWORD):
    if not user.is_admin:
        return dict(c=ERR_USER_AUTH[0], m=ERR_USER_AUTH[1])
    account = Account.objects.filter(del_flag=FLAG_NO, username=username).first()
    if not account:
        account = Account.objects.create_user(username, password=password)
        account.name = name if name else ""
        if sex:  account.sex = sex
        if re.match('[gl]', username, re.I):
            account.code = username
        elif len(username) == 11 and username.isdigit():
            account.mobile = username
        if activity_mask:
            activity_mask = int(activity_mask)
            account.activity_mask = activity_mask
        account.save()
        if area_id:
            resp = config_user(user, account_id=account.id, activity_mask=activity_mask, area_id=area_id)
            if resp['c'] != SUCCESS[0]:
                raise Exception(u"{}".format(resp["m"]))
        return dict(c=SUCCESS[0], m=SUCCESS[1], d=[account.id])
    else:
        return dict(c=ERR_USER_EXISTS[0], m=ERR_USER_EXISTS[1])


def get_current_user(user, area_id):
    if user.is_admin:
        root_user = User.objects.filter(del_flag=FLAG_NO, account=user).first()
        user_info = {"id": root_user.id, "username": user.username, "account_name": user.name,
                     "user_name": user.name,
                     "account_id": user.id,
                     "activity_mask": user.activity_mask,
                     "image_url": get_image_url(user.image.url) if user.image else "",
                     "is_activity_admin": 0,
                     "auth": 0}
        return dict(c=SUCCESS[0], m=SUCCESS[1], d=[user_info])

    else:
        if area_id:
            cur_area = Area.objects.filter(del_flag=FLAG_NO, id=area_id, manage_direct=FLAG_NO).first()
        else:
            cur_area = Area.objects.filter(del_flag=FLAG_NO, parent=None, manage_direct=FLAG_NO).first()
        if not cur_area:
            raise Exception(u"地域的id-{}不存在".format(area_id))
        user_of_area = User.objects.filter(del_flag=FLAG_NO, account=user, area=cur_area).first()
        if not user_of_area:
            raise Exception(u"找不到相关的用户")

        user_info = {"id": user_of_area.id, "username": user_of_area.account.username,
                     "account_name": user_of_area.account.name, "user_name": user_of_area.name,
                     "account_id": user_of_area.account_id,
                     "activity_mask": user.activity_mask,
                     "image_url": get_image_url(user.image.url) if user.image else "",
                     "is_activity_admin": 1 if cur_area.area_level & user.activity_mask else 0,
                     "auth": 1 if cur_area.area_level & user.activity_mask else 0}
        return dict(c=SUCCESS[0], m=SUCCESS[1], d=[user_info])


def api_get_activity_user(user, activity_id):
    result = {
        "activity_user_type": "",
        "cur_user_id": "",
        "expert_id": "",
        "is_leader": "",
        "user_area_manage_direct": "",
        "user_flag": 0,
        "expert_flag": 0,
    }

    activity_user_type = USER_TYPE_NONE

    try:
        activity_role = get_account_activity_role(activity_id, user.id)
    except BusinessException as ex:
        # 如果异常说明用户没有权限
        result['activity_user_type'] = USER_TYPE_NONE
        return result

    for usertype in (USER_TYPE_ACTIVITY_ADMIN, USER_TYPE_NATION_ADMIN, USER_TYPE_PROVINCE_ADMIN, USER_TYPE_CITY_ADMIN, USER_TYPE_COUNTRY_ADMIN,
                     USER_TYPE_INSTITUTION_ADMIN, USER_TYPE_NORMAL, USER_TYPE_EXPERT):
        # 用户只能有一个权限，权限从大到小排，只要有其中一个，就退出判断。
        if usertype in activity_role:
            activity_user_type = usertype
            break

    # 查询当前用户id
    role = Role.objects.filter(activity_id=activity_id, user__account=user, del_flag=FALSE_INT).first()
    if role:
        cur_user_id = role.user_id
    else:
        cur_user_id = ''

    # 查询当前用户是不是专家
    expertactivity = ExpertActivity.objects.filter(activity_id=activity_id, expert__account=user, expert__del_flag=FALSE_INT)
    if expertactivity:
        expert_id = expertactivity.first().expert_id
        teamexpert = TeamExpert.objects.filter(team__activity_id=activity_id, expert_id=expert_id, expert__del_flag=FALSE_INT)
        if not teamexpert:
            # 还没分组，所有专家都不是组长。
            is_leader = 0
        else:
            # 如果已经分组，取分组内的组长标识
            is_leader = teamexpert.first().is_leader
    else:
        expert_id = ''
        is_leader = ''
    result['activity_user_type'] = activity_user_type
    result['cur_user_id'] = cur_user_id
    result['expert_id'] = expert_id
    result['is_leader'] = is_leader
    result['user_area_manage_direct'] = role.user.area.manage_direct if role else expertactivity.first().expert.area.manage_direct
    result['user_flag'] = 1 if role else 0
    result['expert_flag'] = 1 if expertactivity else 0
    return result


def get_area_onhome():
    area_list = Activity.objects.filter(del_flag=FLAG_NO, stage__gte=1, template_flag=FLAG_NO,
                                        user__area__manage_direct=FLAG_NO). \
        values("user__area_id", "user__area__area_code", "user__area__area_name", "user__area__area_level",
               "user__area__parent").order_by("-user__area__area_level")
    data = []
    id_statistic = []
    for item in list(area_list):
        area_id = item["user__area_id"]
        if area_id in id_statistic:
            continue
        else:
            id_statistic.append(area_id)
            area_level = item["user__area__area_level"]
            if area_level >= LEVEL_COUNTY:
                parent = item['user__area__parent']
                if parent:
                    parent_area = Area.objects.get(id=parent)
                    parent_area_name = parent_area.area_name
                else:
                    parent_area_name = ""
                dict_tmp = {"area_id": item["user__area_id"],
                            "area_code": item["user__area__area_code"],
                            "area_name": item["user__area__area_name"] if area_level > LEVEL_COUNTY else parent_area_name + item[
                                "user__area__area_name"],
                            "area_parent": parent_area_name,
                            "area_level": area_level,
                            }
                data.append(dict_tmp)
    return dict(c=SUCCESS[0], m=SUCCESS[1], d=data)


def get_area_byid(area_id):
    # 根据id查询区域信息
    result = dict()
    if area_id:
        area_qry = Area.objects.filter(id=area_id, del_flag=FALSE_INT).first()
    else:
        area_qry = Area.objects.filter(parent_id__isnull=True, del_flag=FALSE_INT).first()

    if not area_qry:
        raise BusinessException(ERR_AREA_ERROR)

    result['area_id'] = area_qry.id
    result['area_code'] = area_qry.area_code
    result['area_level'] = area_qry.area_level
    result['area_name'] = area_qry.area_name
    result['manage_direct'] = area_qry.manage_direct
    result['parent_id'] = area_qry.parent_id if area_qry.parent_id else ''

    return dict(c=SUCCESS[0], m=SUCCESS[1], d=result)

#
# def import_user(user, file_obj="", user_flag="", area_id="", cur_user_id=""):
#     if not file_obj.name.endswith('xlsx') and not file_obj.name.endswith('xls'):
#         raise BusinessException(ERR_FILE_FORMAT_NOT_SUPPORTED)
#
#     user_flag = int(user_flag)
#     ret_data = []
#
#     # #得到关于导入者的信息
#     # importer_id = user.id
#     # importer_dict = get_account_info(activity_id=activity_id,account_id=importer_id)
#     # if importer_dict["c"] != ERR_SUCCESS[0]:
#     #     return {"c": ERR_USER_AUTH[0], "m": ERR_USER_AUTH[1], "d": []}
#     # importer_info = importer_dict["d"][0]
#     # importer_type = importer_info["account_type"]
#     cur_area = Area.objects.filter(id=area_id, del_flag=FALSE_INT).first()
#     if not cur_area:
#         raise BusinessException(ERR_AREA_ERROR)
#
#     # write file data to tmp path
#     file_path = gen_path()
#     with open(file_path, 'wb+') as destination:
#         for chunk in file_obj.chunks():
#             destination.write(chunk)
#
#     wb = openpyxl.load_workbook(filename=file_path, read_only=True)
#     ws = wb.get_active_sheet()
#     if cur_area.area_level == LEVEL_PROVINCE:
#         property_display_list = [u"用户名", u"姓名", u"性别", u"市州", u"机构"]
#         property_name_list = ["username", "name", "sex", "city", "institution"]
#     else:
#         property_display_list = [u"用户名", u"姓名", u"性别", u"市州", u"区县", u"机构"]
#         property_name_list = ["username", "name", "sex", "city", "county", "institution"]
#     property_len = len(property_name_list)
#
#     # check the first row
#     for i in range(property_len):
#         try:
#             if not re.search(property_display_list[i], ws.cell(column=i + 1, row=1).value):
#                 raise BusinessException(ERR_FILE_TEMPLATE_ERROR)
#         except:
#             raise BusinessException(ERR_FILE_TEMPLATE_ERROR)
#
#     # 检查表格数据是否有误，若有误，则返回错误数据提示信息
#     row_num = 0
#     for row in ws.rows:
#         row_num += 1
#         if row_num == 1:
#             # 第一行为标题栏
#             continue
#         col_num = 0
#         account_info = {}
#         for cell in row:
#             if col_num >= property_len:
#                 break
#             property_name = property_name_list[col_num]
#             value = ""
#             if isinstance(cell.value, (int, long)):
#                 value = str(cell.value)
#             elif isinstance(cell.value, basestring):
#                 value = cell.value.strip()
#             account_info[property_name] = value
#             col_num += 1
#         # 忽略空白行
#         if not account_info["username"] and not account_info["name"]:
#             continue
#         # check not null field
#         if not account_info["username"]:
#             err_msg = u"第%d行: 用户名为空" % row_num
#             ret_data.append(err_msg)
#             continue
#
#         # 用户名格式的检查
#         resp_code = __check_username_valid(account_info["username"])
#         if resp_code['c'] != ERR_SUCCESS[0]:
#             err_msg = u"第%d行: 用户名格式不正确" % row_num
#             ret_data.append(err_msg)
#
#         # 检查表格中市州区县是否存在
#         province = Area.objects.get(area_level=LEVEL_PROVINCE, del_flag=FALSE_INT)
#         city_list = get_child_by_areaid(province.id, manage_direct='0')
#
#         city_name_list = []
#         for each_city in city_list:
#             city_name_list.append(each_city['area_name'])
#
#         if account_info["city"] not in city_name_list:
#             err_msg = u"第%d行: 市州信息有误" % row_num
#             ret_data.append(err_msg)
#             continue
#
#         county_name_list = get_rank_child_area_name(account_info["city"], LEVEL_CITY)
#         if account_info.has_key("county") and account_info["county"] not in county_name_list:
#             err_msg = u"第%d行: 区县信息有误" % row_num
#             ret_data.append(err_msg)
#             continue
#
#         # 如果为省直属或者市直属，机构字段不允许为空
#         if account_info["city"] == u'省直属' and not account_info["institution"]:
#             err_msg = u"第%d行: 省直属必须填写直属的机构名称" % row_num
#             ret_data.append(err_msg)
#             continue
#         if account_info.has_key("county") and account_info["county"] == u'市直属' and not account_info["institution"]:
#             err_msg = u"第%d行: 市直属必须填写直属的机构名称" % row_num
#             ret_data.append(err_msg)
#             continue
#
#         # 专家导入必须填写机构
#         if user_flag == 1 and not account_info["institution"]:
#             err_msg = u"第%d行: 导入专家必须填写专家归属机构名称" % row_num
#             ret_data.append(err_msg)
#             continue
#
#         # #如果导入者是超级管理员
#         # if importer_type == USER_TYPE_SUPER_ADMIN:
#         #     if user_flag != 2:
#         #         #如果不是导入的市州或直属机构的管理员
#         #         if not (((account_info["city"] in CITY_LIST) and account_info["city"] != u"省直属" and account_info["institution"]=="") or
#         #                     (account_info["city"]==u"省直属" and account_info["institution"] != "")):
#         #             err_msg = u"第%d行: 地区信息有误" % row_num
#         #             ret_data.append(err_msg)
#         #             continue
#         #     else:
#         #         if not ((account_info["city"] in CITY_LIST) and account_info["institution"] != ""):
#         #             err_msg = u"第%d行: 地区信息有误" % row_num
#         #             ret_data.append(err_msg)
#         #             continue
#         #
#         # #如果导入者是市州管理员
#         # elif importer_type == USER_TYPE_CITY_ADMIN:
#         #     #如果不是创建的区县或直属机构的管理员
#         #     if not ((account_info["city"] == importer_info["city"] and (account_info["country"] in CITY_COUNTRY_MAPPING[importer_info["city"]]) and account_info["country"]!=u"市直属" and account_info["institution"]=="")
#         #                or (account_info["city"]==importer_info["city"] and account_info["country"]==u"市直属" and account_info["institution"] != "")):
#         #         err_msg = u"第%d行: 地区信息有误" % row_num
#         #         ret_data.append(err_msg)
#         #         continue
#         #
#         # #如果创建者是区县管理员
#         # elif importer_type == USER_TYPE_COUNTRY_ADMIN:
#         #     #如果不是创建机构的管理员
#         #     if not (account_info["city"] == importer_info["city"] and account_info["country"]==importer_info["country"]
#         #                and account_info["institution"] != ""):
#         #         err_msg = u"第%d行: 地区信息有误" % row_num
#         #         ret_data.append(err_msg)
#         #         continue
#         #
#         # #如果创建者是机构管理员
#         # elif importer_type == USER_TYPE_INSTITUTION_ADMIN:
#         #     if not (account_info["city"] == importer_info["city"] and account_info["country"] == importer_info["country"]
#         #                and account_info["institution"] == importer_info["institution"]):
#         #         err_msg = u"第%d行: 地区信息有误" % row_num
#         #         ret_data.append(err_msg)
#         #         continue
#
#     if ret_data:
#         return {"c": -1, "m": u"表格数据有误，数据没有导入，请修改后重新导入", "d": ret_data}
#
#     # 如果表格数据无误，则导入数据
#     row_num = 0
#     for row in ws.rows:
#         row_num += 1
#         if row_num == 1:
#             # 第一行为标题栏
#             continue
#         col_num = 0
#         account_info = {}
#         for cell in row:
#             if col_num >= property_len:
#                 break
#             property_name = property_name_list[col_num]
#             value = ""
#             if isinstance(cell.value, (int, long)):
#                 value = str(cell.value)
#             elif isinstance(cell.value, basestring):
#                 value = cell.value.strip()
#             account_info[property_name] = value
#             col_num += 1
#         # 忽略空白行
#         if not account_info["username"] and not account_info["name"]:
#             continue
#
#         if account_info.get("county") is None:
#             account_info["county"] = ""
#
#         if account_info["city"] == u'省直属' or account_info["county"] == u'市直属':
#             area_name = account_info["institution"]
#             manage_direct = "1"
#         elif account_info["county"]:
#             area_name = account_info["county"]
#             manage_direct = None
#         else:
#             area_name = account_info["city"]
#             manage_direct = None
#
#         try:
#             if user_flag == 0:
#                 err_code = add_user(user, account_info["username"], account_info["name"], account_info["sex"], area_name, cur_user_id, manage_direct)
#             else:
#                 err_code = add_expert_user(user, cur_user_id, account_info["username"], account_info["name"], account_info["sex"], area_name,
#                                            account_info["institution"], None, None, None, manage_direct)
#
#             if err_code['c'] != ERR_SUCCESS[0]:
#                 err_msg = u"第%d行[%s]: " % (row_num, account_info["username"])
#                 err_msg += err_code['m']
#                 ret_data.append(err_msg)
#
#         except Exception as e:
#             err_msg = u"第%d行[%s]: " % (row_num, account_info["username"])
#             err_msg += str(e)
#             ret_data.append(err_msg)
#         # err_code = add_user(user=user, activity_id=activity_id, user_flag=user_flag,name=account_info["name"],
#         #                     sex=account_info["sex"], username=account_info["username"],
#         #                     city=account_info["city"], country=account_info["country"], institution=account_info["institution"])
#
#     if ret_data:
#         return {"c": ERR_IMPORT_DATA_ERROR[0], "m": ERR_IMPORT_DATA_ERROR[1], "d": ret_data}
#     else:
#         return {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": []}
#
#
# def import_activity_user(user, file_obj="", activity_id="", user_flag="", cur_user_id=""):
#     if not file_obj.name.endswith('xlsx') and not file_obj.name.endswith('xls'):
#         raise BusinessException(ERR_FILE_FORMAT_NOT_SUPPORTED)
#
#     activity_id = int(activity_id)
#     user_flag = int(user_flag)
#     ret_data = []
#
#     # 得到关于导入者的信息
#     # importer_id = user.id
#     # importer_dict = get_account_info(activity_id=activity_id,account_id=importer_id)
#     # if importer_dict["c"] != ERR_SUCCESS[0]:
#     #     return {"c": ERR_USER_AUTH[0], "m": ERR_USER_AUTH[1], "d": []}
#     # importer_info = importer_dict["d"][0]
#     # importer_type = importer_info["account_type"]
#
#     # write file data to tmp path
#     file_path = gen_path()
#     with open(file_path, 'wb+') as destination:
#         for chunk in file_obj.chunks():
#             destination.write(chunk)
#
#     wb = openpyxl.load_workbook(filename=file_path, read_only=True)
#     ws = wb.get_active_sheet()
#     property_display_list = [u"用户名"]
#     property_name_list = ["username"]
#     property_len = len(property_name_list)
#
#     # check the first row
#     for i in range(property_len):
#         try:
#             if not re.search(property_display_list[i], ws.cell(column=i + 1, row=1).value):
#                 raise BusinessException(ERR_FILE_TEMPLATE_ERROR)
#         except:
#             raise BusinessException(ERR_FILE_TEMPLATE_ERROR)
#
#     # 检查表格数据是否有误，若有误，则返回错误数据提示信息
#     row_num = 0
#     for row in ws.rows:
#         row_num += 1
#         if row_num == 1:
#             # 第一行为标题栏
#             continue
#         col_num = 0
#         account_info = {}
#         for cell in row:
#             if col_num >= property_len:
#                 break
#             property_name = property_name_list[col_num]
#             value = ""
#             if isinstance(cell.value, (int, long)):
#                 value = str(cell.value)
#             elif isinstance(cell.value, basestring):
#                 value = cell.value.strip()
#             account_info[property_name] = value
#             col_num += 1
#         # 忽略空白行
#         if not account_info["username"] and not account_info["name"]:
#             continue
#         # check not null field
#         if not account_info["username"]:
#             err_msg = u"第%d行: 用户名为空" % row_num
#             ret_data.append(err_msg)
#             continue
#         # 用户名格式的检查
#         resp_code = __check_username_valid(account_info["username"])
#         if resp_code['c'] != ERR_SUCCESS[0]:
#             err_msg = u"第%d行: 用户名格式不正确" % row_num
#             ret_data.append(err_msg)
#
#     if ret_data:
#         return {"c": -1, "m": u"表格数据有误，数据没有导入，请修改后重新导入", "d": ret_data}
#
#     # 获取可添加用户列表
#     add_id_list = []
#     if user_flag == 0:
#         # 先获取所有可添加的用户列表
#         rolecanadd = list_activity_role(user, activity_id, None, None, cur_user_id, True, None)
#         if rolecanadd['c'] == ERR_SUCCESS[0]:
#             usercanadd = rolecanadd['d']
#         else:
#             raise BusinessException(ERR_CHECK_CANADDUSER)
#     else:
#         # 获取可添加专家列表
#         expertcanadd = list_activity_expert(user, activity_id, None, None, cur_user_id, True, None, None)
#         if expertcanadd['c'] == ERR_SUCCESS[0]:
#             usercanadd = expertcanadd['d']
#         else:
#             raise BusinessException(ERR_CHECK_CANADDEXPERT)
#
#     # ero_code = add_activity_role(user, activity_id, user_id_list, cur_user_id)
#     # err_code = add_user(user=user, activity_id=activity_id, user_flag=user_flag, name=account_info["name"],
#     #                     sex=account_info["sex"], username=account_info["username"],
#     #                     city=account_info["city"], country=account_info["country"], institution=account_info["institution"])
#
#     # 如果表格数据无误，则导入课程数据
#     row_num = 0
#     for row in ws.rows:
#         row_num += 1
#         if row_num == 1:
#             # 第一行为标题栏
#             continue
#         col_num = 0
#         account_info = {}
#         for cell in row:
#             if col_num >= property_len:
#                 break
#             property_name = property_name_list[col_num]
#             value = ""
#             if isinstance(cell.value, (int, long)):
#                 value = str(cell.value)
#             elif isinstance(cell.value, basestring):
#                 value = cell.value.strip()
#             account_info[property_name] = value
#             col_num += 1
#         # 忽略空白行
#         if not account_info["username"] and not account_info["name"]:
#             continue
#
#         importdata_in_rolecanadd = False
#         for each_data in usercanadd:
#             if account_info["username"] == each_data['username']:
#                 if user_flag == 0:
#                     add_id_list.append(each_data['user_id'])
#                 else:
#                     add_id_list.append(each_data['expert_id'])
#                 importdata_in_rolecanadd = True
#                 break
#
#         if not importdata_in_rolecanadd:
#             err_msg = u"第%d行[%s]: " % (row_num, account_info["username"])
#             err_msg += u'用户不在可添加用户列表中，不允许添加！'
#             ret_data.append(err_msg)
#
#     if not add_id_list:
#         return {"c": ERR_NODATA_IMPORT[0], "m": ERR_NODATA_IMPORT[1], "d": ret_data}
#
#     if user_flag == 0:
#         err_code = add_activity_role(user, activity_id, json.dumps(add_id_list), cur_user_id)
#     else:
#         err_code = add_activity_expert(user, activity_id, json.dumps(add_id_list), cur_user_id)
#
#     if err_code['c'] != ERR_SUCCESS[0]:
#         err_msg = u"添加成员失败！"
#         ret_data.append(err_msg)
#
#     if ret_data:
#         return {"c": ERR_IMPORT_DATA_ERROR[0], "m": ERR_IMPORT_DATA_ERROR[1], "d": ret_data}
#     else:
#         return {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": []}
#
#
# def api_export_user(user, user_id, area_name, name, item_id_list):
#     # 用户时
#     if item_id_list:
#         item_id_json = json.loads(item_id_list)
#     else:
#         item_id_json = None
#
#     user_list = list_user(user, user_id, area_name, name)
#     if user_list['c'] != ERR_SUCCESS[0]:
#         return user_list
#
#     user_list = user_list['d']
#
#     result_row_head = [u'用户名', u'姓名', u'性别', u'地区']
#     result_data_list = []
#     result_data_list.append(result_row_head)
#
#     for each_user in user_list:
#         data_row = [each_user['username'], each_user['name'], each_user['sex'], each_user['area_name']]
#         if item_id_json:
#             if each_user['id'] in item_id_json or str(each_user['id']) in item_id_json:
#                 result_data_list.append(data_row)
#         else:
#             result_data_list.append(data_row)
#
#     file_path = gen_path(suffix='.xls')
#     save_array_excel(result_data_list, 'Sheet', file_path, hasrowhead=True, hascolhead=False)
#     return dict(c=SUCCESS[0], m=SUCCESS[1], d=file_path)
#
#
# def api_export_expert(user, cur_user_id, name, area_id, manage_direct, item_id_list):
#     # 专家时
#     if item_id_list:
#         item_id_json = json.loads(item_id_list)
#     else:
#         item_id_json = None
#
#     expert_list = list_expert_user(user, cur_user_id, name, area_id, manage_direct)
#     if expert_list['c'] != ERR_SUCCESS[0]:
#         return expert_list
#
#     expert_list = expert_list['d']
#
#     result_row_head = [u'用户名', u'姓名', u'性别', u'地区', u'机构']
#     result_data_list = [result_row_head]
#
#     for each_expert in expert_list:
#         data_row = [each_expert['username'], each_expert['name'], each_expert['sex'], each_expert['area_name'], each_expert['institution'], ]
#         if item_id_json:
#             if each_expert['id'] in item_id_json or str(each_expert['id']) in item_id_json:
#                 result_data_list.append(data_row)
#         else:
#             result_data_list.append(data_row)
#
#     file_path = gen_path(suffix='.xls')
#     save_array_excel(result_data_list, 'Sheet', file_path, hasrowhead=True, hascolhead=False)
#     return dict(c=SUCCESS[0], m=SUCCESS[1], d=file_path)
#
#
# def export_activity_user(user, activity_id, account_id_list, name, user_flag, area_id, cur_user_id, direct_level, item_id_list, qry_all_user):
#     # 用户时
#     if item_id_list:
#         item_id_json = json.loads(item_id_list)
#     else:
#         item_id_json = None
#
#     activity_role_list = list_activity_role(user, activity_id, area_id, name, cur_user_id, None, direct_level, qry_all_user)
#     if activity_role_list['c'] != ERR_SUCCESS[0]:
#         return activity_role_list
#
#     activity_role_list = activity_role_list['d']
#
#     result_row_head = [u'用户名', u'姓名', u'性别', u'地区', u'作品提交数量', u'作品未审批量']
#     result_data_list = [result_row_head]
#
#     for each_role in activity_role_list:
#         if not account_id_list or each_role['account_id'] in account_id_list:
#             data_row = [each_role['username'], each_role['name'], each_role['sex'], each_role['area_name'], each_role['user_approve_number'],
#                         each_role['user_noapprove_number'], ]
#             if item_id_json:
#                 if each_role['id'] in item_id_json or str(each_role['id']) in item_id_json:
#                     result_data_list.append(data_row)
#             else:
#                 result_data_list.append(data_row)
#
#     file_path = gen_path(suffix='.xls')
#     save_array_excel(result_data_list, 'Sheet', file_path, hasrowhead=True, hascolhead=False)
#     return dict(c=SUCCESS[0], m=SUCCESS[1], d=file_path)


# def export_activity_expert(user, activity_id, area_id, name, cur_user_id, only_can_add, direct_level, institution, item_id_list):
#     # 专家时
#     if item_id_list:
#         item_id_json = json.loads(item_id_list)
#     else:
#         item_id_json = None
#
#     activity_expert_list = list_activity_expert(user, activity_id, area_id, name, cur_user_id, only_can_add, direct_level, institution)
#     if activity_expert_list['c'] != SUCCESS[0]:
#         return activity_expert_list
#
#     activity_expert_list = activity_expert_list['d']
#
#     result_row_head = [u'用户名', u'姓名', u'性别', u'地区', u'机构']
#     result_data_list = [result_row_head]
#
#     for each_expert in activity_expert_list:
#         data_row = [each_expert['username'], each_expert['name'], each_expert['sex'], each_expert['area_name'], each_expert['institution'], ]
#         # 如果限定导出指定的id，则id在指定列表中才导出。如果未限制id，则所有列表都导出。
#         if item_id_json:
#             if each_expert['id'] in item_id_json or str(each_expert['id']) in item_id_json:
#                 result_data_list.append(data_row)
#         else:
#             result_data_list.append(data_row)
#
#     file_path = gen_path(suffix='.xls')
#     save_array_excel(result_data_list, 'Sheet', file_path, hasrowhead=True, hascolhead=False)
#     return dict(c=SUCCESS[0], m=SUCCESS[1], d=file_path)
#
#
# def download_user_template(user, area_id="", user_flag="0"):
#     cur_area = Area.objects.filter(id=area_id, del_flag=FALSE_INT).first()
#     if not cur_area:
#         raise BusinessException(ERR_AREA_ERROR)
#     downloader_area_detail = get_areadetail_by_id(area_id)
#
#     user_flag = int(user_flag)
#
#     if user_flag == 0:
#         if cur_area.area_level == LEVEL_PROVINCE:
#             USER_PROPERTY_LIST = [u'用户名*', u'姓名*', u'性别*', u'市州*', u'机构']
#         elif cur_area.area_level == LEVEL_CITY:
#             USER_PROPERTY_LIST = [u'用户名*', u'姓名*', u'性别*', u'市州*', u'区县*', u'机构']
#         else:
#             USER_PROPERTY_LIST = [u'用户名*', u'姓名*', u'性别*', u'市州*', u'区县*', u'机构*']
#     elif user_flag == 1:
#         if cur_area.area_level == LEVEL_PROVINCE:
#             USER_PROPERTY_LIST = [u'用户名*', u'姓名*', u'性别*', u'市州*', u'机构*']
#         else:
#             USER_PROPERTY_LIST = [u'用户名*', u'姓名*', u'性别*', u'市州*', u'区县*', u'机构*']
#     else:
#         raise BusinessException(ERR_USER_FLAG)
#
#     # 获取文件路径
#     file_path = gen_path()
#     wb = Workbook(encoding='utf-8')
#     ws = wb.active
#     title_font = Font(b=True)
#
#     # 设置市州-区县对应关系
#     if cur_area.area_level != LEVEL_PROVINCE:
#         __set_city_country_data_validation(wb, ws, USER_PROPERTY_LIST.index(u"市州*") + 1, USER_PROPERTY_LIST.index(u"区县*") + 1)
#     else:
#         __set_city_data_validation(wb, ws, USER_PROPERTY_LIST.index(u"市州*") + 1)
#
#     # 模板的首行设置
#     col = 1
#     row = 1
#     for property_name in USER_PROPERTY_LIST:
#         cell = ws.cell(column=col, row=row, value=property_name)
#         cell.font = title_font
#         col += 1
#
#     # 设置第二、三行的前三列固定数据
#     ws.cell(column=1, row=2, value="13966668888")
#     ws.cell(column=2, row=2, value=u"张三")
#     ws.cell(column=3, row=2, value=u"男")
#
#     ws.cell(column=1, row=3, value="15600339999")
#     ws.cell(column=2, row=3, value=u"李四")
#     ws.cell(column=3, row=3, value=u"女")
#
#     # 根据用户类型，设置二、三行的后几列数据
#     if user_flag == 0:
#         if cur_area.area_level == LEVEL_PROVINCE:
#             ws.cell(column=4, row=2, value=u"省直属")
#             ws.cell(column=5, row=2, value=u"省教育局")
#             ws.cell(column=4, row=3, value=u"武汉")
#             ws.cell(column=5, row=3, value=u"")
#
#         elif cur_area.area_level == LEVEL_CITY:
#             ws.cell(column=4, row=2, value=downloader_area_detail["city"])
#             ws.cell(column=5, row=2, value=u"市直属")
#             ws.cell(column=6, row=2, value=u"市教育局")
#             ws.cell(column=4, row=3, value=downloader_area_detail["city"])
#             ws.cell(column=5, row=3, value=get_rank_child_area_name(downloader_area_detail["city"], LEVEL_CITY)[0] if get_rank_child_area_name(
#                 downloader_area_detail["city"], LEVEL_CITY) else '')
#             ws.cell(column=6, row=3, value=u"")
#
#         elif cur_area.area_level == LEVEL_COUNTY:
#             ws.cell(column=4, row=2, value=downloader_area_detail["city"])
#             ws.cell(column=5, row=2, value=downloader_area_detail["country"])
#             ws.cell(column=6, row=2, value=downloader_area_detail["country"] + u"教育局")
#             ws.cell(column=4, row=3, value=downloader_area_detail["city"])
#             ws.cell(column=5, row=3, value=downloader_area_detail["country"])
#             ws.cell(column=6, row=3, value=downloader_area_detail["country"] + u"教科所")
#
#         elif cur_area.area_level == LEVEL_INSTITUTION:
#             ws.cell(column=4, row=2, value=downloader_area_detail["city"])
#             ws.cell(column=5, row=2, value=downloader_area_detail["country"])
#             ws.cell(column=6, row=2, value=downloader_area_detail["institution"])
#             ws.cell(column=4, row=3, value=downloader_area_detail["city"])
#             ws.cell(column=5, row=3, value=downloader_area_detail["country"])
#             ws.cell(column=6, row=3, value=downloader_area_detail["institution"])
#
#             ws.cell(column=1, row=3, value="G130879200308080030")
#
#         else:
#             raise BusinessException(ERR_USER_FLAG)
#
#     elif user_flag == 1:
#         if cur_area.area_level == LEVEL_PROVINCE:
#             ws.cell(column=4, row=2, value=u"省直属")
#             ws.cell(column=5, row=2, value=u"省教育局")
#             ws.cell(column=4, row=3, value=u"武汉")
#             ws.cell(column=5, row=3, value=u"武汉市教育局")
#         else:
#             ws.cell(column=4, row=2, value=downloader_area_detail["city"])
#             ws.cell(column=5, row=2, value=downloader_area_detail["country"])
#             ws.cell(column=6, row=2, value=downloader_area_detail["country"] + u"教育局")
#             ws.cell(column=4, row=3, value=downloader_area_detail["city"])
#             ws.cell(column=5, row=3, value=downloader_area_detail["country"])
#             ws.cell(column=6, row=3, value=downloader_area_detail["country"] + u"教科所")
#
#
#     else:
#         return {"c": ERR_USER_FLAG[0], "m": ERR_USER_FLAG[1], "d": []}
#
#     wb.save(file_path)
#     return {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": [file_path]}
#
#
# def download_activity_user_template():
#     # 新版的活动中的用户中心只能选择用户，并不能新增用户，所以没必要再填写其它内容了。填个用户名就可以了。
#
#     # 获取文件路径
#     file_path = gen_path()
#     wb = Workbook(encoding='utf-8')
#     ws = wb.active
#     title_font = Font(b=True)
#     USER_PROPERTY_LIST = ['用户名']
#
#     # 模板的首行设置
#     col = 1
#     row = 1
#     for property_name in USER_PROPERTY_LIST:
#         cell = ws.cell(column=col, row=row, value=property_name)
#         cell.font = title_font
#         col += 1
#
#     # 设置第二、三行的前三列固定数据
#     ws.cell(column=1, row=2, value="13888888888")
#     ws.cell(column=1, row=3, value="13999999999")
#
#     wb.save(file_path)
#     return {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": [file_path]}
#
#     pass
#
#
# def download_activity_user_template_old(user, activity_id="", user_flag="0"):
#     activity_id = int(activity_id)
#     activity_obj = Activity.objects.filter(id=activity_id, del_flag=0).first()
#     if not activity_obj:
#         raise BusinessException(ERR_ACTIVITY_ID_ERROR)
#
#     user_flag = int(user_flag)
#
#     # 得到下载者的信息
#     downloader_id = user.id
#     downloader_role = get_account_activity_role(activity_id, downloader_id)
#
#     # 获取下载者的地区信息
#     activity_user = api_get_activity_user(user, activity_id)
#     if activity_user['cur_user_id']:
#         cur_user = User.objects.get(id=activity_user['cur_user_id'])
#         area_id = cur_user.area_id
#     else:
#         cur_expert = Expert.objects.get(id=activity_user['expert_id'])
#         area_id = cur_expert.area_id
#
#     downloader_area_detail = get_areadetail_by_id(area_id)
#     # if downloader_dict["c"] != ERR_SUCCESS[0]:
#     #     return {"c": ERR_NODATA_FOUND[0], "m": ERR_NODATA_FOUND[1], "d": []}
#     # downloader_info = downloader_dict["d"][0]
#     # downloader_type = downloader_info["account_type"]
#
#     if user_flag == 0:
#         if USER_TYPE_PROVINCE_ADMIN in downloader_role:
#             USER_PROPERTY_LIST = [u'用户名*', u'姓名*', u'性别*', u'市州*', u'机构']
#         elif USER_TYPE_CITY_ADMIN in downloader_role:
#             USER_PROPERTY_LIST = [u'用户名*', u'姓名*', u'性别*', u'市州*', u'区县*', u'机构']
#         else:
#             USER_PROPERTY_LIST = [u'用户名*', u'姓名*', u'性别*', u'市州*', u'区县*', u'机构*']
#     elif user_flag == 1:
#         # todo 专家需要分层级返回不同的列名
#         USER_PROPERTY_LIST = [u'用户名*', u'姓名*', u'性别*', u'市州*', u'机构*']
#     else:
#         raise BusinessException(ERR_USER_FLAG)
#
#     # 获取文件路径
#     file_path = gen_path()
#     wb = Workbook(encoding='utf-8')
#     ws = wb.active
#     title_font = Font(b=True)
#
#     # 设置市州-区县对应关系
#     if user_flag == 0 and USER_TYPE_PROVINCE_ADMIN not in downloader_role:
#         __set_city_country_data_validation(wb, ws, USER_PROPERTY_LIST.index(u"市州*") + 1, USER_PROPERTY_LIST.index(u"区县*") + 1)
#     else:
#         __set_city_data_validation(wb, ws, USER_PROPERTY_LIST.index(u"市州*") + 1)
#
#     # 模板的首行设置
#     col = 1
#     row = 1
#     for property_name in USER_PROPERTY_LIST:
#         cell = ws.cell(column=col, row=row, value=property_name)
#         cell.font = title_font
#         col += 1
#
#     # 设置第二、三行的前三列固定数据
#     ws.cell(column=1, row=2, value="13966668888")
#     ws.cell(column=2, row=2, value=u"张三")
#     ws.cell(column=3, row=2, value=u"男")
#
#     ws.cell(column=1, row=3, value="15600339999")
#     ws.cell(column=2, row=3, value=u"李四")
#     ws.cell(column=3, row=3, value=u"女")
#
#     # 根据用户类型，设置二、三行的后几列数据
#     if user_flag == 0:
#         if USER_TYPE_PROVINCE_ADMIN in downloader_role:
#             ws.cell(column=4, row=2, value=u"省直属")
#             ws.cell(column=5, row=2, value=u"省教育局")
#             ws.cell(column=4, row=3, value=u"武汉")
#             ws.cell(column=5, row=3, value=u"")
#
#         elif USER_TYPE_CITY_ADMIN in downloader_role:
#             ws.cell(column=4, row=2, value=downloader_area_detail["city"])
#             ws.cell(column=5, row=2, value=u"市直属")
#             ws.cell(column=6, row=2, value=u"市教育局")
#             ws.cell(column=4, row=3, value=downloader_area_detail["city"])
#             ws.cell(column=5, row=3, value=get_rank_child_area_name(downloader_area_detail["city"])[1] if get_rank_child_area_name(
#                 downloader_area_detail["city"]) else '')
#             ws.cell(column=6, row=3, value=u"")
#
#         elif USER_TYPE_COUNTRY_ADMIN in downloader_role:
#             ws.cell(column=4, row=2, value=downloader_area_detail["city"])
#             ws.cell(column=5, row=2, value=downloader_area_detail["country"])
#             ws.cell(column=6, row=2, value=downloader_area_detail["country"] + u"教育局")
#             ws.cell(column=4, row=3, value=downloader_area_detail["city"])
#             ws.cell(column=5, row=3, value=downloader_area_detail["country"])
#             ws.cell(column=6, row=3, value=downloader_area_detail["country"] + u"教科所")
#
#         elif USER_TYPE_INSTITUTION_ADMIN in downloader_role:
#             ws.cell(column=4, row=2, value=downloader_area_detail["city"])
#             ws.cell(column=5, row=2, value=downloader_area_detail["country"])
#             ws.cell(column=6, row=2, value=downloader_area_detail["institution"])
#             ws.cell(column=4, row=3, value=downloader_area_detail["city"])
#             ws.cell(column=5, row=3, value=downloader_area_detail["country"])
#             ws.cell(column=6, row=3, value=downloader_area_detail["institution"])
#
#             ws.cell(column=1, row=3, value="G130879200308080030")
#
#         else:
#             raise BusinessException(ERR_USER_FLAG)
#
#     elif user_flag == 1:
#         # todo 此处需要判断一下是哪级的专家
#         ws.cell(column=4, row=2, value=u"省直属")
#         ws.cell(column=5, row=2, value=u"省教育局")
#         ws.cell(column=4, row=3, value=u"武汉")
#         ws.cell(column=5, row=3, value=u"武汉市教育局")
#
#     else:
#         return {"c": ERR_USER_FLAG[0], "m": ERR_USER_FLAG[1], "d": []}
#
#     wb.save(file_path)
#     return {"c": ERR_SUCCESS[0], "m": ERR_SUCCESS[1], "d": [file_path]}


# 创建市州区县的二级联动效果
def __set_city_country_data_validation(wb, ws, city_data_col, country_data_col):
    city_start_row = 1
    country_start_row = 2
    city_country_start_col = 50

    city_row = city_start_row
    city_col = city_country_start_col

    # 将市州、区县对应的数据写入excel
    province = Area.objects.get(area_level=LEVEL_PROVINCE, del_flag=FALSE_INT)
    city_list = get_child_by_areaid(province.id, manage_direct='0')
    for city in city_list:
        if not city['area_id']:
            continue

        country_list = get_child_by_areaid(city['area_id'], manage_direct='0')

        # 市州数据写入
        ws.cell(row=city_row, column=city_col, value=city['area_name'])

        country_row = country_start_row
        for country in country_list:
            # 对应市州的区县数据写入
            ws.cell(row=country_row, column=city_col, value=country['area_name'])
            country_row += 1

        # 对于市州中无区县（省直）的情况，做特殊处理
        if country_row == country_start_row:
            country_row += 1

        # 创建区县 name_range
        country_col_letter = get_column_letter(city_col)
        cell_range = u"%s%d:%s%d" % (country_col_letter, country_start_row, country_col_letter, country_row - 1)
        wb.create_named_range(name=city['area_name'], worksheet=ws, range=cell_range)

        city_col += 1

    # 创建市州的DV
    city_start_col_letter = get_column_letter(city_country_start_col)
    city_end_col_letter = get_column_letter(city_col - 1)
    cell_range = u"%s!$%s$%d:$%s$%d" % (ws.title, city_start_col_letter, 1, city_end_col_letter, 1)
    dv = DataValidation(type="list", formula1=cell_range.format(quote_sheetname("dv_sheet")))
    dv.error = u'请输入列表中的选项'
    dv.errorTitle = u'错误输入'
    ws.add_data_validation(dv)

    # apply the city validation to a range of cells
    city_data_col_letter = get_column_letter(city_data_col)
    cell_range = "%s%d:%s%d" % (city_data_col_letter, 2, city_data_col_letter, 1048576)
    dv.ranges.append(cell_range)

    # 创建 区县 DV
    for row in xrange(1000):
        if row < country_start_row:
            continue
        city_data_col_letter = get_column_letter(city_data_col)
        formula_str = "=INDIRECT($%s%d)" % (city_data_col_letter, row)
        dv = DataValidation(type="list", formula1=formula_str, allow_blank=True)
        dv.error = u'请输入列表中的选项'
        dv.errorTitle = u'错误输入'
        ws.add_data_validation(dv)

        # apply the country validation to a range of cells
        country_data_col_letter = get_column_letter(country_data_col)
        cell_name = "%s%d" % (country_data_col_letter, row)
        cell = ws[cell_name]
        dv.add(cell)


# 创建市州的下拉框选择效果
def __set_city_data_validation(wb, ws, city_data_col):
    city_start_row = 1
    city_start_col = 50

    city_row = city_start_row
    city_col = city_start_col

    province = Area.objects.get(area_level=LEVEL_PROVINCE, del_flag=FALSE_INT)
    city_list = get_child_by_areaid(province.id, manage_direct='0')
    for city in city_list:
        ws.cell(row=city_row, column=city_col, value=city['area_name'])
        city_col += 1

    # 创建市州的DV
    city_start_col_letter = get_column_letter(city_start_col)
    city_end_col_letter = get_column_letter(city_col - 1)
    cell_range = "%s!$%s$%d:$%s$%d" % (ws.title, city_start_col_letter, 1, city_end_col_letter, 1)
    dv = DataValidation(type="list", formula1=cell_range.format(quote_sheetname("dv_sheet")))
    dv.error = u'请输入列表中的选项'
    dv.errorTitle = u'错误输入'
    ws.add_data_validation(dv)

    # apply the city validation to a range of cells
    city_data_col_letter = get_column_letter(city_data_col)
    cell_range = "%s%d:%s%d" % (city_data_col_letter, 2, city_data_col_letter, 1048576)
    dv.ranges.append(cell_range)


def api_list_account_right(request, name_or_mobile, only_qry_my_right, rows, page, last_id):
    result = dict()
    # 检查是否运维人员
    if not only_qry_my_right and not request.user.auth:
        raise BusinessException(AUTH_WRONG_TYPE)

    account_rights = AccountRight.objects.filter(del_flag=FALSE_INT)

    # 非运维人员只能查询自己的
    if only_qry_my_right:
        account_rights = account_rights.filter(account=request.user)

    if name_or_mobile:
        account_rights = account_rights.filter(Q(account__name__contains=name_or_mobile) | Q(account__mobile__contains=name_or_mobile))

    # 分页
    cnt = len(account_rights)
    num_pages = 1
    if rows and page:
        num_pages, cur_start, cur_end = get_pages(cnt, page, rows)
        account_rights = account_rights[cur_start:cur_end]
    elif rows:
        cur_start = get_lastid_index(account_rights, last_id)
        account_rights = account_rights[cur_start:cur_start + int(rows)]
        pass

    result["max_page"] = num_pages
    result["total"] = cnt

    account_rights_list = list()
    for each_accountright in account_rights:
        account_rights_dict = {
            'id': each_accountright.id,
            'account_name': each_accountright.account.name,
            'account_mobile': each_accountright.account.mobile,
            'area_id': each_accountright.area_id,
            'area_fullname': each_accountright.area_fullname,
            'create_count': each_accountright.create_count,
        }
        account_rights_list.append(account_rights_dict)

    result['account_rights_list'] = account_rights_list
    return result


def api_list_account(request, name_or_mobile, rows, page, last_id):
    """
    根据名称过滤帐户列表
    :param request:
    :param name_or_mobile:
    :return:
    """
    result = dict()
    # 查询全部用户慢，对需求也没必要暂时不支持。
    if not name_or_mobile:
        return result

    accounts = Account.objects.filter(del_flag=FALSE_INT)
    if name_or_mobile:
        accounts = accounts.filter(Q(name__contains=name_or_mobile) | Q(mobile__contains=name_or_mobile))

    # 分页
    cnt = len(accounts)
    num_pages = 1
    if rows and page:
        num_pages, cur_start, cur_end = get_pages(cnt, page, rows)
        accounts = accounts[cur_start:cur_end]
    elif rows:
        cur_start = get_lastid_index(accounts, last_id)
        accounts = accounts[cur_start:cur_start + int(rows)]
        pass

    result["max_page"] = num_pages
    result["total"] = cnt

    account_list = list()
    for each_account in accounts:
        account_dict = {
            'id': each_account.id,
            'name': each_account.name,
            'mobile': each_account.mobile,
        }
        account_list.append(account_dict)

    result['account_list'] = account_list
    return result


def api_add_account_right(request, account_id, area_id):
    result = dict()
    # 检查是否运维人员
    if not request.user.auth:
        raise BusinessException(AUTH_WRONG_TYPE)

    # 检查是否已有此权限
    accountright = AccountRight.objects.filter(account_id=account_id, area_id=area_id, del_flag=FALSE_INT)
    if accountright:
        raise BusinessException(ERR_USER_HAS_PERMISSION)

    # 检查account_id是否正确
    accounts = Account.objects.filter(id=account_id, del_flag=FALSE_INT).first()
    if not accounts:
        raise BusinessException(ERR_USER_NOT_EXIST)

    # 查询area_id的所有上级
    area = get_areadetail_by_id(area_id)
    if area['province']:
        area_fullname = area['province']
    else:
        area_fullname = area['nation']

    area_fullname = '%s-%s-%s-%s' % (area_fullname, area['city'], area['country'], area['institution'])
    area_fullname = area_fullname.strip('-')

    accountright = AccountRight()
    accountright.account_id = account_id
    accountright.area_fullname = area_fullname
    accountright.area_id = area_id
    accountright.admit_account = request.user
    accountright.save()

    return result


def api_del_account_right(request, ids):
    result = dict()
    # 检查是否运维人员
    if not request.user.auth:
        raise BusinessException(AUTH_WRONG_TYPE)

    if not ids:
        raise BusinessException(REQUEST_PARAM_ERROR)
    id_list = ids.strip(',').split(',')

    rows = AccountRight.objects.filter(id__in=id_list, del_flag=FALSE_INT).update(del_flag=TRUE_INT)
    result['rows'] = rows
    return result


def api_account_dataconfirm(request, name, sex, area_id, area_name, manage_direct, institution, position, region_id, email):
    """
    用户首次登陆时，确认用户信息。因为用户可能是别人导入进去的，资料不是自己填的
    :param request:
    :param name:
    :param sex:
    :param area_id:
    :param area_name:
    :param manage_direct:
    :param institution:
    :param position:
    :return:
    """
    # 检查用户资料是否已经确认。
    if request.user.is_data_confirm:
        raise BusinessException(ERR_ACCOUNTDATA_IS_CONFIRM)

    api_modi_accountdata(request.user, name, sex, area_id, area_name, manage_direct, institution, position, region_id, email, is_self_reg='')
    return True


def api_modi_accountdata(user, name, sex, area_id, area_name, manage_direct, institution, position, region_id, email, is_self_reg):
    # 注意：具体地区几个参数的传参方法见接口api_account_dataconfirm
    # 先保存帐户资料
    if name:
        user.name = name
    if sex:
        user.sex = sex
    if region_id:
        user.region_id = region_id
    if email:
        user.email = email
    if is_self_reg:
        user.is_self_reg = is_self_reg
    user.save()

    # 检查area_id是否存在
    area = None
    if area_id:
        area = Area.objects.filter(del_flag=FALSE_INT, id=area_id).first()
        if not area:
            raise BusinessException(ERR_AREA_ERROR)

    # 检查用户是不是专家再保存专家信息
    expert = Expert.objects.filter(del_flag=FALSE_INT, account=user).first()
    if expert:
        if not area:
            raise BusinessException(ERR_ACCOUNTDATA_NOT_COMPLETE)

        if name:
            expert.name = name
        if sex:
            expert.sex = sex
        if institution:
            expert.institution = institution
        if position:
            expert.position = position

        # 如果是直属，需要先创建直属区域信息
        if manage_direct and int(manage_direct):
            area_zone, _ = Area.objects.get_or_create(del_flag=FLAG_NO, area_name=institution, parent=area,
                                                      manage_direct=FLAG_YES, area_level=area.area_level >> 1)
        else:
            if area.area_level > LEVEL_COUNTY:
                area_zone = Area.objects.filter(del_flag=FLAG_NO, area_name=area_name, parent=area).first()
            else:
                area_zone, _ = Area.objects.get_or_create(area_name=institution, parent=area, del_flag=FLAG_NO,
                                                          area_level=area.area_level >> 1)
        if not area_zone:
            raise BusinessException(REQUEST_PARAM_ERROR)

        expert.area = area_zone

        expert.save()

    user.is_data_confirm = 1
    user.save()

    return True


def api_send_smsverifycode(user, mobile):
    # 检查电话号码是否正常
    if not is_mobile(mobile):
        raise BusinessException(ERR_ERR_MOBILE)

    # 检查用户是否注册过本系统
    # account = Account.objects.filter(mobile=mobile, del_flag=FALSE_INT)
    # if not account:
    #     raise BusinessException(ERR_USER_NOT_EXIST)

    # 先将原来的验证码失效，再生成新的验证码。
    VerifyCode.objects.filter(mobile=mobile, del_flag=FLAG_NO).update(del_flag=FLAG_YES)
    VerifyCode.objects.create(mobile=mobile, IMCode_status=FLAG_YES)
    # 发送验证码
    send_message(mobile)
    return ''


def api_check_smsverifycode(user, mobile, smscode):
    # 检查电话号码是否正常
    if not is_mobile(mobile):
        raise BusinessException(ERR_ERR_MOBILE)

    # 检查验证码
    verify_messagecode(mobile, smscode)
    return ''


def api_reset_forget_password(user, mobile, smscode, new_password):
    """
    忘记密码后重置密码, 避免使用token麻烦，此处就只使用smscode进行验证
    :param user:
    :param mobile:
    :param smscode:
    :param new_password:
    :return:
    """
    # 检查验证码是否正确
    verify_messagecode(mobile, smscode)

    # 检查用户是否注册过本系统
    account = Account.objects.filter(mobile=mobile, del_flag=FALSE_INT).first()
    if not account:
        raise BusinessException(ERR_USER_NOT_EXIST)

    # 修改密码
    account.set_password(new_password)
    account.encoded_pwd = xor_crypt_string(data=new_password, encode=True)
    account.save()

    # 删除验证码
    del_messagecode(mobile)

    return ''


def api_update_region_fullname(region_id):
    regions = Region.objects.all()
    if region_id:
        regions = regions.filter(id=region_id)

    for each_region in regions:
        region_detail = get_regiondetail_by_id(each_region.id)
        region_fullname = '%s%s%s%s' % (region_detail['province'], region_detail['city'], region_detail['country'], region_detail['institution'])
        each_region.region_fullname = region_fullname
        each_region.save()

    return 'ok'


def api_update_area_fullname(area_id):
    update_area_fullname(area_id)
    return 'ok'


def api_list_region(request, region_name, region_level, rows, page, last_id):
    result = dict()
    regions = Region.objects.filter(region_fullname__contains=region_name, del_flag=FALSE_INT)
    if region_level:
        regions = regions.filter(region_level=region_level)

    # 分页
    cnt = len(regions)
    num_pages = 1
    if rows and page:
        num_pages, cur_start, cur_end = get_pages(cnt, page, rows)
        regions = regions[cur_start:cur_end]
    elif rows:
        cur_start = get_lastid_index(regions, last_id)
        regions = regions[cur_start:cur_start + int(rows)]
        pass

    result["max_page"] = num_pages
    result["total"] = cnt
    region_list = list()
    for each_region in regions:
        region_dict = {
            'id': each_region.id,
            'region_code': each_region.region_code,
            'region_level': each_region.region_level,
            'region_name': each_region.region_name,
            'is_school': each_region.is_school,
            'region_fullname': each_region.region_fullname,
        }
        region_list.append(region_dict)
    result['region_list'] = region_list
    return result


def api_mod_account(request, name, sex, area_id, area_name, manage_direct, institution, position, region_id, email, is_self_reg):
    result = api_modi_accountdata(request.user, name, sex, area_id, area_name, manage_direct, institution, position, region_id, email, is_self_reg)
    return result


def check_account_reg_param(request, name, sex, area_id, area_name, manage_direct, institution, position, region_id, email, mobile, smscode, password, confirm_code):
    if area_id:
        area = Area.objects.filter(id=area_id)
        if not area:
            raise BusinessException(ERR_AREA_ERROR)

    if region_id:
        region = Region.objects.filter(id=region_id)
        if not region:
            raise BusinessException(ERR_SCHOOL_REGION_ERROR)

    if mobile or smscode:
        ret = __check_username_valid(mobile)
        if ret['c'] != SUCCESS[0]:
            raise BusinessException(ERR_ERR_MOBILE)

        if not is_mobile(mobile):
            raise BusinessException(ERR_ERR_MOBILE)

    if smscode:
        verify_messagecode(mobile, smscode)

    if password:
        if len(password) < 6:
            raise BusinessException(ERR_ERR_SHORT_PWD)

    if confirm_code:
        # 检查微信码是否正确
        weixinscanconfirm = WeixinScanConfirm.objects.filter(code=confirm_code, del_flag=FALSE_INT, busitype=WEIXIN_SCAN_TYPE_REG).first()
        if not weixinscanconfirm:
            raise BusinessException(ERR_QRCODE_TIMEOUT)

        # 检查微信是否已经扫过码了
        if not weixinscanconfirm.openid or not weixinscanconfirm.openid_fh:
            raise BusinessException(ERR_WEIXIN_NOT_SCAN)

    return 'ok'


@transaction.atomic
def api_account_reg(request, name, sex, area_id, area_name, manage_direct, institution, position, region_id, email, mobile, smscode, password, confirm_code):
    # 先进行参数检查
    if not region_id or not mobile or not password or not name:
        raise BusinessException(REQUEST_PARAM_ERROR)

    check_account_reg_param(request, name, sex, area_id, area_name, manage_direct, institution, position, region_id, email, mobile, smscode, password, confirm_code)

    # 查询帐户是否已经存在
    username = mobile
    account = Account.objects.filter(del_flag=FLAG_NO, username=username).first()
    if account:
        raise BusinessException(ERR_USER_EXISTS)

    # 创建帐号
    account = Account.objects.create_user(username, password=password, mobile=mobile)

    api_modi_accountdata(account, name, sex, area_id, area_name, manage_direct, institution, position, region_id, email, is_self_reg=1)

    # 先取消该微信绑定的其它帐户，再将新微信绑定到新帐户
    if confirm_code:
        weixinscanconfirm = WeixinScanConfirm.objects.filter(code=confirm_code, del_flag=FALSE_INT).first()
        if not weixinscanconfirm:
            raise BusinessException(ERR_QRCODE_TIMEOUT)

        WeixinAccount.objects.filter(openid=weixinscanconfirm.openid, del_flag=FALSE_INT).update(del_flag=TRUE_INT)

        weixinscanconfirm.account = account
        weixinscanconfirm.save()

        weixinaccount = WeixinAccount()
        weixinaccount.account = weixinscanconfirm.account
        weixinaccount.openid = weixinscanconfirm.openid
        weixinaccount.openid_fh = weixinscanconfirm.openid_fh
        weixinaccount.save()

    # 删除验证码
    del_messagecode(mobile)

    return account.id


def api_qry_institution_reg(user, cur_user_id, rows, page, last_id):
    result = dict()
    cur_user = User.objects.filter(del_flag=FLAG_NO, id=cur_user_id).first()
    if cur_user.account != user:
        return dict(c=ERR_USER_AUTH[0], m=ERR_USER_AUTH[1])

    # 检查当前用户是否为学校管理员
    # if cur_user.area.area_level != LEVEL_INSTITUTION:
    #     raise BusinessException(ERR_ERR_NOT_INSTITUTION_ADMIN)

    if not cur_user.area.region:
        raise BusinessException(ERR_ERR_NOT_INSTITUTION_ADMIN)

    institution_accounts = Account.objects.filter(del_flag=FALSE_INT, region=cur_user.area.region)

    # 分页
    cnt = len(institution_accounts)
    num_pages = 1
    if rows and page:
        num_pages, cur_start, cur_end = get_pages(cnt, page, rows)
        institution_accounts = institution_accounts[cur_start:cur_end]
    elif rows:
        cur_start = get_lastid_index(institution_accounts, last_id)
        institution_accounts = institution_accounts[cur_start:cur_start + int(rows)]
        pass

    result["max_page"] = num_pages
    result["total"] = cnt

    account_list = list()
    for each_account in institution_accounts:
        row_dict = {
            "account_id": each_account.id,
            "account_name": each_account.name,
            "account_sex": each_account.sex,
            "account_mobile": each_account.mobile,
        }
        account_list.append(row_dict)
    result["account_list"] = account_list
    return result


def api_qry_area_dropdowndetail(user, area_id):
    result = dict()
    area_detail = get_areadetail_by_id(area_id)

    result['nation'] = area_detail['nation']
    result['area_id_nation'] = area_detail['area_id_nation']
    # result['area_detail_nation'] = ''
    result['province'] = area_detail['province']
    result['area_id_province'] = area_detail['area_id_province']
    result['area_detail_province'] = ''
    result['city'] = area_detail['city']
    result['area_id_city'] = area_detail['area_id_city']
    result['area_detail_city'] = ''
    result['country'] = area_detail['country']
    result['area_id_country'] = area_detail['area_id_country']
    result['area_detail_country'] = ''
    result['institution'] = area_detail['institution']
    result['area_id_institution'] = area_detail['area_id_institution']
    result['area_detail_institution'] = ''

    result['area_detail_province'] = list_sub_area(user=user, cur_user_id='', area_id=area_detail['area_id_nation'], manage_direct='', area_name='',
                                                   is_school='')['d'] if area_detail['area_id_nation'] else ''
    result['area_detail_city'] = list_sub_area(user=user, cur_user_id='', area_id=area_detail['area_id_province'], manage_direct='', area_name='',
                                               is_school='')['d'] if area_detail['area_id_province'] else ''
    result['area_detail_country'] = list_sub_area(user=user, cur_user_id='', area_id=area_detail['area_id_city'], manage_direct='', area_name='',
                                                  is_school='')['d'] if area_detail['area_id_city'] else ''
    result['area_detail_institution'] = list_sub_area(user=user, cur_user_id='', area_id=area_detail['area_id_country'], manage_direct='',
                                                      area_name='', is_school='')['d'] if area_detail['area_id_country'] else ''

    return result
