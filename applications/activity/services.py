#!/usr/bin/env python
# coding=utf-8
import os
import datetime
import json
import copy
import uuid
from io import BytesIO
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import quote_sheetname

from django.db import transaction
from django.db.models import Q, Count, F

from applications.activity.guess import GuessArea
from applications.activity.models import *
from applications.activity.share import area_struct, is_activity_owner, is_devops, expert, china
from applications.common.services import area_name, get_areadetail_by_id
from applications.subjudge.models import SubJudgeExpert, SubJudge, SubJudgeTeamExpert, SubJudgeRule
from applications.user.models import *
from applications.work.models import *
from utils.const_def import *
from utils.const_err import *
from utils.public_fun import str_p_datetime, datetime_f_str, paging_by_page, get_pages, get_lastid_index, const_stage, fname_unique, excel_cell2str
from utils.file_fun import get_image_url, gen_path
from utils.store_file import save_file
from utils.utils_except import BusinessException
from utils.utils_type import str2bool, bool2str


AREA_DEF = {
    # 表头显示  / 下拉框特殊项  / 是否要补充机构
    0b10000: (u'省或直辖市', u'国直属', True),
    0b01000: (u'市州', u'省直属', True),
    0b00100: (u'区县', u'市直属', True),
    0b00010: (u'学校或机构', u'', False),
}


def list_activity_create_area(account):
    result = []
    auth_list = AccountRight.objects.filter(del_flag=FLAG_NO, account=account)
    for each in auth_list:
        result.append(area_struct(each.area))
    return dict(c=SUCCESS[0], m=SUCCESS[1], d=result)


def _time_display(w, start, end):
    # 赛事豆腐块时间展示方式
    s = datetime_f_str(start, str_format='%Y-%m-%d') if start else '~'
    e = datetime_f_str(end, str_format='%Y-%m-%d') if end else '~'
    return '%s: %s - %s' % (w, s, e)


def my_activities_create(account):
    # 我创建的活动列表信息
    activity_create_list = Activity.objects.filter(
        user__account=account, template_flag=FLAG_NO,
        stage__in=[ACTIVITY_STAGE_UPLOAD, ACTIVITY_STAGE_GROUP, ACTIVITY_STAGE_REVIEW, ACTIVITY_STAGE_PUBLIC, ACTIVITY_STAGE_ARCHIVE],
    ).order_by('-open_time', '-create_time').values(
        "id", "name", "user_id", "stage", "browse_count", "work_count",
        "user__area_id", "user__area__area_name", "user__area__area_code", "user__area__area_level", "user__area__manage_direct",
        "user__area__parent__id", "user__area__parent__area_name", "user__area__parent__area_level",
        "upload_time", "public_time", "group_time", "review_time", "archive_time", "organizers", "banner__url")
    # 我的草稿箱信息活动列表信息
    activity_draft_list = Activity.objects.filter(
        user__account=account, template_flag=FLAG_NO, stage=ACTIVITY_STAGE_EDIT,) \
        .values("id", "name", "organizers", "banner__url").order_by('-update_time', '-create_time')
    # 组装返回的数据
    result = {
        'doing': list(),
        'done': list(),
        'draft': list(),
    }
    for c in activity_create_list:
        sub_lv, sub_lv_desc = subordinate_area_level(c['user__area__area_level'])
        data = dict(
                activity_id=str(c['id']), activity_name=c['name'], stage=str(c['stage']),
                browse_count=str(c['browse_count']), work_count=str(c['work_count']),
                time_display=_time_display(u'赛事时间', c['upload_time'], c['archive_time']),
                organizer=c['organizers'], banner_url=get_image_url(c['banner__url']),

                area_id=str(c['user__area_id']), area_code=c['user__area__area_code'],
                area_level=str(c['user__area__area_level']), area_is_direct=str(c['user__area__manage_direct']),
                area_name_simple=area_name(c['user__area_id'], full=False),
                area_name_full=area_name(c['user__area_id']),
                area_direct_level=str(c['user__area__parent__area_level']) if c['user__area__parent__area_level'] else '',
                area_parent_id=str(c['user__area__parent__id']) if c['user__area__parent__id'] else '',
                area_parent_name_simple=c['user__area__parent__area_name'] if c['user__area__parent__area_name'] else '',

                type=str(WHO.CREATOR), user_id=str(c['user_id']),
                user_area_id=str(c['user__area_id']),
                subordinate_area_level=str(sub_lv), subordinate_area_level_name=sub_lv_desc,
        )
        if c['stage'] != ACTIVITY_STAGE_ARCHIVE:
            result['doing'].append(data)
        else:
            result['done'].append(data)
    for d in activity_draft_list:
        data = dict(activity_id=str(d['id']), activity_name=d['name'], organizer=d['organizers'], banner_url=get_image_url(d['banner__url']))
        result['draft'].append(data)
    result['is_empty'] = bool2str(len(activity_create_list) == 0)
    return result


def _decide_subjudge_info(activity_id, user_area):
    result = {}
    need_decide_judge = \
        user_area.manage_direct == FALSE_INT \
        and user_area.area_level in (LEVEL_PROVINCE, LEVEL_CITY, LEVEL_COUNTY)

    # 获取该活动的评审模式
    judge_rule = ''
    subjudge_id = ''
    sbj = SubJudge.objects.filter(activity_id=int(activity_id), area_id=user_area.id).first()
    is_decide_judge = TRUE_STR if sbj else FALSE_STR
    if sbj:
        if sbj.is_active == TRUE_INT:
            subjudge_id = str(sbj.id)
            subjg_rule = SubJudgeRule.objects.filter(subjudge=sbj).first()
            if subjg_rule:
                judge_rule = str(subjg_rule.code)
    result['need_decide_judge'] = bool2str(need_decide_judge)
    result['is_decide_judge'] = is_decide_judge
    result['subjudge_id'] = subjudge_id
    result['judge_rule'] = judge_rule
    return result


def my_activities_join(account):
    # 我参加的活动列表信息
    activity_join_list = Role.objects.filter(user__account=account, activity__stage__in=[
            ACTIVITY_STAGE_UPLOAD, ACTIVITY_STAGE_GROUP, ACTIVITY_STAGE_REVIEW, ACTIVITY_STAGE_PUBLIC, ACTIVITY_STAGE_ARCHIVE]). \
        exclude(activity__user__account=account).\
        order_by('-activity__open_time', '-activity__create_time').values(
        "activity_id", "activity__name", "user_id", "activity__stage", "activity__browse_count", "activity__work_count",
        "activity__upload_time", "activity__public_time", "activity__group_time", "activity__review_time", "activity__archive_time",
        "activity__user__area_id", "activity__user__area__area_name", "activity__user__area__area_code", "activity__user__area__area_level",
        "activity__user__area__manage_direct", "activity__user__area__parent__id",
        "activity__user__area__parent__area_name", "activity__user__area__parent__area_level",
        "user__area_id", "user__area__area_name", "user__area__area_code", "user__area__area_level", "user__area__manage_direct",
        "activity__organizers", "activity__banner__url")
    result = {
        'doing': list(),
        'done': list(),
    }
    for each in activity_join_list:
        sub_lv, sub_lv_desc = subordinate_area_level(each['user__area__area_level'])

        user_area = Area.objects.filter(id=int(each['user__area_id'])).first()
        # 是否需要选择审批模式
        subjudge_info_dict = _decide_subjudge_info(each['activity_id'], user_area)

        # prepare addtional info
        area = Area.objects.filter(id=int(each['user__area_id'])).first()
        # has_account_registered = Account.objects.filter(del_flag=FALSE_INT, region=area.region).exclude(region__isnull=True).exists()
        if (account.region != None) or (area.area_level == 0) or (area.parent.manage_direct == TRUE_INT):
            additional_info = 1  # 是末端用户
        elif (area.manage_direct == TRUE_INT or (area.manage_direct == FALSE_INT and area.area_level == 0b00001))\
                and area.region != None:
            additional_info = 2  # 是学校管理员
        else:
            additional_info = 0

        # 全局关闭了子级评审功能
        if not settings.USE_SUBJUDGE:
            subjg, is_created = SubJudge.objects.update_or_create(
                activity__id=int(each['activity_id']), user__id=int(each['user_id']), area__id=int(each['user__area_id']),
                defaults={'is_active': FALSE_INT}, )
            subjudge_info_dict['need_decide_judge'] = FALSE_STR

        data = dict(
            activity_id=str(each['activity_id']), activity_name=each['activity__name'], stage=str(each['activity__stage']),
            browse_count=str(each['activity__browse_count']), work_count=str(each['activity__work_count']),
            time_display=_time_display(u'参赛时间', each['activity__upload_time'], each['activity__group_time']),
            organizer=each['activity__organizers'], banner_url=get_image_url(each['activity__banner__url']),

            area_id=str(each['activity__user__area_id']), area_code=each['activity__user__area__area_code'],
            area_level=str(each['activity__user__area__area_level']), area_is_direct=str(each['activity__user__area__manage_direct']),
            area_name_simple=each['activity__user__area__area_name'],
            area_name_full=area_name(each['activity__user__area_id']),
            area_direct_level=str(each['activity__user__area__parent__area_level']) if each['activity__user__area__parent__area_level'] else '',
            area_parent_id=str(each['activity__user__area__parent__id']) if each['activity__user__area__parent__id'] else '',
            area_parent_name_simple=each['activity__user__area__parent__area_name'] if each['activity__user__area__parent__area_name'] else '',

            type=str(WHO.PARTICIPANT), user_id=str(each['user_id']),
            additional_info=str(additional_info),
            user_area_is_direct=str(each['user__area__manage_direct']),
            user_area_id=str(each['user__area_id']), user_area_level=str(each['user__area__area_level']),
            user_area_name_simple=each['user__area__area_name'], user_area_name_full=area_name(each['user__area_id']),
            subordinate_area_level=str(sub_lv), subordinate_area_level_name=sub_lv_desc,

            need_decide_judge=subjudge_info_dict['need_decide_judge'],
            is_decide_judge=subjudge_info_dict['is_decide_judge'],
            subjudge_id=subjudge_info_dict['subjudge_id'],
            judge_rule=subjudge_info_dict['judge_rule'],
        )

        if each['activity__stage'] != ACTIVITY_STAGE_ARCHIVE:
            result['doing'].append(data)
        else:
            result['done'].append(data)
    result['is_empty'] = bool2str(len(activity_join_list) == 0)
    return result


def my_activities_judge(account):
    # 我参加评审的活动列表信息
    qs = ExpertActivity.objects.filter(expert__del_flag=FALSE_INT, expert__account=account)
    # if not settings.DEBUG:
    if True:
        # 豆腐块仅可见“正在评审”的活动
        qs = qs.filter(activity__stage=ACTIVITY_STAGE_REVIEW)
    activity_judge_list = qs.order_by('-activity__open_time', '-activity__create_time').values(
            "activity_id", "activity__name", "expert_id", "activity__stage", "activity__browse_count", "activity__work_count",
            "activity__upload_time", "activity__public_time", "activity__group_time", "activity__review_time", "activity__archive_time",
            "activity__user__area_id", "activity__user__area__area_code",
            "activity__user__area__area_name", "activity__user__area__area_level",
            "activity__user__area__manage_direct", "activity__user__area__parent__id",
            "activity__user__area__parent__area_name", "activity__user__area__parent__area_level",
            "activity__organizers", "activity__banner__url"
    )

    # 我参加子级评审的活动列表信息
    # 只有整体在上传阶段子级评审专家才可以看见子级评审
    qs = SubJudgeExpert.objects.filter(expert__account=account,
                subjudge__is_active=TRUE_INT, subjudge__status=TRUE_INT, subjudge__activity__stage=ACTIVITY_STAGE_UPLOAD)
    activity_subjudge_list = qs.order_by('-subjudge__activity__open_time').values(
            "subjudge__activity_id", "subjudge__activity__name", "subjudge__activity__stage", "expert_id",
            "subjudge__activity__browse_count", "subjudge__activity__work_count",
            "subjudge__activity__upload_time", "subjudge__activity__public_time", "subjudge__activity__group_time",
            "subjudge__activity__review_time", "subjudge__activity__archive_time",
            "subjudge__activity__user__area_id", "subjudge__activity__user__area__area_code",
            "subjudge__activity__user__area__area_name", "subjudge__activity__user__area__area_level",
            "subjudge__activity__user__area__manage_direct", "subjudge__activity__user__area__parent__id",
            "subjudge__activity__user__area__parent__area_name", "subjudge__activity__user__area__parent__area_level",

            "subjudge_id",
            "subjudge__area_id", "subjudge__area__area_code", "subjudge__area__area_name", "subjudge__area__area_level",
            "subjudge__activity__organizers", "subjudge__activity__banner__url"
    )
    result = {
        'doing': list(),
        'done': list(),
    }
    for each in activity_judge_list:
        te = TeamExpert.objects.filter(del_flag=FALSE_INT, expert_id=each['expert_id']).first()
        is_leader = str(te.is_leader) if te else FALSE_STR
        data = dict(
            type=str(WHO.EXPERT), expert_id=str(each['expert_id']), is_leader=is_leader,
            activity_id=str(each['activity_id']), activity_name=each['activity__name'],
            browse_count=str(each['activity__browse_count']), work_count=str(each['activity__work_count']),
            stage=str(each['activity__stage']),
            time_display=_time_display(u'评审时间', each['activity__review_time'], each['activity__public_time']),
            organizer=each['activity__organizers'], banner_url=get_image_url(each['activity__banner__url']),

            area_id=str(each['activity__user__area_id']), area_name=each['activity__user__area__area_name'],
            area_code=each['activity__user__area__area_code'], area_level=str(each['activity__user__area__area_level']),
            area_fullname=area_name(each['activity__user__area_id']),
            area_is_direct=str(each['activity__user__area__manage_direct']),
            area_direct_level=str(each['activity__user__area__parent__area_level']) if each['activity__user__area__parent__area_level'] else '',
            area_parent_id=str(each['activity__user__area__parent__id']) if each['activity__user__area__parent__id'] else '',
            area_parent_name_simple=each['activity__user__area__parent__area_name'] if each['activity__user__area__parent__area_name'] else '',

            judge_flag=str(JUDGE_MAIN), subjudge_id='', subjudge_area_id='',
            subjudge_area_name_full='', subjudge_area_name_simple='',
            subjudge_area_level='',
        )
        if each['activity__stage'] != ACTIVITY_STAGE_ARCHIVE:
            result['doing'].append(data)
        else:
            result['done'].append(data)
    for each in activity_subjudge_list:
        sjte = SubJudgeTeamExpert.objects.filter(expert_id=each['expert_id'], subjudge_id=each['subjudge_id']).first()
        is_leader = str(sjte.is_leader) if sjte else FALSE_STR
        data = dict(
            type=str(WHO.EXPERT), expert_id=str(each['expert_id']), is_leader=is_leader,

            activity_id=str(each['subjudge__activity_id']), activity_name=each['subjudge__activity__name'],
            browse_count=str(each['subjudge__activity__browse_count']), work_count=str(each['subjudge__activity__work_count']),
            stage=str(each['subjudge__activity__stage']), time_display=u'评审截止时间请联系赛事管理员',
            organizer=each['subjudge__activity__organizers'], banner_url=get_image_url(each['subjudge__activity__banner__url']),

            area_id=str(each['subjudge__activity__user__area_id']), area_name_simple=each['subjudge__activity__user__area__area_name'],
            area_code=each['subjudge__activity__user__area__area_code'], area_level=str(each['subjudge__activity__user__area__area_level']),
            area_name_full=area_name(each['subjudge__activity__user__area_id'], joint=''),
            area_is_direct=str(each['subjudge__activity__user__area__manage_direct']),
            area_direct_level=str(each['subjudge__activity__user__area__parent__area_level']) if each['subjudge__activity__user__area__parent__area_level'] else '',
            area_parent_id=str(each['subjudge__activity__user__area__parent__id']) if each['subjudge__activity__user__area__parent__id'] else '',
            area_parent_name_simple=each['subjudge__activity__user__area__parent__area_name'] if each['subjudge__activity__user__area__parent__area_name'] else '',

            judge_flag=str(JUDGE_SUB), subjudge_id=str(each['subjudge_id']),
            subjudge_area_name_full=area_name(each['subjudge__area_id']), subjudge_area_id=str(each['subjudge__area_id']),
            subjudge_area_name_simple=each['subjudge__area__area_name'], subjudge_area_level=each['subjudge__area__area_level'],
        )
        if each['subjudge__activity__stage'] != ACTIVITY_STAGE_ARCHIVE:
            result['doing'].append(data)
        else:
            result['done'].append(data)
    result['is_empty'] = bool2str((len(activity_judge_list)==0 and len(activity_subjudge_list)==0))
    return result


def list_activity_category(account, category):
    categories = ['1', '2', '3'] if not category or not category.strip() else category.strip().split(',')
    result = dict()
    if '1' in categories:
        result['create'] = my_activities_create(account)
    if '2' in categories:
        result['join'] = my_activities_join(account)
    if '3' in categories:
        result['judge'] = my_activities_judge(account)
    return result


def stage_check(activity, new_stage):
    if abs(new_stage - activity.stage) > 1:
        raise BusinessException(ERR_STAGE_JUMP)

    if new_stage == ACTIVITY_STAGE_REVIEW:
        # 存在缺少的组长组不能进入评审
        teams = Team.objects.filter(activity=activity).all()
        for each in teams:
            if not TeamExpert.objects.filter(team=each, is_leader=TRUE_INT).exists():
                raise BusinessException(ERR_TEAM_WITHOUT_LEADER)


# 修改模板  修改活动  切换活动阶段
@transaction.atomic()
def edit_activity(account, name, activity_id, stage,
                        start_time, upload_time, group_time, review_time, public_time, archive_time,
                        organizer, participator, banner_id, attachment_id,
                        introduction, author_count, tutor_count,
                        base_info_value, copyright, is_top, is_minor, genre):
    # 增加查询锁，防止并行调用覆盖更新
    activity = Activity.objects.filter(id=int(activity_id)).select_for_update().first()

    # 检查权限
    if not is_devops(account) and not is_activity_owner(activity, account):
        raise BusinessException(ERR_USER_AUTH)

    if name:
        if Activity.objects.filter(name=name).exclude(id=activity.id).exists():
            raise BusinessException(ERR_ACTIVITY_NAME_CONFLICT)
        activity.name = name

    if upload_time:
        activity.upload_time = str_p_datetime(upload_time)

    if group_time:
        activity.group_time = str_p_datetime(group_time)

    if review_time:
        activity.review_time = str_p_datetime(review_time)

    if public_time:
        activity.public_time = str_p_datetime(public_time)

    if archive_time:
        activity.archive_time = str_p_datetime(archive_time)

    if attachment_id:
        activity.attachment_id = int(attachment_id)
    else:
        activity.attachment = None   # 取消附件

    if banner_id:
        activity.banner_id = int(banner_id)
    else:
        activity.banner = None   # 取消题图

    if introduction:
        activity.introduction = introduction
    else:
        if introduction == '':
            activity.introduction = ''

    if author_count:
        activity.author_count = int(author_count)

    if tutor_count:
        activity.tutor_count = int(tutor_count)

    if organizer:
        activity.organizers = organizer
    else:
        if organizer == '':
            activity.organizers = ''

    if participator:
        activity.participator = participator
    else:
        if participator == '':
            activity.participator = ''

    if stage:
        stage_check(activity, int(stage))
        activity.stage = int(stage)

    if base_info_value:
        activity.base_info_value = base_info_value

    if copyright:
        activity.copyright = copyright
    else:
        if copyright == '':
            activity.copyright = ''

    if is_top:
        activity.is_top = int(is_top)

    if is_minor:
        activity.is_top = int(is_minor)

    if genre:
        activity.genre = int(genre)

    activity.save()

    if name:
        logger.info('after name change ---> name: %s, copyright: %s' % (activity.name, activity.copyright))
    if copyright:
        logger.info('after copyright change ---> name: %s, copyright: %s' % (activity.name, activity.copyright))

    return {'activity_id': str(activity.id)}


def activity_info_update_status(activity, stage_list):
    if activity.stage in stage_list:
        if ACTIVITY_STAGE_UPLOAD in stage_list and ACTIVITY_STAGE_UPLOAD == activity.stage:
            if Work.objects.filter(del_flag=DEL_FLAG_NO, activity=activity).exists():
                return False
        return True
    else:
        return False


def _gen_work_attr(activity, category, info_list, count=1, mandatory_first=True):
    if not info_list:
        pass
    for grp in xrange(1, count+1):
        for i, each in enumerate(info_list):
            if mandatory_first:
                mandatory = int(each["mandatory"]) if grp == 1 else 0
            else:
                mandatory = int(each["mandatory"])
            WorkAttr.objects.create(activity=activity, category=category, group_sn=grp,
                                    sn=i + 1, name=each["name"], type=int(each["type"]),
                                    values=each["values"], mandatory=mandatory)


@transaction.atomic()
def edit_work_attr(account, activity_id, category, work_info_list, count):
    category = int(category)

    # 增加查询锁，防止edit_activity时并行调用覆盖更新
    activity = Activity.objects.filter(id=activity_id).select_for_update().first()
    if not is_devops(account) and not is_activity_owner(activity, account):
        raise BusinessException(ERR_USER_AUTH)

    WorkAttr.objects.filter(activity=activity, category=category, del_flag=DEL_FLAG_NO).update(del_flag=DEL_FLAG_YES)

    if ACTIVITY_CATEGORY_WORK_ATTR == category:
        _gen_work_attr(activity, category, work_info_list, mandatory_first=False)
    elif ACTIVITY_CATEGORY_AUTHOR_ATTR == category:
        activity.author_count = int(count)
        activity.save()
        _gen_work_attr(activity, category, work_info_list, count=int(count))
    elif ACTIVITY_CATEGORY_TUTOR_ATTR == category:
        activity.tutor_count = int(count)
        activity.save()
        _gen_work_attr(activity, category, work_info_list, count=int(count))
    return dict(c=SUCCESS[0], m=SUCCESS[1], d=TRUE_STR)


def edit_work_attr_schema_bulk(account, activity_id, bulk_list):
    bulk = json.loads(bulk_list)
    for each in bulk:
        category = each['category']
        work_info_list = each['work_info_list']
        count = each['count']
        edit_work_attr(account, activity_id, category, work_info_list, count)
    return dict(c=SUCCESS[0], m=SUCCESS[1], d=TRUE_STR)


@transaction.atomic()
def edit_ranks(account, activity, ranks_list):
    # 检查权限
    if not is_devops(account) and not is_activity_owner(activity, account):
        raise BusinessException(ERR_USER_AUTH)

    ranks_list = json.loads(ranks_list)
    Ranks.objects.filter(activity=activity).update(del_flag=FLAG_YES)
    for item in ranks_list:
        rank_name = item['rank_name']
        all_allowed = item['all_allow']
        Ranks.objects.create(activity=activity, name=rank_name, all_allowed=all_allowed)
    return dict(c=SUCCESS[0], m=SUCCESS[1], d=[])


def detail_rule(activity):
    rule = Rule.objects.filter(activity=activity).first()
    if not rule:
        return dict(c=SUCCESS[0], m=SUCCESS[1], d={})
    d = json.loads(rule.content)
    code = str(d['code'])
    return dict(c=SUCCESS[0], m=SUCCESS[1], d={
        'code': code,
        'rule': d
    })


@transaction.atomic()
def edit_rule(account, activity, rule_json):
    # 检查权限
    if not is_devops(account) and not is_activity_owner(activity, account):
        raise BusinessException(ERR_USER_AUTH)

    Rule.objects.filter(activity=activity).update(del_flag=FLAG_YES)
    ''' 组长制
        {
            'code': 1,
            'judge_count': 6,
            'max': 100
        }
     OR 平均分制
        {
            'code': 2,
            'judge_count': 6,
            'max': 100
            'ignore_maxmin': 0
        }
    '''
    if rule_json:
        rule_obj = json.loads(rule_json)
        rule_code = rule_obj['code']
        Rule.objects.create(activity=activity, code=rule_code, content=rule_json)
    return dict(c=SUCCESS[0], m=SUCCESS[1], d=[])


def all_activity(stages, is_home, level, rows, page):
    qs = Activity.objects.filter(template_flag=FLAG_NO).order_by('-open_time', '-create_time')
    if str2bool(is_home):
        qs = qs.exclude(is_minor=TRUE_INT)
    if stages and stages.strip('').strip(','):
        stage_list = [int(x) for x in stages.strip('').strip(',').split(',')]
        if not all([s in const_stage().all() for s in stage_list]):
            raise BusinessException(ERR_INVALID_PHASE)
        qs = qs.filter(stage__in=stage_list)
    if level:
        qs = qs.filter(user__area__area_level=int(level))
    qs_values = qs.values(
        "id", "name", "upload_time", "archive_time", "create_time", "browse_count", "work_count",
        "banner__url", "user__area_id", "stage", "organizers", "banner__url",
        "is_top", "is_minor",
        "user__area_id", "user__area__area_name", "user__area__area_code", "user__area__area_level",
        "user__area__manage_direct", "user__area__parent__area_level", "user__area__parent__id",
        "user__area__parent__area_name",
    )
    paged_data, result = paging_by_page(list(qs_values), rows, page)
    for each in paged_data:
        result['items'].append(dict(activity_id=str(each['id']), activity_name=each['name'],
            browse_count=str(each['browse_count']), work_count=str(each['work_count']),
            stage=str(each['stage']), time_display=_time_display(u'赛事时间', each['upload_time'], each['archive_time']),
            organizer=each['organizers'], banner_url=get_image_url(each['banner__url']),
            year=str(each['create_time'].year),

            area_id=str(each['user__area_id']), area_code=each['user__area__area_code'], area_level=str(each['user__area__area_level']),
            area_name_simple=area_name(each['user__area_id'], full=False),   #each['user__area__area_name'],
            area_name_full=area_name(each['user__area_id']),
            area_is_direct=str(each['user__area__manage_direct']),
            area_direct_level=str(each['user__area__parent__area_level']) if each['user__area__parent__area_level'] else '',
            area_parent_id=str(each['user__area__parent__id']) if each['user__area__parent__id'] else '',
            area_parent_name_simple=each['user__area__parent__area_name'] if each['user__area__parent__area_name'] else '',

            is_top=str(each['is_top']), is_minor=str(each['is_minor']),
         ))

    return dict(c=SUCCESS[0], m=SUCCESS[1], d=result)


def detail_activity(account, activity, with_introduction):
    rule = Rule.objects.filter(activity=activity).first()
    # judger_max = rule.parse_rule().expert_count()

    banner_url = activity.banner.url if activity.banner else ""
    banner_name = activity.banner.name if activity.banner else ""
    if banner_url:
        banner_url = get_image_url(banner_url)
    banner_id = str(activity.banner.id) if activity.banner else ""
    attachment_url = activity.attachment.url if activity.attachment else ""
    attachment_name = activity.attachment.name if activity.attachment else ""
    if attachment_url:
        attachment_url = get_image_url(attachment_url)
    attachment_id = str(activity.attachment.id) if activity.attachment else ""

    identity = list()
    if not account.is_anonymous():
        if is_activity_owner(activity, account):
            # 归档阶段不允许使用“我要管理”
            if activity.stage != ACTIVITY_STAGE_ARCHIVE:
                identity.append(str(WHO.CREATOR))
        if Role.objects.filter(activity=activity, user__account__id=account.id).exclude(parent_role=None).exists():
            # 仅上传阶段可用“我要参赛”
            if activity.stage == ACTIVITY_STAGE_UPLOAD:
                identity.append(str(WHO.PARTICIPANT))
        if (ExpertActivity.objects.filter(activity=activity, expert__account=account, expert__del_flag=FALSE_INT).exists() \
                and activity.stage == ACTIVITY_STAGE_REVIEW) \
                    or \
           (SubJudgeExpert.objects.filter(subjudge__activity=activity, expert__account=account, expert__del_flag=FALSE_INT).exists()
                and activity.stage == ACTIVITY_STAGE_UPLOAD):   # 只有整体在上传阶段子级评审专家才可以看见子级评审
                identity.append(str(WHO.EXPERT))
    identity_str = ','.join(identity) if len(identity) > 0 else ''

    is_upload_winner_info = Winner.objects.filter(activity=activity).exists()

    dict_resp = {
        "activity_id": str(activity.id),
        "is_template": str(activity.template_flag),
        "name": activity.name,
        "stage": str(activity.stage),
        "organizers": activity.organizers,
        "participator": activity.participator,

        "banner_url": banner_url,
        "banner_id": str(banner_id),
        'banner_name': banner_name,

        "attachment_url": attachment_url,
        "attachment_id": str(attachment_id),
        "attachment_name": attachment_name,

        "open_time": datetime_f_str(activity.open_time) if activity.open_time else "",
        "upload_time": datetime_f_str(activity.upload_time) if activity.upload_time else "",
        "group_time": datetime_f_str(activity.group_time) if activity.group_time else "",
        "review_time": datetime_f_str(activity.review_time) if activity.review_time else "",
        "public_time": datetime_f_str(activity.public_time) if activity.public_time else "",
        "archive_time": datetime_f_str(activity.archive_time) if activity.archive_time else "",

        "introduction": activity.introduction if str(with_introduction) == TRUE_STR else '',
        "author_count": str(activity.author_count),
        "tutor_count": str(activity.tutor_count),

        "base_info_value": activity.base_info_value,

        "copyright": activity.copyright or '',
        "is_top": str(activity.is_top),
        "is_minor": str(activity.is_minor),
        "genre": str(activity.genre),
        "browse_count": str(activity.browse_count),
        "work_count": str(activity.work_count),

        'identity': identity_str,
        'rule': json.loads(rule.content) if rule else '',

        # 是否上传了获奖名单
        'is_upload_winner_info': bool2str(is_upload_winner_info),
        'year': str(activity.create_time.year),
    }
    return {"c": SUCCESS[0], "m": SUCCESS[1],  "d": [dict_resp,]}


def detail_work_attr(activity):
    result = list()
    work_attrs = WorkAttr.objects.filter(activity=activity, del_flag=DEL_FLAG_NO)
    if work_attrs:
        qs = work_attrs.values("id", "category", 'group_sn', "sn", "name", "type", "values", "mandatory")
        data = {}
        for each in qs:
            if each["category"] not in data:
                data[each["category"]] = []
            data[each["category"]].append(
                        {"id": str(each["id"]), "group_sn": str(each["group_sn"]), "name": each["name"],
                         "type": str(each["type"]), "values": each["values"], "mandatory": str(each["mandatory"]),
                        })
        for d in data:
            result.append({"category": d, "item": data[d]})
    return dict(c=SUCCESS[0], m=SUCCESS[1], d=result)


def detail_work_attr_schema(activity):
    result = list()
    work_attrs = WorkAttr.objects.filter(activity=activity, del_flag=DEL_FLAG_NO, group_sn=1)
    qs = work_attrs.values("id", "category", 'group_sn', "sn", "name", "type", "values", "mandatory")
    data = [
        {'category': '1', 'count': '0', 'items': list()},
        {'category': '2', 'count': '0', 'items': list()},
        {'category': '3', 'count': '0', 'items': list()}
    ]
    for each in qs:
        if each["category"] == 2:
            count = activity.author_count
        elif each["category"] == 3:
            count = activity.tutor_count
        else:
            count = 1
        data[each["category"]-1]['count'] = count
        data[each["category"]-1]['items'].append(
                    {"id": str(each["id"]), "name": each["name"],
                     "type": str(each["type"]), "values": each["values"], "mandatory": str(each["mandatory"]),
                    })
    for each in data:
        result.append({"category": each['category'], "count": each['count'], "item": each['items']})
    return dict(c=SUCCESS[0], m=SUCCESS[1], d=result)


def detail_ranks(activity, account):
    qs_values = Ranks.objects.filter(activity=activity)\
        .annotate(rank_id=F('id')).values("rank_id", "name", "all_allowed")
    if not is_devops(account) and not is_activity_owner(activity, account):
        qs_values = qs_values.filter(all_allowed=TRUE_INT)
    return dict(c=SUCCESS[0], m=SUCCESS[1], d=list(qs_values))


def brother_role(role, exclude_myself=False):
    # 找到兄弟Role
    qs = Role.objects.filter(activity=role.activity, user__area=role.user.area)
    if exclude_myself:
        qs = qs.exclude(id=role.id)
    return list(qs)


def list_role_in_activity(
        account, activity, user, role, keyword=None, area=None, direct_area=None, all=False,
        rows=10, page=1, with_work_stats=True):

    brother = brother_role(role)
    qs = Role.objects.filter(activity=activity)
    if not all:
        qs = qs.filter(parent_role__in=brother)
    else:
        if not is_activity_owner(activity, account):
            raise BusinessException(ERR_USER_AUTH)

    if keyword:  # 支持账号和姓名的模糊查找
        qs = qs.filter(Q(user__account__name__contains=keyword) | Q(user__account__username__contains=keyword))
    if area:
        qs = qs.filter(user__area=area)
    if direct_area:
        qs = qs.filter(user__area__manage_direct=TRUE_INT, user__area__parent=direct_area)

    all = list(qs)
    paged_data, result = paging_by_page(all, rows, page)

    for each in paged_data:
        c_total = c_apprv = c_non_apprv = 0
        if with_work_stats:
            # 获取该用户上传的作品数量
            qs = Work.objects.filter(del_flag=FALSE_INT, uploader=each.user)
            c_total = qs.count()
            c_apprv = qs.filter().count()
            c_non_apprv = qs.filter().count()
        d = {
            "role_id": str(each.id),
            "user_id": str(each.user.id),
            "account_id": str(each.user.account.id),
            "username": str(each.user.account.username),
            "name": str(each.user.account.name),
            "sex": str(each.user.account.sex),
            "area_id": str(each.user.area.id),
            "area_name_full": area_name(str(each.user.area.id)),
            "area_name_simple": area_name(str(each.user.area.id), full=False),
            "manage_direct": str(each.user.area.manage_direct),
            "user_work_number": str(c_total),
            "user_approve_number": str(c_apprv),
            "user_noapprove_number": str(c_non_apprv),
            'max_work': str(each.max_work),
            "is_data_confirm": str(each.user.account.is_data_confirm)
        }
        result['items'].append(d)
    result['id_all'] = [str(each.id) for each in all]
    return result


def add_activity_role_registered(account, activity, user, account_id_list):
    import applications.user.services as u_service
    stu_accounts = Account.objects.filter(del_flag=FALSE_INT, id__in=[int(a) for a in account_id_list])
    user_id_list = []
    for stu in stu_accounts:
        result_add_user = u_service.add_user(account, stu.username, stu.name, stu.sex, user.area.id, None, None)
        user_id_list.append(result_add_user['d']['user_id'])
    add_activity_role_exist(account, activity, user, user_id_list)
    return dict(c=SUCCESS[0], m=SUCCESS[1])


def add_activity_role_new(account, activity, user, mobile, name, sex, area, direct_area, institution, max_work):
    import applications.user.services as u_service

    if user.area.manage_direct == TRUE_INT or user.area.area_level == 0b00001:
        # 校或机构管理员、直属组织的管理员创建新用户时，只能创建末端用户
        direct_area = user.area
    else:
        if not direct_area:
            # Area必须是当前用户的子级
            if user.area != area.parent:
                raise BusinessException(ERR_AREA_ERROR)

    direct_area_id = direct_area.id if direct_area else ''
    area_id = area.id if area else ''
    create_result = u_service.add_user(account, mobile, name, sex, area_id, direct_area_id, institution)
    user_id = int(create_result['d']['user_id'])
    new_user = User.objects.filter(del_flag=FALSE_INT, id=int(user_id)).first()
    my_role = Role.objects.filter(activity=activity, user=user).first()
    if max_work:
        obj, is_create = Role.objects.get_or_create(activity=activity, user=new_user, parent_role=my_role, max_work=int(max_work))
    else:
        obj, is_create = Role.objects.get_or_create(activity=activity, user=new_user, parent_role=my_role)
    if not is_create:
        logger.warn('role %s is already exist, no need create again' % obj.id)
    return dict(c=SUCCESS[0], m=SUCCESS[1], d={'role_id': str(obj.id), 'user_id': str(user_id)})


@transaction.atomic()
def import_activity_role(account, user, role, activity, excel):
    ext = (excel.name[excel.name.rfind('.') + 1:]).lower()   # 后缀(不含.)
    if ext != 'xlsx':
        raise BusinessException(u'仅支持xlsx格式EXCEL')

    excel_store_name = 'import_role_%s_%s_%s.%s' % (
        datetime.datetime.now().strftime('%y%m%d%H%M%S'), account.id, activity.id, ext)

    sheet = load_workbook(filename=BytesIO(excel.read())).active
    content = list(sheet.rows)
    content.pop(0)  # skip TIPS
    content.pop(0)  # skip HEAD
    for i, row in enumerate(content):
        # 用户名 / 姓名 / 性别 / 地区 / 直属机构
        username = str(row[0].value)
        name = str(row[1].value)
        sex = str(row[2].value)
        cur_area_lv = user.area.area_level
        new_area = None
        new_direct_area = None
        new_institution = None
        # 国/省/市/区 导入用户
        if cur_area_lv in AREA_DEF:
            area_input = str(row[3].value)
            if area_input == AREA_DEF[cur_area_lv][1]:
                new_institution = str(row[4].value)
                new_direct_area = user.area
            else:
                new_area = Area.objects.filter(del_flag=FALSE_INT, area_name=area_input, parent=user.area).first()
        # 学校或机构导入用户
        else:
            new_direct_area = user.area
        add_result = add_activity_role_new(
                    account, activity, user, username, name, sex, new_area, new_direct_area, new_institution, None)

    path_r = os.path.join(settings.MEDIA_KEY_EXCEL, excel_store_name)
    path_a = os.path.join(settings.BASE_DIR, path_r)
    save_file(excel, path_a)  # 将用户导入的EXCEL保存下来

    return TRUE_STR


def export_activity_expert(account, activity, user, role, expert_list):
    wb = Workbook()
    sheet = wb.active

    sheet.cell(column=1, row=1, value=u'用户名（手机号）')
    sheet.cell(column=2, row=1, value=u'姓名')
    sheet.cell(column=3, row=1, value=u'性别')
    sheet.cell(column=4, row=1, value=u'地区')
    sheet.cell(column=5, row=1, value=u'岗位或职务')

    for i, each in enumerate(expert_list):
        sheet.cell(column=1, row=1 + 1 + i, value=each.account.username)
        sheet.cell(column=2, row=1 + 1 + i, value=each.account.name)
        sheet.cell(column=3, row=1 + 1 + i, value=each.account.sex)
        sheet.cell(column=4, row=1 + 1 + i, value=each.area.area_fullname)
        sheet.cell(column=5, row=1 + 1 + i, value=each.position)

    return wb


def gen_download_account_template(account, activity, user, role, with_area=True, tips=None):
    from applications.user.services import list_sub_area
    wb = Workbook()
    sheet = wb.active
    area_lv = user.area.area_level   # 当前执行人所在区域

    HEAD_R_IDX = 2

    sheet.cell(column=1, row=HEAD_R_IDX, value=u'手机号*')
    sheet.cell(column=2, row=HEAD_R_IDX, value=u'姓名*')
    sheet.cell(column=3, row=HEAD_R_IDX, value=u'性别*')
    end_col = 3
    if with_area and area_lv in AREA_DEF:
        end_col = 4
        sheet.cell(column=end_col, row=HEAD_R_IDX, value=(AREA_DEF[area_lv][0]+'*'))
        enum = [each['name'] for each in list_sub_area(None, None, user.area.id, None, None, None)['d']]
        if AREA_DEF[area_lv][2]:
            end_col = 5
            sheet.cell(column=end_col, row=HEAD_R_IDX, value=u'直属机构')
            # enum.append(area_def[area_lv][1])
        enum_str = '"%s"' % (','.join(enum))
        dv = DataValidation(type="list", formula1=enum_str, allow_blank=False)
        dv.error = u'请选择区域而不要自己填写'
        dv.prompt = u'请选择区域'
        sheet.add_data_validation(dv)
        area_sel_col = get_column_letter(4)
        for r in xrange(HEAD_R_IDX+1, EXCEL_ROW_MAX):
            dv.add(sheet['%s%s'% (area_sel_col, r)])

    tips_show = tips or u'打*必填，如果选择“XX直属”，请填写直属机构'
    sheet.cell(column=1, row=1, value=tips_show)
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    return wb, HEAD_R_IDX, end_col


def download_expert_template(account, activity, user, role):
    wb, head_row, end_col = gen_download_account_template(account, activity, user, role,
                    with_area=False, tips=u'请准确填写地区，例如：湖北省武汉市洪山区')
    wb.active.cell(column=end_col + 1, row=head_row, value=u'地区')
    wb.active.cell(column=end_col + 2, row=head_row, value=u'岗位或职务')
    # fill sample data
    # TODO
    return wb


def download_activity_role_template(account, activity, user, role):
    wb, head_row, end_col = gen_download_account_template(account, activity, user, role)
    # fill sample data
    # TODO
    return wb


def export_activity_role(account, activity, user, role, role_list):
    wb = Workbook()
    sheet = wb.active

    sheet.cell(column=1, row=1, value=u'用户名（手机号）')
    sheet.cell(column=2, row=1, value=u'姓名')
    sheet.cell(column=3, row=1, value=u'性别')
    sheet.cell(column=4, row=1, value=u'地区')

    for i, each in enumerate(role_list):
        sheet.cell(column=1, row=1 + 1 + i, value=each.user.account.username)
        sheet.cell(column=2, row=1 + 1 + i, value=each.user.account.name)
        sheet.cell(column=3, row=1 + 1 + i, value=each.user.account.sex)
        sheet.cell(column=4, row=1 + 1 + i, value=each.user.area.area_fullname)

    return wb


def add_activity_role_exist(account, activity, user, user_id_list, keyword=None, all=None):
    # 检查操作权限
    role = Role.objects.filter(activity=activity, user=user).first()
    if (not role) or (user.area.area_level==0):
        raise BusinessException(ERR_USER_AUTH)

    choose_user_id_list = []
    if user_id_list:
        choose_user_id_list = map(lambda x: int(x), user_id_list)

    # 新加的用户必须在可添加列表内
    # available = [int(each['user_id'])
    #              for each in available_user_add_role(account, user, role, activity, keyword, 1, BIG)['items']]
    # illegal = set(choose_user_id_list) - set(available)
    # if len(illegal) > 0:
    #     logger.warn('users %s is not available to add into activity.' % (', '.join([str(each) for each in illegal])))
    # choose_user_id_list = set(choose_user_id_list) & set(available)

    # 当一个用户出现在多区域时，请仅选择一个区域的用户
    choose_user_list = [User.objects.filter(del_flag=FALSE_INT, id=u).first() for u in choose_user_id_list if u]
    choose_account_list = list()
    dup_err_list = list()
    for each in choose_user_list:
        if each.account in choose_account_list:
            dup_err_list.append(each.account)
        choose_account_list.append(each.account)
    if len(dup_err_list) > 0:
        raise BusinessException(ERR_USER_ADD_IN_ACTI_DUP % ','.join([a.name for a in dup_err_list]))

    # # 单一用户不能被在一次赛事中多次加入（评审除外）
    # qs_check_exist = Role.objects.filter(activity=activity, user__account__in=choose_account_list)
    # if qs_check_exist.exists():
    #     raise BusinessException(ERR_USER_ADD_ALREADY_IN_ACTI % ','.join([r.user.account.name for r in qs_check_exist]))

    for each in choose_user_list:
        Role.objects.get_or_create(activity=activity, user=each, parent_role=role)
    return dict(c=SUCCESS[0], m=SUCCESS[1], d=[])


def available_user_add_role(account, user, role, activity, keyword, page, rows):
    import applications.user.services as u_service
    user_lib = u_service.list_user(account, user.id, None, keyword)
    all = [{'user_id': str(each['id']), 'name': each['name'], 'mobile': each['mobile'],
            'area_name': area_name(int(each['area_id']), full=False, pov=user.area),
            }
           for each in user_lib['d']]
    role_in_acti = list_role_in_activity(account, activity, user, role, rows=BIG, page=1)
    user_in_acti = [str(r['user_id']) for r in role_in_acti['items']]
    filtered = list()
    for each in all:
        if each['user_id'] not in user_in_acti:
            filtered.append(each)
    paged_data, result = paging_by_page(filtered, rows, page)
    result['items'] = paged_data
    return result


def available_registered_add_role(account, user, role, activity, keyword, page, rows):
    # 所有注册到当前用户所在区域的末端用户
    qs = Account.objects.filter(del_flag=FALSE_INT, region=user.area.region).exclude(region__isnull=True)
    if keyword:
        qs = qs.filter(Q(name__contains=keyword) | Q(mobile__contains=keyword))

    role_in_acti = list_role_in_activity(account, activity, user, role, rows=BIG, page=1)
    account_in_acti = [int(r['account_id']) for r in role_in_acti['items']]
    filtered = list()
    for each in qs:
        if each.id not in account_in_acti:
            filtered.append(each)
    paged_data, result = paging_by_page(filtered, rows, page)
    items = list()
    for each in paged_data:
        items.append({'account_id': str(each.id), 'name': each.name, 'mobile': each.mobile})
    result['items'] = items
    return result


@transaction.atomic()
def edit_activity_role(activity, user, role, modify_role, new_username, new_name, new_sex, new_max):
    from applications.user.services import mod_user

    new_username = new_username or modify_role.user.account.username
    new_name = new_name or modify_role.user.account.name
    new_sex = new_sex or modify_role.user.account.sex
    old_area_name = modify_role.user.area.area_name  # 修改用户是不允许修改所在区域

    mod_user(user.account, modify_role.user.id, new_username, new_name, new_sex, modify_role.user.area.id, None, None)

    if new_max:
        if modify_role.approve_work > int(new_max):
            raise BusinessException(ERR_ROLE_MAX_EXIST_EXCEED)
        modify_role.max_work = new_max
        modify_role.save()
    return TRUE_STR


def detail_activity_role(activity, user, role, role_retrieved):
    area = role_retrieved.user.area
    area_info = get_areadetail_by_id(area.id)
    return {
        'user_id': str(role_retrieved.user.id),
        'role_id': str(role_retrieved.id),
        'username': role_retrieved.user.account.username,
        'name': role_retrieved.user.account.name,
        'sex': role_retrieved.user.account.sex,
        'mobile': role_retrieved.user.account.mobile,
        'area_id': str(area_info['area_id']),
        'area_province': area_info['province'],
        'area_city': area_info['city'],
        'area_district': area_info['country'],
        'area_school_or_institution': area_info['institution'],
        'is_direct': str(area_info['manage_direct']),
        'direct_level': str(role_retrieved.user.area.area_level),
        'max_work': str(role_retrieved.max_work),
        'approve_work': str(role_retrieved.approve_work),
    }


def remove_activity_role(account, activity, role, roles_list):
    qs_del_roles = Role.objects.filter(activity=activity, id__in=roles_list)
    if qs_del_roles.exclude(parent_role=role).exists():
        raise BusinessException(ERR_USER_AUTH)

    role_work = defaultdict(lambda: list())
    for w in Work.objects.filter(activity=activity):
        role_work[w.uploader.id].append(w)

    fail_count = 0
    succ_count = 0
    for each in qs_del_roles:
        if len(role_work[each.user.id]) > 0:
            fail_count += 1
        else:
            each.del_flag = TRUE_INT
            each.save()
            succ_count += 1

    msg1 = (u'%s个用户从活动中移除成功. ' % succ_count) if succ_count > 0 else ''
    msg2 = (u'%s个用户从活动中移除失败，可能该用户已经上传了作品' % fail_count) if fail_count > 0 else ''
    return dict(c=SUCCESS[0], m=SUCCESS[1], d=msg1 + msg2)


def add_activity_expert_exist(account, activity, user, expert_id_list):
    if activity.user != user:
        raise BusinessException(ERR_USER_AUTH)
    expert_id_in_activity = [ea.expert.id for ea in ExpertActivity.objects.filter(activity=activity, expert__del_flag=FALSE_INT)]

    expert_exist = set(expert_id_list) & set(expert_id_in_activity)
    if len(expert_exist) > 0:
        logger.info('following expert is already in activity, no need to join again: %s' % ','.join(
            [str(each) for each in expert_exist]
        ))

    expert_id_can_add = set(expert_id_list) - set(expert_id_in_activity)
    expert_can_add = Expert.objects.filter(del_flag=FALSE_INT, id__in=expert_id_can_add)
    create_obj_list = list()
    for each in expert_can_add:
        create_obj_list.append(ExpertActivity(activity=activity, expert=each))
    ExpertActivity.objects.bulk_create(create_obj_list)
    return dict(c=SUCCESS[0], m=SUCCESS[1], d=[])


def add_activity_expert_new(account, activity, user, mobile, name, sex, area, direct_area, institution, position):
    from applications.expert.services import add_expert_user
    area_id = area.id if area else None
    direct_area_id = direct_area.id if direct_area else None
    result = add_expert_user(account, mobile, name, sex, institution, None, position, None, area_id, direct_area_id)
    expert_id = result['d']['expert_id']
    add_activity_expert_exist(account, activity, user, [int(expert_id), ])
    return {'expert_id': expert_id}


@transaction.atomic()
def import_activity_expert(account, user, role, activity, excel):
    from applications.expert.services import add_expert_user
    ext = (excel.name[excel.name.rfind('.') + 1:]).lower()   # 后缀(不含.)
    if ext != 'xlsx':
        raise BusinessException(u'仅支持xlsx格式EXCEL')

    excel_store_name = 'import_expert_%s_%s_%s.%s' % (
        datetime.datetime.now().strftime('%y%m%d%H%M%S'), account.id, activity.id, ext)

    sheet = load_workbook(filename=BytesIO(excel.read())).active
    content = list(sheet.rows)
    content.pop(0)  # skip TIPS
    content.pop(0)  # skip HEAD
    for i, row in enumerate(content):
        # 用户名 / 姓名 / 性别 / 地区 / 岗位或职务
        username = str(row[0].value)
        name = str(row[1].value)
        sex = str(row[2].value)

        # 猜测导入专家的区域
        guess_result, area, institution = GuessArea().guess(str(row[3].value))
        para_institution = None
        para_area_id = None
        para_direct_area_id = None
        if guess_result == GuessArea.RESULT_STATUS_FIND:
            para_area_id = area.id
        if guess_result == GuessArea.RESULT_STATUS_FIND_DIRECT:
            para_direct_area_id = area.id
            para_institution = institution
        if guess_result == GuessArea.RESULT_STATUS_NOT_FIND:
            para_area_id = china().id

        position = str(row[4].value)
        logger.info('call service to add user. username: %s, name:%s, sex:%s, area_id:%s, d_area_id:%s, insti:%s, posi:%s' %
                    (username, name, sex, para_area_id, para_direct_area_id, para_institution, position))
        add_result = add_expert_user(account, username, name, sex, para_institution, None, position, None,
                        para_area_id, para_direct_area_id)
        if add_result['c'] != 0:
            raise BusinessException(ERR_ADD_USER)
        expert_id = add_result['d']['expert_id']

        add_activity_expert_exist(account, activity, user, [int(expert_id), ])

    path_r = os.path.join(settings.MEDIA_KEY_EXCEL, excel_store_name)
    path_a = os.path.join(settings.BASE_DIR, path_r)
    save_file(excel, path_a)  # 将用户导入的EXCEL保存下来

    return TRUE_STR


def list_activity_expert(account, activity, user, keyword, rows, page):
    qualified_id_list = list()
    qs = ExpertActivity.objects.filter(activity=activity, expert__del_flag=FALSE_INT).order_by('-update_time')
    if keyword:
        qs = qs.filter(Q(expert__name__contains=keyword) |
                       Q(expert__position__contains=keyword) |
                       Q(expert__account__username__contains=keyword))
    # if direct_level:
    #     qs = qs.filter(expert__area__manage_direct=TRUE_INT, expert__area__area_level=direct_level)
    # if area_id:
    #     qs = qs.filter(expert__area_id=int(area_id))
    # if institution:
    #     qs = qs.filter(expert__institution__contains=institution)

    filtered_list = list(qs)
    for each in filtered_list:
        qualified_id_list.append(str(each.expert.id))

    paged_data, result = paging_by_page(filtered_list, rows, page)
    for each in paged_data:
        result['items'].append(expert(each.expert))

    result['id_all'] = ','.join(qualified_id_list).strip(',')
    return result


def available_add_activity_expert(account, activity, user, keyword, rows, page):
    expert_in_activity = list_activity_expert(account, activity, user, None, BIG, 1)['items']
    expert_id_list_in_activity = [each['expert_id'] for each in expert_in_activity]
    all = Expert.objects.filter(del_flag=FALSE_INT)
    if keyword:
        all = all.filter(Q(account__username__contains=keyword) | Q(account__name__contains=keyword))
    all = all.exclude(id__in=expert_id_list_in_activity)
    paged_data, result = paging_by_page(list(all), rows, page)
    for each in paged_data:
        result['items'].append(expert(each))
    return dict(c=SUCCESS[0], m=SUCCESS[1], d=result)


def remove_activity_expert(account, activity, user, expert_id_list):
    if activity.user != user:
        raise BusinessException(ERR_USER_AUTH)
    expert_remove_list = Expert.objects.filter(del_flag=FALSE_INT, id__in=expert_id_list)
    expert_in_team = TeamExpert.objects.filter(del_flag=FALSE_INT, expert__in=expert_remove_list, team__activity=activity)
    fail_list = list()
    if expert_in_team.exists():
        dup_name = ','.join([each.expert.account.name for each in expert_in_team])
        fail_list.append(each.expert)
        logger.warn(ERR_CANT_DEL_EXPERT_INTEAM[1] % dup_name)
    expert_can_remove = set(expert_remove_list) - set(fail_list)
    ExpertActivity.objects.filter(activity=activity, expert__del_flag=FALSE_INT, expert__in=expert_can_remove).update(del_flag=TRUE_INT)

    msg1 = (u'%d个专家成功从活动中移除' % len(expert_can_remove)) if len(expert_can_remove) > 0 else ''
    msg2 = (u'%d个专家已经在评审组中，因而无法从活动中移除' % len(fail_list)) if len(fail_list) > 0 else ''
    return dict(c=SUCCESS[0], m=SUCCESS[1], d=msg1 + msg2)


def _copy_activity(source, template_flag=FLAG_YES, name='', user=None, stage=ACTIVITY_STAGE_EDIT):
    destiny = copy.deepcopy(source)
    destiny.pk = None
    destiny.attachment_id = None
    destiny.template_flag = template_flag
    destiny.name = name
    destiny.user = user
    destiny.stage = stage
    if template_flag == FLAG_NO:
        destiny.template_based = source.id
    destiny.save()

    # 活动对应的WorkAttr
    work_attr = WorkAttr.objects.filter(activity=source).all()
    for item in list(work_attr):
        new_item = copy.deepcopy(item)
        new_item.pk = None
        new_item.activity = destiny
        new_item.save()

    # 活动对应的评审规则
    rule = Rule.objects.filter(activity=source).all()
    for item in list(rule):
        new_item = copy.deepcopy(item)
        new_item.pk = None
        new_item.activity = destiny
        new_item.save()

    # 活动对应的Ranks
    ranks = Ranks.objects.filter(activity=source).all()
    for item in list(ranks):
        new_item = copy.deepcopy(item)
        new_item.pk = None
        new_item.activity = destiny
        new_item.save()

    return destiny


def add_template(account, activity_id, name):
    if not is_devops(account):
        raise BusinessException(ERR_USER_AUTH)
    if not activity_id:
        # 创建新模板
        if not name:
            raise BusinessException(ERR_DATA_NOT_FOUND)
        if Activity.objects.filter(name=name, template_flag=FLAG_YES).exists():
            raise BusinessException(ERR_ACTIVITY_NAME_CONFLICT)
        activity = Activity.objects.create(name=name, template_flag=FLAG_YES)
    else:
        # 通过活动添加模板
        exist_activity = Activity.objects.filter(id=int(activity_id)).first()
        if not exist_activity:
            raise BusinessException(ERR_DATA_NOT_FOUND)
        activity = _copy_activity(exist_activity)

    return dict(c=SUCCESS[0], m=SUCCESS[1], d=[activity.id])


def list_template(account):
    templates = Activity.objects.filter(template_flag=FLAG_YES).values("id", "name")
    data = []
    for each in list(templates):
        count = Activity.objects.filter(template_flag=FALSE_INT, template_based=int(each['id'])).count()
        dict_tmp = {'template_id': str(each['id']),
                    'count': str(count),
                    'template_name': each['name']}
        data.append(dict_tmp)
    return dict(c=SUCCESS[0], m=SUCCESS[1], d=data)


@transaction.atomic()
def del_activity(activity):
    activity.del_flag=FLAG_YES
    activity.save()
    WorkAttr.objects.filter(activity=activity, del_flag=FLAG_NO).update(del_flag=FLAG_YES)
    Ranks.objects.filter(activity=activity, del_flag=FLAG_NO).update(del_flag=FLAG_YES)
    Rule.objects.filter(activity=activity, del_flag=FLAG_NO).update(del_flag=FLAG_YES)
    # 活动中的用户、专家、作品不删除
    return {'activity_id': str(activity.id)}


@transaction.atomic()
def add_activity(account, area, template):
    if not template:
        raise BusinessException(ERR_TEMPLATE_NOT_EXIST)
    # 验证该用户有该区域的赛事创建权限
    if not AccountRight.objects.filter(del_flag=FLAG_NO, account=account, area=area).exists():
        raise BusinessException(ERR_USER_AUTH)

    # 为创建者设定User对象，并创建Role
    user, _ = User.objects.get_or_create(del_flag=FALSE_INT, account=account, area=area, parent_account=None)
    activity = _copy_activity(template, template_flag=FLAG_NO, user=user)
    Role.objects.create(activity=activity, user=user, parent_role=None)

    return dict(c=SUCCESS[0], m=SUCCESS[1], d={'activity_id': str(activity.id)})


def import_winner(user, activity_id, file):
    # 导入获奖名单
    # 检查当前用户是否活动创建者
    activity = Activity.objects.filter(id=activity_id, del_flag=FALSE_INT, user__account=user)
    if not activity:
        raise BusinessException(ERR_USER_AUTH)

    # 将文件写入临时目录
    file_path = gen_path()
    with open(file_path, 'wb+') as destination:
        for chunk in file.chunks():
            destination.write(chunk)

    wb = load_workbook(file_path, read_only=True)
    ws = wb.get_active_sheet()

    # 获取数据，为了便于后期可能对数据进行处理，将数据获取和记录分开。
    randid = uuid.uuid4().hex
    file_data_list = list()
    for each_row in ws.rows:
        col_value_list = list()
        for cell in each_row:
            col_value_list.append(excel_cell2str(cell.value))
        file_data_list.append(col_value_list)

    # 将原记录失效
    Winner.objects.filter(activity_id=activity_id, del_flag=FALSE_INT).update(del_flag=TRUE_INT)
    sn = 0
    for each_datarow in file_data_list:
        sn += 1
        new_winner = Winner()
        new_winner.activity_id = activity_id
        new_winner.uuid = randid
        new_winner.sn = sn
        new_winner.attr = json.dumps(each_datarow, ensure_ascii=False)
        new_winner.save()

    return dict(c=SUCCESS[0], m=SUCCESS[1], d=[])


def list_winner(user, activity_id, rows, page, last_id):
    # 返回获奖名单的列表，注意此列表分页与其它页面不一样，第一行为表头，每次返回都要带回表头。
    result = dict()
    winners = Winner.objects.filter(activity_id=activity_id, del_flag=FALSE_INT).order_by("sn")
    if not len(winners):
        result["max_page"] = 1
        result["total"] = 0
        result['winner_list'] = []
        return dict(c=SUCCESS[0], m=SUCCESS[1], d=result)

    winner_head = [winners[0]]
    winner_data = winners[1:]

    # 分页
    cnt = len(winner_data)
    num_pages = 1
    if rows and page:
        num_pages, cur_start, cur_end = get_pages(cnt, page, rows)
        winner_data = winner_data[cur_start:cur_end]
    elif rows:
        cur_start = get_lastid_index(winner_data, last_id)
        winner_data = winner_data[cur_start:cur_start + int(rows)]
        pass

    result["max_page"] = num_pages
    result["total"] = cnt
    result["page"] = page

    # winner_detail = winner_head | winner_data

    winner_list = list()
    for each_winner in winner_head:
        winner_dict = dict()
        winner_dict['id'] = each_winner.id
        winner_dict['sn'] = each_winner.sn
        winner_dict['attr'] = json.loads(each_winner.attr)
        winner_list.append(winner_dict)

    for each_winner in winner_data:
        winner_dict = dict()
        winner_dict['id'] = each_winner.id
        winner_dict['sn'] = each_winner.sn
        winner_dict['attr'] = json.loads(each_winner.attr)
        winner_list.append(winner_dict)
    result['winner_list'] = winner_list
    return dict(c=SUCCESS[0], m=SUCCESS[1], d=result)


def tag_alias(account, activity, tag):
    find = Alias.objects.filter(activity=activity, attr=tag).first()
    return {'alias': find.alias if find else ''}

