#!/usr/bin/env python
# coding=utf-8

import json
import copy
import logging
import os
from io import BytesIO
import datetime

from django.db import transaction
from django.db.models import Q, Count, F
from openpyxl import Workbook, load_workbook

from applications.activity.guess import GuessArea
from applications.activity.models import Role
from applications.activity.share import is_activity_owner, china
from applications.expert.models import Expert
from applications.subjudge.share import struct_work_subjudge, struct_work_subjudge_leader, struct_subjudge_expert
from applications.user.models import User, Area
from applications.work.models import Work
from applications.work.share import works, struct_work, works_manager_can_see, get_user
from utils.const_def import *
from utils.const_err import *
from utils.public_fun import str_p_datetime, datetime_f_str, paging_by_page, seq2list, is_1
from utils.file_fun import get_image_url
from utils.store_file import save_file
from utils.utils_except import BusinessException
from utils.utils_type import str2bool, bool2str
from applications.common.services import area_name, area_by_pov
from applications.subjudge.models import SubJudgeExpert, SubJudge, SubJudgeTeamExpert, SubJudgeRank, SubJudgeRule, SubJudgeTeamWork, SubJudgeTeam, SubJudgeScore

logger = logging.getLogger(__name__)


def works_4_expert(subjudge, subjg_expert, sub_activity=None, phase=None, project=None,
                   subject=None, keyword=None, subjg_team=None, as_leader=False):
    qs = works(subjudge.activity, sub_activity=sub_activity, paragraph=phase,
               paragraph_project=project, subject=subject, keyword=keyword)
    my_team_qs = SubJudgeTeamExpert.objects.filter(subjudge=subjudge, expert=subjg_expert.expert, expert__del_flag=FALSE_INT)
    if subjg_team:
        my_team_qs = my_team_qs.filter(subjudge_team=subjg_team)
    if as_leader:
        my_team_qs = my_team_qs.filter(is_leader=TRUE_INT)
    my_team = my_team_qs.values_list('subjudge_team__id', flat=True)
    my_work = SubJudgeTeamWork.objects.filter(subjudge=subjudge, subjudge_team__id__in=my_team).values_list('work__id', flat=True)
    qs = qs.filter(id__in=my_work)
    return qs


def _work_subjudge_info(subjudge, subjg_expert, work):
    t_w = SubJudgeTeamWork.objects.filter(subjudge=subjudge, work=work).first()
    t_e = SubJudgeTeamExpert.objects.filter(subjudge_team=t_w.subjudge_team, expert=subjg_expert.expert, expert__del_flag=FALSE_INT).first()
    subjudge_score = SubJudgeScore.objects.filter(subjudge_team_work=t_w, subjudge_team_expert=t_e).first()
    is_unfinish = False
    if subjudge_score:
        if subjudge_score.status == FALSE_INT:
            is_unfinish = True
            j_status = WORK_SHOW_STATUS_NOT_SUBMIT
        else:
            j_status = WORK_SHOW_STATUS_SUBMITED
    else:
        is_unfinish = True
        j_status = WORK_SHOW_STATUS_NON_JUDGED
    return {'work': work, 'is_unfinish': is_unfinish, 'judge_status': j_status, 'score': subjudge_score, 'team': t_w.subjudge_team}


def control(account, subjudge, status):
    subjudge.status = TRUE_INT if str(status) == TRUE_STR else FALSE_INT
    subjudge.save()


def list_work_for_expert(account, subjudge, subjg_expert, sub_activity, phase, project,
            subject, keyword, subjg_team, is_unfinish, page, rows):
    rule = SubJudgeRule.objects.filter(subjudge=subjudge).first()

    qs = works_4_expert(subjudge, subjg_expert, sub_activity=sub_activity, phase=phase,
                project=project, subject=subject, keyword=keyword, subjg_team=subjg_team)

    filter_list = list()
    for each in qs:
        work_info = _work_subjudge_info(subjudge, subjg_expert, each)
        if not work_info['is_unfinish'] and is_1(is_unfinish):
            continue
        filter_list.append({'work': work_info['work'], 'judge_status': work_info['judge_status'],
                            'score': work_info['score'], 'team': work_info['team']})
    paged_data, result = paging_by_page(filter_list, rows, page)
    for d in paged_data:
        result['items'].append(struct_work_subjudge(d['work'], d['judge_status'], d['score'], d['team'], rule))
    return result


def list_work_for_leader(account, subjudge, subjg_expert, sub_activity, phase, project,
            subject, keyword, subjg_team, is_unfinish, page, rows):
    rule = SubJudgeRule.objects.filter(subjudge=subjudge).first()
    judger_max = rule.parse_rule().expert_count()

    qs = works_4_expert(subjudge, subjg_expert, sub_activity=sub_activity, phase=phase,
                project=project, subject=subject, keyword=keyword, subjg_team=subjg_team, as_leader=True)

    qualified_id_list = list()
    filter_list = list()

    for each in qs:
        work_info = _work_subjudge_info(subjudge, subjg_expert, each)
        if not work_info['is_unfinish'] and is_1(is_unfinish):
            continue
        # 获取其它专家的评分
        expert_score = dict()
        score_qs = SubJudgeScore.objects.filter(
            subjudge_team_work__work=each, subjudge_team_expert__in=SubJudgeTeamExpert.objects.filter(subjudge_team=work_info['team']))
        for s in score_qs:
            expert_score[int(s.subjudge_team_expert.sn) - 1] = s
        expert_in_team_count = SubJudgeTeamExpert.objects.filter(subjudge_team=work_info['team']).count()
        filter_list.append({
            'work': work_info['work'], 'judge_status': work_info['judge_status'],
            'score': work_info['score'], 'team': work_info['team'],
            'expert_in_team_count': expert_in_team_count, 'expert_score': expert_score})
        qualified_id_list.append(str(each.id))
    paged_data, result = paging_by_page(filter_list, rows, page)
    for d in paged_data:
        result['items'].append(struct_work_subjudge_leader(
            rule, d['work'], d['judge_status'], d['score'], d['team'], d['expert_in_team_count'], d['expert_score']))

    result['id_all'] = ','.join(qualified_id_list).strip(',')
    result['judger_max'] = str(judger_max)
    return result


@transaction.atomic()
def decide_subjudge(account, user, role, activity, is_active, data):
    # 活动创建者、末端用户均不可创建子级评审
    if is_activity_owner(activity, account):
        raise BusinessException(ERR_USER_AUTH)
    if user.area.area_level == 0:
        raise BusinessException(ERR_USER_AUTH)

    subjg, is_created = SubJudge.objects.update_or_create(
            activity=activity, user=user, area=user.area,
            defaults={'is_active': int(is_active)}, )
    if int(is_active) == TRUE_INT:  # 评审后审批模式
        '''
        {
            'rank': [
                {'sort': '1', 'desc':'xxx'},
                {'sort': '2', 'desc':'xxx'},
                {'sort': '3', 'desc':'xxx'},
            ],
            'rule': {
                'code': '2',
                'judge_count': '6',
                'max': '100'
                'ignore_maxmin': '0'
            }
        }
        '''
        d = json.loads(data, strict=False)
        rank_data = d['rank']
        rule_data = d['rule']
        for rk in rank_data:
            SubJudgeRank.objects.create(subjudge=subjg, sn=int(rk['sort']), name=rk['desc'])
        SubJudgeRule.objects.create(
                subjudge=subjg, code=int(rule_data['code']), content=json.dumps(rule_data))
    return {'subjudge_id': str(subjg.id)}


def detail_subjudge(subjg):
    result = dict()
    result['is_active'] = str(subjg.is_active)
    result['rank'] = list()
    qs_rank = SubJudgeRank.objects.filter(subjudge=subjg)
    for each in qs_rank:
        result['rank'].append({
            'id': str(each.id),
            'sort': str(each.sn),
            'desc': each.name,
        })
    subjg_rule = SubJudgeRule.objects.filter(subjudge=subjg).first()
    result['rule'] = json.loads(subjg_rule.content)
    result['rule_code'] = str(subjg_rule.code)
    return result


def _team_expert_title(subjg_team_expert):
    sn = NUM[str(subjg_team_expert.sn)]
    leader = ''
    if subjg_team_expert.is_leader == TRUE_INT:
        leader = u'(组长)'
    return u'评审专家%s%s' % (sn, leader)


def detail_team(team):
    rule = SubJudgeRule.objects.filter(subjudge=team.subjudge, ).first()
    if not rule:
        raise BusinessException(ERR_RULE_NOT_EXIST)
    max = rule.parse_rule().expert_count()
    result = dict(judger_list=list(),
                  subjudge_team_id=str(team.id),
                  subjudge_team_name=team.name,
                  subjudge_team_work_count=str(team.work_count,))
    experts = SubJudgeTeamExpert.objects.filter(subjudge_team=team).order_by('sn')
    judger_list = list()
    for i in xrange(max):
        for each in experts:
            if each.sn == i+1:
                judger_list.append({
                    'subjudge_id': str(team.subjudge.id),
                    'expert_id': str(each.expert.id),
                    'expert_name': each.expert.name,
                    'sn': str(each.sn),
                    'is_leader': str(each.is_leader),
                    'expert_position': each.expert.position,
                    'expert_area_id': str(each.expert.area_id),
                    'expert_area_name_full': area_name(each.expert.area_id),
                })
                break
        else:
            judger_list.append({
                'subjudge_id': '',
                'expert_id': '',
                'expert_name': '',
                'sn': str(i+1),
                'is_leader': '',
                'expert_position': '',
                'expert_area_id': '',
                'expert_area_name_full': '',
            })
    result['judger_list'] = judger_list
    return result


def list_team(subjg, keyword, judger_keyword, rows, page, subjudge_team_id_list=None):
    rule = SubJudgeRule.objects.filter(subjudge=subjg, ).first()
    if not rule:
        raise BusinessException(ERR_RULE_NOT_EXIST)
    max = rule.parse_rule().expert_count()
    qs = SubJudgeTeam.objects.filter(subjudge=subjg).order_by('-create_time')
    if subjudge_team_id_list:
        qs = qs.filter(id__in=subjudge_team_id_list)
    if keyword:
        qs = qs.filter(name__contains=keyword)
    if judger_keyword:
        with_judger_team = SubJudgeTeamExpert.objects.filter(
                expert__account__name__contains=judger_keyword, subjudge=subjg).values_list('subjudge_team__id', flat=True)
        qs = qs.filter(id__in=with_judger_team)
    paged_data, result = paging_by_page(list(qs), rows, page)
    for each in paged_data:
        # work_count = SubJudgeWorkTeam.objects.filter(subjudge_team=each).count()
        # qs_expert_in_team = SubJudgeTeamExpert.objects.filter(subjudge_team=each)
        # expert_list = [{'expert_name': e.expert.name, 'expert_position': e.expert.position} for e in qs_expert_in_team]
        result['items'].append(detail_team(each))
    result['judger_max'] = str(max)
    result['id_all'] = ','.join([str(each.id) for each in qs])
    return result


def list_team_by_judger(account, subjg, as_leader):
    expert = Expert.objects.filter(del_flag=FALSE_INT, account=account).first()
    if not expert:
        raise BusinessException(ERR_USER_AUTH)

    teams_filterd = list()
    if as_leader and str(as_leader) == TRUE_STR:
        teams_filterd = SubJudgeTeamExpert.objects.filter(
            subjudge=subjg, expert=expert, is_leader=TRUE_INT).values_list('subjudge_team__id', flat=True)
    else:
        teams_filterd = SubJudgeTeamExpert.objects.filter(
            subjudge=subjg, expert=expert).values_list('subjudge_team__id', flat=True)

    subj_teams = SubJudgeTeam.objects.filter(id__in=teams_filterd)

    result = list()
    for each in subj_teams.values('id', 'name'):
        result.append({
            'subjudge_team_id': str(each['id']),
            'subjudge_team_name': each['name'],
        })
    return result


def edit_team(subjg, team, name):
    if team:
        team.name = name
        team.save()
    else:
        team = SubJudgeTeam.objects.create(subjudge=subjg, name=name)
    return str(team.id)


@transaction.atomic()
def delete_team(account, subjudge_team_obj_list):
    succ_count = 0
    for each in subjudge_team_obj_list:
        # 仅子级评审的维护者可使用
        if not is_subjudge_manager(each.subjudge, account):
            continue
        if SubJudgeTeamWork.objects.filter(subjudge_team=each).exists():
            continue
        each.del_flag = TRUE_INT
        each.save()
        SubJudgeTeamExpert.objects.filter(subjudge_team=each).update(del_flag=TRUE_INT)
        succ_count += 1
    msg1 = (u'%d个分组删除成功. ' % succ_count) if succ_count > 0 else ''
    fail_count = len(subjudge_team_obj_list) - succ_count
    msg2 = (u'%d个分组删除失败或不支持删除' % fail_count) if fail_count > 0 else ''
    return msg1 + msg2


def work_in_team(account, subjudge, subjudge_team, sub_activity, phase,
                project, area, direct_area, subject, keyword, page, rows):
    my_user, role = get_user(subjudge.activity, account)
    logger.info('user %s area %s wanna check works in subjudge_team %s' % (my_user.id, my_user.area.id, subjudge_team.id))
    qs = works(subjudge.activity, sub_activity=sub_activity, paragraph=phase,
               paragraph_project=project, subject=subject, keyword=keyword,
               area=area, direct_area=direct_area)
    work_id_list = [w.work.id for w in SubJudgeTeamWork.objects.filter(subjudge_team=subjudge_team)]
    qs = qs.filter(id__in=work_id_list)
    paged_data, result = paging_by_page(qs, rows, page)
    for each in paged_data:
        result['items'].append(struct_work(each, my_user.area))
    return result


def work_availale_add_team(account, subjudge, subjudge_team, sub_activity, phase, project,
                           area, direct_area, subject, keyword, ignore_area_list, page, rows):
    my_user, role = get_user(subjudge.activity, account)
    # 该级别管理员可见的作品
    all, non_approved = works_manager_can_see(
        my_user, subjudge.activity, sub_activity, phase, project, area, direct_area, subject, keyword)

    # 哪些作品可以加入分组（两种可能的规则：所有可见作品均可，仅未审核的可加入）
    handle = all

    subjg_t_w_work = [each.work for each in SubJudgeTeamWork.objects.all()]
    final_handle = filter(lambda x: x['work'] not in subjg_t_w_work, handle)
    paged_data, result = paging_by_page(final_handle, rows, page)
    for each in paged_data:
        result['items'].append(struct_work(each['work'], my_user.area))
    return result


def _update_team_work_count_stats(team):
    count = SubJudgeTeamWork.objects.filter(subjudge_team=team).count()
    team.work_count = count
    team.save()


@transaction.atomic()
def add_team_work(team, work_id_list):
    works = list(Work.objects.filter(id__in=work_id_list))
    works_in_team = [tw.work for tw in SubJudgeTeamWork.objects.filter(subjudge_team=team)]
    adding = set(works) - set(works_in_team)

    select_invalid = set(works) & set(works_in_team)
    if len(select_invalid)>0:
        logger.warn('works %s already in subjudge team, can not add them into subjudge team %s'
                    % (','.join([w.name for w in select_invalid]), team.name) )

    adding_list = list()
    for each in adding:
        subj_tw = SubJudgeTeamWork()
        subj_tw.subjudge = team.subjudge
        subj_tw.subjudge_team = team
        subj_tw.work = each
        adding_list.append(subj_tw)
    new_added = SubJudgeTeamWork.objects.bulk_create(adding_list)

    # 更新分组内的作品数统计
    _update_team_work_count_stats(team)

    msg1 = (u'%s个作品加入分组成功. ' % len(new_added)) if len(new_added) > 0 else ''
    msg2 = (u'%s个作品加入分组失败，可能该作品已在分组中' % len(select_invalid)) if len(select_invalid) > 0 else ''
    return msg1 + msg2


@transaction.atomic()
def remove_team_work(team, work_id_list):
    update_rows = SubJudgeTeamWork.objects.filter(
                subjudge_team=team, work__id__in=work_id_list).update(del_flag=TRUE_INT)

    SubJudgeScore.objects.filter(subjudge_team_work__work__id__in=work_id_list).update(del_flag=TRUE_INT)

    # 更新分组内的作品数统计
    _update_team_work_count_stats(team)
    return u'%s个作品移除成功' % update_rows


def export_team(account, subjudge, user, role, subjudge_team_id_list):
    teams = list_team(subjudge, None, None, BIG, 1, subjudge_team_id_list=subjudge_team_id_list)['items']
    rule = SubJudgeRule.objects.filter(subjudge=subjudge).first()
    if not rule:
        raise BusinessException(ERR_RULE_NOT_EXIST)
    max = rule.parse_rule().expert_count()

    wb = Workbook()
    sheet = wb.active

    expert_info_map = {}

    sheet.cell(column=1, row=1, value=u'组名')
    sheet.cell(column=2, row=1, value=u'组内作品数')
    for i in xrange(max):
        col = 3+i
        sn = 1+i
        sheet.cell(column=col, row=1, value=u'评审专家%s'%(sn))
        expert_info_map[sn] = col

    for r, each in enumerate(teams):
        cur_row = 2+r
        judger_dict = {int(judger['sn']):judger['expert_name'] for judger in each['judger_list']}

        sheet.cell(column=1, row=cur_row, value=each['subjudge_team_name'])
        sheet.cell(column=2, row=cur_row, value=each['subjudge_team_work_count'])
        for i in xrange(max):
            sn = i+1
            col = expert_info_map[sn]
            sheet.cell(column=col, row=cur_row, value=judger_dict[sn])
    return wb


def area_stats_team_expert(subjudge_team, user, role):
    # 仅子级评审的维护者可使用
    if not is_subjudge_manager(subjudge_team.subjudge, user.account):
        raise BusinessException(ERR_USER_AUTH)

    result = list()
    area_id_list_raw = set(SubJudgeTeamExpert.objects.filter(subjudge_team=subjudge_team).values_list('expert__area__id', flat=True))
    area_list_raw = Area.objects.filter(del_flag=FALSE_INT, id__in=area_id_list_raw)
    for each in area_list_raw:
        area = area_by_pov(each, None)
        if area:
            result.append({
                'area_id': str(area.id), 'area_name': area.area_name
            })
    return result


def available_add_team_expert(subjg_team, expert_in_same_team, keyword, page, rows):
    expert_in_subjg = SubJudgeExpert.objects.filter(subjudge=subjg_team.subjudge, expert__del_flag=FALSE_INT)
    if keyword:
        expert_in_subjg = expert_in_subjg.filter(Q(expert__account__username__contains=keyword) | Q(expert__account__name__contains=keyword))
    expert_in_subjg = expert_in_subjg.values_list('expert__id', flat=True)
    expert_exist_in_other_team = SubJudgeTeamExpert.objects.filter(subjudge=subjg_team.subjudge)\
                        .exclude(subjudge_team=subjg_team).values_list('expert__id', flat=True)
    in_team_page = [int(each) for each in seq2list(expert_in_same_team)] if expert_in_same_team else list()

    result_id_list = set(expert_in_subjg) - set(expert_exist_in_other_team) - set(in_team_page)
    qs = Expert.objects.filter(id__in=result_id_list, del_flag=FALSE_INT)
    paged_data, result = paging_by_page(qs, rows, page)
    for each in qs:
        result['items'].append({
            'subjudge_id': str(subjg_team.subjudge.id),
            'expert_id': str(each.id),
            'expert_name': each.account.name,
            'expert_mobile': each.account.username,
            'expert_position': each.position,
            'expert_area_id': str(each.area.id) if each.area else '',
            'expert_area_name_full': area_name(each.area.id) if each.area else '',
        })
    return result


@transaction.atomic()
def update_team_expert(team, data):
    d1 = json.loads(data)
    """
    [
        {'expert_id': 'x', 'sn': 'x'},
        ......
    ]
    """
    # 原有专家
    expert_old = [each.expert for each in SubJudgeTeamExpert.objects.filter(subjudge_team=team)]
    expert_team_old = {each.expert.id : each for each in SubJudgeTeamExpert.objects.filter(subjudge_team=team)}

    # 现有专家
    expert_new = [Expert.objects.filter(del_flag=FALSE_INT, id=int(each['expert_id'])).first() for each in d1]
    d2 = {int(each['expert_id']): int(each['sn']) for each in d1}

    # 处理保留的专家
    expert_keep = set(expert_old) & set(expert_new)
    for each in expert_keep:
        subjg_team_expert = expert_team_old[each.id]
        subjg_team_expert.sn = d2[each.id]
        is_leader = TRUE_INT if d2[each.id] == 1 else FALSE_INT
        subjg_team_expert.is_leader = is_leader
        subjg_team_expert.save()

    # 处理新增的专家
    expert_new_add = set(expert_new) - set(expert_old)
    for each in expert_new_add:
        sn = d2[each.id]
        is_leader = TRUE_INT if d2[each.id] == 1 else FALSE_INT
        SubJudgeTeamExpert.objects.create(
                subjudge=team.subjudge, subjudge_team=team, expert=each, sn=sn, is_leader=is_leader)

    # 处理删除的专家
    expert_remove = set(expert_old) - set(expert_new)
    with transaction.atomic():
        SubJudgeTeamExpert.objects.filter(
                expert__in=expert_remove, subjudge=team.subjudge, subjudge_team=team).update(del_flag=TRUE_INT)
        SubJudgeScore.objects.filter(subjudge_team=team, subjudge_team_expert__expert__in=expert_remove).update(del_flag=TRUE_INT)

    return TRUE_STR


def add_expert(account, subjg, expert_id_list):
    exist = SubJudgeExpert.objects.filter(subjudge=subjg, expert__id__in=expert_id_list)\
                .values_list('expert__id', flat=True)
    need_add = set(expert_id_list) - set(exist)
    experts = Expert.objects.filter(del_flag=FALSE_INT, id__in=need_add)
    bulk = list()
    for each in experts:
        # 是否支持将自己加入自己评审
        # if each.account == account:
        #     raise BusinessException(ERR_SUBJG_MANAGER_EXPERT_CONFILICT)
        b = SubJudgeExpert()
        b.subjudge = subjg
        b.expert = each
        bulk.append(b)
    SubJudgeExpert.objects.bulk_create(bulk)
    se_list = SubJudgeExpert.objects.filter(subjudge=subjg).values_list('expert__id', flat=True)
    return ','.join([str(each) for each in  se_list]) if se_list else ''


def new_expert(account, subjudge, user, mobile, name, sex, area, direct_area, institution, position):
    from applications.expert.services import add_expert_user
    area_id = area.id if area else None
    direct_area_id = direct_area.id if direct_area else None
    result = add_expert_user(account, mobile, name, sex, institution, None, position, None, area_id, direct_area_id)
    expert_id = result['d']['expert_id']
    add_expert(account, subjudge, [int(expert_id), ])
    return {'expert_id': expert_id}


def list_expert(account, subjg, keyword, rows, page):
    qualified_id_list = list()
    qs = SubJudgeExpert.objects.filter(subjudge=subjg).order_by('-create_time')

    if keyword:
        qs = qs.filter(Q(expert__account__username__contains=keyword) |
                       Q(expert__account__name__contains=keyword) |
                       Q(expert__position__contains=keyword))

    filtered_list = list(qs)
    for each in filtered_list:
        qualified_id_list.append(str(each.expert.id))

    paged_data, result = paging_by_page(filtered_list, rows, page)
    for each in paged_data:
        result['items'].append(struct_subjudge_expert(each.subjudge, each.expert))
    result['id_all'] = ','.join(qualified_id_list).strip(',')
    return result


def available_add_expert(account, subjg, keyword, rows, page):
    expert_in_subjudge = list_expert(account, subjg, keyword, BIG, 1)['items']
    expert_id_list_in_subjudge = [int(each['expert_id']) for each in expert_in_subjudge]
    all = Expert.objects.filter(del_flag=FALSE_INT)
    if keyword:
        all = all.filter(Q(account__username__contains=keyword) | Q(account__name__contains=keyword))
    all = all.exclude(id__in=expert_id_list_in_subjudge)
    paged_data, result = paging_by_page(list(all), rows, page)
    for each in paged_data:
        result['items'].append(struct_subjudge_expert(subjg, each))
    return result


def remove_expert(account, subjg, expert_id_list):
    expert_in_team = SubJudgeTeamExpert.objects.filter(subjudge=subjg, expert__id__in=expert_id_list)
    in_team = [each.expert for each in expert_in_team]
    all = SubJudgeExpert.objects.filter(expert__id__in=expert_id_list)
    fail_count = 0
    succ_count = 0
    for each in all:
        if each.expert in in_team:
            # 已经进组的专家不能删除
            fail_count += 1
        else:
            each.del_flag = TRUE_INT
            each.save()
            succ_count += 1
    msg1 = (u'%d个专家成功从评审中移除' % succ_count) if succ_count > 0 else ''
    msg2 = (u'%d个专家已经在评审组中，因而无法从评审中移除' % fail_count) if fail_count > 0 else ''
    return dict(c=SUCCESS[0], m=SUCCESS[1], d=msg1 + msg2)


def is_subjudge_manager(subjudge, account):
    return subjudge.user.account == account


def score(account, subjudge, subjudge_team_work, data, commit):
    rule = SubJudgeRule.objects.filter(subjudge=subjudge).first().parse_rule()
    expert = Expert.objects.filter(account=account, del_flag=FALSE_INT).first()
    # 权限校验
    is_super = is_subjudge_manager(subjudge, account)
    if not is_super:
        if not expert:
            raise BusinessException(ERR_USER_AUTH)
        # 非子级评审创建者则必须是本组专家才可以打分
        if not SubJudgeTeamExpert.objects.filter(
                subjudge_team=subjudge_team_work.subjudge_team, expert=expert).exists():
            raise BusinessException(ERR_USER_AUTH)

    id = rule.score(subjudge_team_work, data, commit, expert=expert)
    return {'score_id': str(id) if id else ''}


def get_score(account, subjudge, subjudge_team_work):
    expert = Expert.objects.filter(del_flag=FALSE_INT, account=account).first()
    te = SubJudgeTeamExpert.objects.filter(expert=expert, subjudge=subjudge, subjudge_team=subjudge_team_work.subjudge_team).first()
    rule = SubJudgeRule.objects.filter(subjudge=subjudge).first().parse_rule()

    # 权限校验
    if not expert or not te:
        logger.warn('account %s wanna get_score for work %s, but he/she is not a expert or not expert in that team'
                    % (account.id, subjudge_team_work.work.id))
        raise BusinessException(ERR_USER_AUTH)

    return rule.get_score(subjudge_team_work, subjudge_team_expert=te, ranks=rule.get_ranks(), account=account)


def load_all_score(account, subjudge, subjudge_team_work):
    expert = Expert.objects.filter(del_flag=FALSE_INT, account=account).first()
    rule = SubJudgeRule.objects.filter(subjudge=subjudge).first().parse_rule()

    # 权限校验
    is_super = is_subjudge_manager(subjudge, account)
    if not is_super:
        if not expert:
            raise BusinessException(ERR_USER_AUTH)
        # 非子级评审创建者则必须是组长才可以看所有得分
        if not SubJudgeTeamExpert.objects.filter(
                subjudge_team=subjudge_team_work.subjudge_team, expert=expert, is_leader=TRUE_INT).exists():
            raise BusinessException(ERR_USER_AUTH)

    return rule.get_score(subjudge_team_work, subjudge_team_expert=None, ranks=rule.get_ranks(), account=account)


def submit_score(account, subjudge, score_id_list):
    # expert = Expert.objects.filter(account=account).first()
    rule = SubJudgeRule.objects.filter(subjudge=subjudge).first().parse_rule()
    qs = SubJudgeScore.objects.filter(status=FALSE_INT, id__in=score_id_list, subjudge_team_work__subjudge=subjudge)
    subjudge_team_work_list = list()

    succ_count = 0
    for each in qs:
        subjudge_team_work_list.append(each.subjudge_team_work)
        each.status = TRUE_INT
        each.save()
        succ_count += 1

    # 更新每一个被提交得分的作品的最终获奖信息
    for each in set(subjudge_team_work_list):
        rule.update_work_status(each)

    fail_count = len(score_id_list) - succ_count
    msg1 = (u'%d个作品提交分数成功' % succ_count) if succ_count > 0 else ''
    msg2 = (u'%d个作品提交分数失败，可能还未评审或已提交' % fail_count) if fail_count > 0 else ''
    return msg1 + msg2


@transaction.atomic()
def import_expert(account, user, role, subjudge, excel):
    from applications.expert.services import add_expert_user
    ext = (excel.name[excel.name.rfind('.') + 1:]).lower()   # 后缀(不含.)
    if ext != 'xlsx':
        raise BusinessException(u'仅支持xlsx格式EXCEL')

    excel_store_name = 'import_subjudge_expert_%s_%s_%s.%s' % (
        datetime.datetime.now().strftime('%y%m%d%H%M%S'), account.id, subjudge.activity.id, ext)

    sheet = load_workbook(filename=BytesIO(excel.read())).active
    content = list(sheet.rows)
    content.pop(0)  # skip TIPS
    content.pop(0)  # skip HEAD
    expert_id_wanna_add = list()
    for i, row in enumerate(content):
        # 用户名 / 姓名 / 性别 / 地区 / 岗位或职务
        username = str(row[0].value)
        name = str(row[1].value)
        sex = str(row[2].value)

        # 猜测导入专家的区域
        guess_result, area, institution = GuessArea().guess(str(row[3].value))
        para_institution = None
        para_area_id = None
        para_direct_area_id = None
        if guess_result == GuessArea.RESULT_STATUS_FIND:
            para_area_id = area.id
        if guess_result == GuessArea.RESULT_STATUS_FIND_DIRECT:
            para_direct_area_id = area.id
            para_institution = institution
        if guess_result == GuessArea.RESULT_STATUS_NOT_FIND:
            para_area_id = china().id

        position = str(row[4].value)
        logger.info('call service to add user. username: %s, name:%s, sex:%s, area_id:%s, d_area_id:%s, insti:%s, posi:%s' %
                    (username, name, sex, para_area_id, para_direct_area_id, para_institution, position))
        add_result = add_expert_user(account, username, name, sex, para_institution, None, position, None, para_area_id, para_direct_area_id)
        if add_result['c'] != 0:
            raise BusinessException(ERR_ADD_USER)
        expert_id = add_result['d']['expert_id']
        expert_id_wanna_add.append(int(expert_id))

    expert_id_in_subjudge = [subj_expt.expert.id for subj_expt in SubJudgeExpert.objects.filter(subjudge=subjudge, expert__del_flag=FALSE_INT)]

    expert_exist = set(expert_id_wanna_add) & set(expert_id_in_subjudge)
    if len(expert_exist) > 0:
        logger.info('following expert is already in subjudge, no need to join again: %s' % ','.join(
            [str(each) for each in expert_exist]
        ))

    expert_id_can_add = set(expert_id_wanna_add) - set(expert_id_in_subjudge)
    expert_can_add = Expert.objects.filter(del_flag=FALSE_INT, id__in=expert_id_can_add)
    create_obj_list = list()
    for each in expert_can_add:
        create_obj_list.append(SubJudgeExpert(subjudge=subjudge, expert=each))
    SubJudgeExpert.objects.bulk_create(create_obj_list)

    path_r = os.path.join(settings.MEDIA_KEY_EXCEL, excel_store_name)
    path_a = os.path.join(settings.BASE_DIR, path_r)
    save_file(excel, path_a)  # 将用户导入的EXCEL保存下来

    return TRUE_STR

