# coding=utf-8
import json
import copy
import re
from io import BytesIO
import traceback
from urlparse import urljoin

import execjs
from lxml import etree

from applications.activity.models import Activity, Role, Ranks, Rule
from applications.activity.share import is_activity_owner
from applications.common.models import TaskTrace, TASK_EXPORT_WORK_BY_CREATOR, TASK_STATUS_WAIT, TASK_STATUS_SUCC, TASK_STATUS_FAIL, TASK_STATUS_DOING
from applications.common.services import get_account_activity_role, get_sub_area_id_list, get_account_activity_area, get_user_activity_role, \
    get_areadetail_by_id
from applications.score.models import Score, FinalScore
from applications.subjudge.models import SubJudgeTeamWork
from applications.team.models import TeamExpert
from applications.upload_resumable.models import FileObj
from applications.user.models import *
from applications.work.export import WorkExcelExport, EXCEL
from applications.work.models import WorkFileObj, WorkAttr, WorkAttrString, Work, WorkFlow, WorkVote
from applications.work.share import work_current_place, works, struct_work, struct_work_approve, struct_work_super, struct_work_upload, struct_work_judge, struct_work_judge_leader, \
    works_manager_can_see, LV_MAP, workflow, next_place, how_many_approve
from utils.const_def import *
from utils.file_fun import get_image_url, gen_path

from utils.public_fun import *
# from applications.user.agents import get_account_info
from applications.data.agents import list_file
from applications.work.task import resize_image_and_upload
from django.db import transaction
from django.db.models import F, Q
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl import Workbook, load_workbook
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import quote_sheetname

from utils.store_file import must_exist_folder
from utils.utils_except import BusinessException, TaskException

logger = logging.getLogger(__name__)


def __check_list_attr_value(activity_id, additions):
    list_attr_list = WorkAttr.objects.filter(activity_id=activity_id, type=WORK_ATTR_TYPE_LIST).\
                            values("id", "name", "values")
    for attr in list_attr_list:
        attr["values"] = attr["values"].split(";")
    list_attr_dict = convert_list_to_dict(list_attr_list, "id")
    for key, value in additions.items():
        attr_id = int(key)
        if attr_id in list_attr_dict.keys():
            available_value_list = list_attr_dict[attr_id][0]["values"]
            if value and value not in available_value_list:
                attr_name = list_attr_dict[attr_id][0]["name"]
                msg = u'作品属性"%s"不正确' % attr_name
                return [ERR_WORK_ATTR_VALUE_ERROR[0], msg]
    return SUCCESS


def __check_mandatory_attr(activity_id, additions):
    mandatory_id_list = WorkAttr.objects.filter(
            activity_id=activity_id, mandatory=TRUE_INT, del_flag=FALSE_INT).values_list("id", flat=True)
    mandatory_id_list = list(mandatory_id_list)
    for key, value in additions.items():
        attr_id = int(key)
        value = value.strip()
        if value and attr_id in mandatory_id_list:
            mandatory_id_list.remove(attr_id)
    if mandatory_id_list:
        return ERR_WORK_INFO_INCOMPLETE_ERROR
    return SUCCESS


def __get_authors(activity_id, work_id):
    work_attr_id_list = WorkAttr.objects.\
        filter(activity_id=activity_id, category=ACTIVITY_CATEGORY_AUTHOR_ATTR, name=WORK_ATTR_AUTHORS_NAME,
               del_flag=FALSE_INT).order_by("group_sn").values_list("id", flat=True)
    work_attr_id_list = list(work_attr_id_list)
    attr_string_list = WorkAttrString.objects.filter(work_id=work_id, attr_id__in=work_attr_id_list, del_flag=FALSE_INT).values("attr_id", "value")
    authors = ""
    for work_attr_id in work_attr_id_list:
        for attr_string in attr_string_list:
            if attr_string["value"] and attr_string["attr_id"] == work_attr_id:
                authors += attr_string["value"]
                authors += "/"
    if len(authors) > 0 and authors[-1] == "/":
        authors = authors[:-1]
    return authors


def __get_authors_from_additions(activity_id, additions):
    work_attr_id_list = WorkAttr.objects.\
        filter(activity_id=activity_id, category=ACTIVITY_CATEGORY_AUTHOR_ATTR, name=WORK_ATTR_AUTHORS_NAME,
               del_flag=FALSE_INT).order_by("group_sn").values_list("id", flat=True)
    work_attr_id_list = list(work_attr_id_list)
    authors = ""
    for work_attr_id in work_attr_id_list:
        for attr_id, value in additions.items():
            if value and int(attr_id) == work_attr_id:
                authors += value
                authors += "/"
    if len(authors) > 0 and authors[-1] == "/":
        authors = authors[:-1]
    return authors


def __get_attr_value_from_additions(activity_id, additions, category, name_list):
    attr_value = ""
    for name in name_list:
        work_attr = WorkAttr.objects.filter(activity_id=activity_id, category=category, name__contains=name,
                                            del_flag=FALSE_INT).order_by("group_sn").first()
        if work_attr:
            break
    if not work_attr:
        return attr_value

    for attr_id, value in additions.items():
        if value and int(attr_id) == work_attr.id:
            attr_value = value
            break
    return attr_value


def __get_attr_value_from_db(activity_id, work_id, category, name_list):
    attr_value = ""
    for name in name_list:
        work_attr = WorkAttr.objects.filter(activity_id=activity_id, category=category, name__contains=name,
                                            del_flag=FALSE_INT).order_by("group_sn").first()
        if work_attr:
            break
    if not work_attr:
        return attr_value
    attr_string = WorkAttrString.objects.filter(work_id=work_id, attr_id=work_attr.id, del_flag=FALSE_INT).first()
    if attr_string:
        attr_value = attr_string.value
    return attr_value


def __add_author_school():
    work_list = Work.objects.filter(del_flag=FALSE_INT, author_school="")
    for work in work_list:
        if not work.author_school:
            work.author_school = __get_attr_value_from_db(work.activity_id, work.id, ACTIVITY_CATEGORY_AUTHOR_ATTR,
                                                          [WORK_ATTR_AUTHORS_SCHOOL, WORK_ATTR_AUTHORS_INSTITUTION])
            logger.debug("work_id %s school %s" % (work.no, work.author_school))
        else:
            work.author_school = ""
        work.save()


@transaction.atomic
def create_work(user, role, activity, work_info):
    # 只有上传阶段可以
    if activity.stage != ACTIVITY_STAGE_UPLOAD:
        raise BusinessException(ERR_PAGE_NOT_ALLOW_APPROVE)

    work_info = json.loads(work_info)
    name = work_info.get("name", "")
    sub_activity = work_info.get("sub_activity", "")
    phase = work_info.get("phase", "")
    project = work_info.get("project", "")
    # subject = work_info.get("subject", "")
    rar_file_id = work_info.get("rar_file_id", "")
    additions = work_info.get("additions", {})
    authors = __get_authors_from_additions(activity.id, additions)
    author_school = __get_attr_value_from_additions(activity.id, additions, ACTIVITY_CATEGORY_AUTHOR_ATTR,
                                                    [WORK_ATTR_AUTHORS_SCHOOL, WORK_ATTR_AUTHORS_INSTITUTION])
    # 作品名称，学段，项目必填
    if not name or not phase or not project:
        raise BusinessException(ERR_WORK_INFO_INCOMPLETE_ERROR)
    if Work.objects.filter(name=name, phase=phase, project=project, uploader=user, authors=authors, del_flag=FALSE_INT).exists():
        return {"c": ERR_WORK_HAVE_EXIST[0], "m": ERR_WORK_HAVE_EXIST[1], "d": []}

    # 检查扩展属性的必填字段
    err_code = __check_mandatory_attr(activity.id, additions)
    if err_code[0] != SUCCESS[0]:
        return {"c": err_code[0], "m": err_code[1], "d": []}

    # 检查扩展属性的列表字段
    err_code = __check_list_attr_value(activity.id, additions)
    if err_code[0] != SUCCESS[0]:
        return {"c": err_code[0], "m": err_code[1], "d": []}

    # 检查学段项目是否正确
    err_code = __check_period_project(sub_activity, phase, project, activity.id)
    if err_code[0] != SUCCESS[0]:
        return {"c": err_code[0], "m": err_code[1], "d": []}

    # create work
    work_obj = Work.objects.create(
            activity=activity, name=name, sub_activity=sub_activity, phase=phase, project=project,
            uploader=user, authors=authors, author_school=author_school, area=user.area)

    for key, value in additions.items():
        attr_id = int(key)
        WorkAttrString.objects.create(work=work_obj, attr_id=attr_id, value=value)
        attr = WorkAttr.objects.filter(id=attr_id).first()
        if attr and attr.name == u'所属学科':
            work_obj.subject = value
            work_obj.save()

    # generate work number
    now = datetime.datetime.now()
    timestamp = now.strftime('%Y%m%d%H%M%S')
    work_no = "%d-%s-%d" % (activity.id, timestamp, work_obj.id)
    work_obj.no = work_no
    work_obj.status = WORK_STATUS_NOT_UPLOAD[0]
    if rar_file_id:
        # TODO check max size
        #
        file_obj = FileObj.objects.filter(id=int(rar_file_id), del_flag=FALSE_INT).first()
        if file_obj:
            work_obj.rar_file = file_obj
            work_obj.status = WORK_STATUS_NOT_SUBMIT[0]
    work_obj.save()

    activity.work_count = F('work_count') + 1
    activity.save()

    # create workflow
    WorkFlow.objects.create(work=work_obj, trigger=role, trigger_fullname=role.user.account.name, event=WORKFLOW_CREATE[0],
                            area=user.area, area_name=user.area.area_name, work_status=work_obj.status)

    return {"c": SUCCESS[0], "m": SUCCESS[1], "d": work_obj.id}


def list_work_super(account, my_user, activity, is_all, sub_activity, phase, project, area,
            direct_area, subject, rank, team, keyword, status_list, is_public, page, rows):
    qs = works(activity, sub_activity=sub_activity, paragraph=phase,
          paragraph_project=project, subject=subject, keyword=keyword,
          area=area, direct_area=direct_area, status_list=status_list)
    if is_public:
        qs = qs.filter(is_public=TRUE_INT)
    if is_all == FALSE_INT:
        qs = qs.filter(status__gte=WORK_STATUS_HAVE_EXAMINED[0])
    if rank:
        qs = qs.filter(ranks=rank)
    if team:
        qs = qs.filter(team=team)
    id_all = ','.join([str(each.id) for each in qs]).strip(',')
    paged_data, result = paging_by_page(qs, rows, page)
    for each in paged_data:
        result['items'].append(struct_work_super(each, my_user.area))
    result['id_all'] = id_all
    return result


def list_work_in_team(account, user, activity, sub_activity, phase, project, area, direct_area, subject, keyword, team, page, rows):
    qs = works(activity, sub_activity=sub_activity, paragraph=phase,
          paragraph_project=project, subject=subject, keyword=keyword,
          area=area, direct_area=direct_area)
    qs = qs.filter(status__gte=WORK_STATUS_REVIEWING[0], team=team)
    id_all = ','.join([str(each.id) for each in qs]).strip(',')
    paged_data, result = paging_by_page(qs, rows, page)
    for each in paged_data:
        result['items'].append(struct_work_super(each, user.area))
    result['id_all'] = id_all
    return result


def list_work_avalable_add_team(account, user, team, activity, sub_activity, phase, project, area,
                    direct_area, subject, keyword, ignore_area_list, page, rows):
    # TODO ignore_area_list check ERR
    # 过滤已经在组内的作品
    # TODO
    qs = works(activity, sub_activity=sub_activity, paragraph=phase,
          paragraph_project=project, subject=subject, keyword=keyword,
          area=area, ignore_area_id_list=ignore_area_list, direct_area=direct_area)
    qs = qs.filter(status=WORK_STATUS_HAVE_EXAMINED[0])
    paged_data, result = paging_by_page(qs, rows, page)
    for each in paged_data:
        result['items'].append(struct_work_super(each, user.area))
    return result


def list_work_expert(account, expert, activity, sub_activity=None, phase=None, project=None, subject=None,
                     keyword=None, team=None, is_unfinish=FALSE_INT, page=1, rows=10):
    qualified_id_list = list()
    qs = works(activity, sub_activity=sub_activity, paragraph=phase,
               paragraph_project=project, subject=subject, keyword=keyword)
    rule = Rule.objects.filter(activity=activity).first()

    teams_join = TeamExpert.objects.filter(expert=expert, team__activity=activity).values_list('team', flat=True)
    qs = qs.filter(team__in=teams_join)
    if team:
        qs = qs.filter(team=team)
    filter_list = list()
    for each in qs:
        te = TeamExpert.objects.filter(team=each.team, expert=expert).first()
        score = Score.objects.filter(work=each, team_expert=te).first()
        if score:
            if score.status == FALSE_INT:
                j_status = WORK_SHOW_STATUS_NOT_SUBMIT
            else:
                j_status = WORK_SHOW_STATUS_SUBMITED
                if is_1(is_unfinish):
                    continue
        else:
            j_status = WORK_SHOW_STATUS_NON_JUDGED
        filter_list.append({'work': each, 'judge_status': j_status, 'score': score, 'team': each.team})
        qualified_id_list.append(str(each.id))
    paged_data, result = paging_by_page(filter_list, rows, page)
    for d in paged_data:
        result['items'].append(struct_work_judge(d['work'], d['judge_status'], d['score'], d['team'], rule))
    result['id_all'] = ','.join(qualified_id_list).strip(',')
    return result


def list_work_leader(account, expert, activity, sub_activity, phase, project, subject, keyword,
                     team, is_unfinish, page, rows):
    rule = Rule.objects.filter(activity=activity).first()
    judger_max = rule.parse_rule().expert_count()
    qs = works(activity, sub_activity=sub_activity, paragraph=phase, paragraph_project=project,
               subject=subject, keyword=keyword)

    qualified_id_list = list()

    teams_join = TeamExpert.objects.filter(expert=expert, team__activity=activity, is_leader=TRUE_INT).values_list('team', flat=True)
    qs = qs.filter(team__in=teams_join)
    if team:
        qs = qs.filter(team=team)
    filter_list = list()

    # 活动专家数目
    rule = Rule.objects.filter(activity=activity).first()
    expert_count = rule.parse_rule().expert_count()

    for each in qs:
        te = TeamExpert.objects.filter(team=each.team, expert=expert).first()
        score = FinalScore.objects.filter(work=each, team_expert=te).first()
        if score:
            if score.status == FALSE_INT:
                j_status = WORK_SHOW_STATUS_NOT_SUBMIT
            else:
                j_status = WORK_SHOW_STATUS_SUBMITED
                if is_1(is_unfinish):
                    continue
        else:
            j_status = WORK_SHOW_STATUS_NON_JUDGED
        # 获取其它专家的评分
        expert_score = dict()
        score_qs = Score.objects.filter(work=each, team_expert__in=TeamExpert.objects.filter(team=each.team))
        for s in score_qs:
            expert_score[int(s.team_expert.sn)-1] = s
        filter_list.append({
            'work': each, 'judge_status': j_status, 'score': score, 'team': each.team, 'expert_score':expert_score})
        qualified_id_list.append(str(each.id))
    paged_data, result = paging_by_page(filter_list, rows, page)
    for d in paged_data:
        result['items'].append(struct_work_judge_leader(
            d['work'], d['judge_status'], d['score'], d['team'],
            expert_count, rule, d['expert_score']))
    result['id_all'] = ','.join(qualified_id_list).strip(',')
    result['judger_max'] = str(judger_max)
    return result


def export_work_super_all(account, activity, work_list_str):
    result = {
        'result': '',
        'result_desc': '',
        'progress': '',
        'output': '',
    }
    if TaskTrace.objects.filter(account=account, activity=activity, name=TASK_EXPORT_WORK_BY_CREATOR).count() > 1:
        logger.warn('find more than one task for account [%s] activity [%s] task [%s]'
                    % (account.id, activity.id, TASK_EXPORT_WORK_BY_CREATOR))
    task = TaskTrace.objects.filter(account=account, activity=activity, name=TASK_EXPORT_WORK_BY_CREATOR).first()
    need_create_new_task = False
    if task:
        if task.status == TASK_STATUS_WAIT:
            result['result'] = FALSE_STR
            result['result_desc'] = u'导出任务正在后台队列中，请稍后再次点击来查看结果'
            result['progress'] = '0%'
        if task.status == TASK_STATUS_SUCC:
            result['result'] = TRUE_STR
            result['progress'] = '100%'
            result['output'] = str(task.id)
        if task.status == TASK_STATUS_FAIL:
            need_create_new_task = True
        if task.status == TASK_STATUS_DOING:
            result['result'] = FALSE_STR
            result['progress'] = task.progress
            result['result_desc'] = u'导出任务正在进行中，当前进度%s%%，请稍后再次点击来查看结果' % task.progress
    else:
        need_create_new_task = True

    if need_create_new_task:
        TaskTrace.objects.create(account=account, activity=activity, name=TASK_EXPORT_WORK_BY_CREATOR, body=work_list_str)
        result['result'] = FALSE_STR
        result['progress'] = '0%'
        result['result_desc'] = u'导出任务已提交后台，需要一段时间处理，请稍后再次点击来查看结果'

    return result


def export_work_super_download(account, activity):
    # WARNNING: 由于没有使用S3存储，该接口暂时不支持分布式部署 (TODO)
    task = TaskTrace.objects.filter(account=account, activity=activity,
                        name=TASK_EXPORT_WORK_BY_CREATOR, status=TASK_STATUS_SUCC).first()
    if not task:
        raise BusinessException()
    task.del_flag = TRUE_INT
    task.save()
    return task.output


def export_work_super_all_task(task):
    activity = task.activity
    work_id_list = [int(each) for each in task.body.strip().strip(',').split(',') if each]
    work_list = list(Work.objects.filter(id__in=work_id_list))
    try:
        path = WorkExcelExport(activity, work_list,
                                (EXCEL.BASIC, EXCEL.STATUS, EXCEL.AUTHOR, EXCEL.TUTOR, EXCEL.FINAL_SCORE, EXCEL.SCORE),
                                task=task) \
            .write().dump_file()['d']
        task.output = path
        task.save()
    except Exception as e:
        another_logger = logging.getLogger('task_trace')
        another_logger.exception(e)
        raise TaskException(u'作品导出失败')


def list_work_manager(account, my_user, activity, is_approve=None, sub_activity=None, phase=None, project=None,
                      area=None, direct_area=None, subject=None, keyword=None, page=1, rows=10):
    rule = Rule.objects.filter(activity=activity).first()
    result_all, result_non_approve = works_manager_can_see(
        my_user, activity, sub_activity, phase, project, area, direct_area, subject, keyword)

    filter_data = result_non_approve if is_approve == FALSE_STR else result_all

    id_all_list = [str(each['work'].id) for each in filter_data]

    paged_data, ret = paging_by_page(filter_data, rows, page)

    from applications.subjudge.models import SubJudge, SubJudgeTeamWork, SubJudgeRule
    for each in paged_data:
        # 获取子级评审的状态、进度、成绩
        subjudge = SubJudge.objects.filter(activity=activity, user=my_user).first()
        progress = is_finish = final = ''
        should_subjudge = False
        if subjudge and subjudge.is_active == TRUE_INT:
            should_subjudge = True
            subjudge_rule = SubJudgeRule.objects.filter(subjudge=subjudge).first()
            # 需要进行子级评审
            is_finish, final, progress = subjudge_rule.parse_rule().get_judge_progress(each['work'])

        ret['items'].append(struct_work_approve(
                    each['work'], my_user.area, each['approve_status'], should_subjudge, progress, is_finish, final))
    ret['id_all'] = ','.join(id_all_list).strip(',')
    return ret


def list_work_upload(account, my_user, activity, sub_activity, phase, project, subject, keyword, status_list, page, rows):
    qs = works(activity, uploader=my_user, sub_activity=sub_activity, paragraph=phase,
                    paragraph_project=project, subject=subject, keyword=keyword, status_list=status_list)
    id_all_list = [str(each.id) for each in qs]
    paged_data, ret = paging_by_page(qs, rows, page)
    for each in paged_data:
        if each.status == WORK_STATUS_NOT_UPLOAD[0]:
            upload_status = WORK_SHOW_STATUS_NOT_UPLOAD
        elif each.status == WORK_STATUS_NOT_SUBMIT[0]:
            upload_status = WORK_SHOW_STATUS_NOT_SUBMIT
        else:
            upload_status = WORK_SHOW_STATUS_SUBMITED
        ret['items'].append(struct_work_upload(each, my_user.area, upload_status))
    ret['id_all'] = ','.join(id_all_list).strip(',')
    return ret


def detail_work(work_id, account=None):
    # TODO check work permission

    Work.objects.filter(id=work_id, del_flag=FALSE_INT).update(pv=F('pv')+1)
    work_info = Work.objects.filter(id=work_id, del_flag=FALSE_INT).\
        values("id", "no", "name", "sub_activity", "phase", "project", "subject", "status", "uploader__name",
               "is_public", "area_id", "area__area_name", "pv", "like", "rar_file__name", "rar_file_id",
               "rar_file__url", "activity_id", "authors", "author_school").first()
    work_info["author_name"] = work_info.pop("authors")
    work_info["uploader"] = work_info.pop("uploader__name")
    work_info["rar_url"] = get_image_url(work_info.pop("rar_file__url"))
    work_info["rar_file_name"] = work_info.pop("rar_file__name")
    work_info.pop("area__area_name")
    area = Area.objects.filter(id=int(work_info['area_id'])).first()
    work_info["area_name"] = area.area_name

    # 添加附件属性信息
    work_attr_list = WorkAttrString.objects.filter(work_id=work_id, del_flag=FALSE_INT).values("attr__id", "value")
    additions = {}
    for work_attr in work_attr_list:
        additions[work_attr["attr__id"]] = work_attr["value"]
    work_info["additions"] = additions

    # 添加文件信息, 新版本返回所有的内含文件，并将可预览文件放在前面
    work = Work.objects.filter(id=work_id).first()
    work_files = WorkFileObj.objects.filter(work=work).order_by('des_file__id')
    files_info_list = list()
    for w in work_files:
        same_list = list()
        similar = api_get_outsimipic(None, None, work.activity.id, w.src_file.id)['d']['same_in_site']
        for each_similar in similar:
            same_list.append({
                'activity_name' : each_similar['activity_name'],
                'work_no': each_similar['work_no'],
                'work_name': each_similar['work_name'],
                'work_authors': each_similar['work_authors'],
            })
        files_info_list.append({
            'work_file_id': str(w.id),
            'can_preview': TRUE_STR if w.des_file else FALSE_STR,
            'src_url': get_image_url(w.src_file.url) if w.src_file else '',
            'des_url': get_image_url(w.des_file.url) if w.des_file else '',
            'img_url': get_image_url(w.img_file.url) if w.img_file else '',
            'file_name': w.src_file.name,
            'size': str(w.src_file.size),
            'type': w.src_file.type,
            'same_count': str(len(same_list)),
            'same_list': same_list,
        })
    work_info["files"] = files_info_list

    return {"c": SUCCESS[0], "m": SUCCESS[1], "d": [work_info]}


@transaction.atomic
def update_work(account, user, role, work_obj, work_info):
    # decode work_info
    if not work_info:
        return {"c": ERR_WORK_INFO_ERROR[0], "m": ERR_WORK_INFO_ERROR[1], "d": []}
    work_info = json.loads(work_info)
    name = work_info.get("name", "")
    sub_activity = work_info.get("sub_activity", "")
    phase = work_info.get("phase", "")
    project = work_info.get("project", "")
    # subject = work_info.get("subject", "")
    zip_id = work_info.get("rar_file_id", '')
    rar_file_id = int(zip_id) if zip_id else ''
    additions = work_info.get("additions", {})

    # 三种情况下可以修改作品:
    # 1. 上传者只有在特定的阶段可以修改自己的作品
    # 2. 活动创建者可以修改作品
    # 3. 上级管理员当作品到达他这儿时也可以修改
    if is_activity_owner(work_obj.activity, account):
        pass
    else:
        where = workflow(work_obj)
        if where:
            handlers = where.handler()
            if user not in handlers:
                logger.warn('UPDATE_WORK check work %s fail, expect handler user %s, but find user %s'
                            % (work_obj.id, ','.join([str(each.id) for each in handlers]), user.id))
                raise BusinessException(ERR_USER_AUTH)
        else:
            # 兼容原有活动
            where = work_current_place(work_obj)
            logger.info('following user can update work %s: %s' % (work_obj.id, ','.join(where['current_handler_name'])))
            if role not in where['current_handler']:
                raise BusinessException(ERR_USER_AUTH)

    # 检查扩展属性的必填字段
    if additions:
        err_code = __check_mandatory_attr(work_obj.activity_id, additions)
        if err_code[0] != SUCCESS[0]:
            return {"c": err_code[0], "m": err_code[1], "d": []}

    # 检查扩展属性的列表字段
    if additions:
        err_code = __check_list_attr_value(work_obj.activity_id, additions)
        if err_code[0] != SUCCESS[0]:
            return {"c": err_code[0], "m": err_code[1], "d": []}

    # 检查学段项目是否正确
    if sub_activity or phase or project:
        err_code = __check_period_project(sub_activity, phase, project, work_obj.activity_id)
        if err_code[0] != SUCCESS[0]:
            return {"c": err_code[0], "m": err_code[1], "d": []}

    # update work info
    if name:
        work_obj.name = name
    if sub_activity:
        work_obj.sub_activity = sub_activity
    if phase:
        work_obj.phase = phase
    if project:
        work_obj.project = project
    # if subject:
    #     work_obj.subject = subject

    if rar_file_id:
        # 检查上传作品大小
        f = FileObj.objects.filter(id=rar_file_id).first()
        size_limit = work_obj.activity.get_upload_size_limit(
                        work_obj.sub_activity, work_obj.phase, work_obj.project)
        if int(f.size) > int(size_limit) * 1024 * 1024:
            raise BusinessException(ERR_WORK_FILE_TOO_BIG % size_limit)

        work_obj.rar_file_id = rar_file_id
        if work_obj.status == WORK_STATUS_NOT_UPLOAD[0]:
            work_obj.status = WORK_STATUS_NOT_SUBMIT[0]
            # create workflow
            WorkFlow.objects.create(
                    work=work_obj, trigger=role, trigger_fullname=role.user.account.name,
                    event=WORKFLOW_UPLOAD[0], area=role.user.area, area_name=role.user.area.area_name,
                    work_status=WORK_STATUS_NOT_SUBMIT[0], pre_flow=workflow(work_obj))
        else:
            work_obj.task_status = TASK_STATUS_NOT_PROCESS[0]
            work_obj.preview_status = WORK_PREVIEW_STATUS_NONE[0]
            # 清空历史文件
            _delete_work_file(work_obj)

    # update addition work info
    for key, value in additions.items():
        work_attr_id = int(key)
        obj, _ = WorkAttrString.objects.update_or_create(
            work_id=work_obj.id, attr_id=work_attr_id, del_flag=FALSE_INT,
            defaults={'value': value},
        )
        attr =  WorkAttr.objects.filter(id=work_attr_id).first()
        if attr and attr.name == u'所属学科':
            work_obj.subject = value
            work_obj.save()

    # update authors
    authors = __get_authors(activity_id=work_obj.activity_id, work_id=work_obj.id)
    author_school = __get_attr_value_from_additions(work_obj.activity_id, additions, ACTIVITY_CATEGORY_AUTHOR_ATTR,
                                                    [WORK_ATTR_AUTHORS_SCHOOL, WORK_ATTR_AUTHORS_INSTITUTION])
    work_obj.authors = authors
    work_obj.author_school = author_school

    # 检查作品是否已经存在
    dup_work = Work.objects.exclude(id=work_obj.id).filter(name=work_obj.name, phase=work_obj.phase, project=work_obj.project,
                                            uploader=work_obj.uploader, authors=work_obj.authors, del_flag=FALSE_INT).first()
    if dup_work:
        logger.warn('UPDATE_WORK, find duplicate work ID:%s NAME:%s' % (dup_work.id, dup_work.name))
        raise BusinessException(ERR_WORK_HAVE_EXIST)

    work_obj.save()

    return {"c": SUCCESS[0], "m": SUCCESS[1], "d": TRUE_STR}


def _delete_work_file(work_obj):
    # 将作品相关文件标记为删除
    work_file_list = WorkFileObj.objects.filter(work=work_obj, del_flag=FALSE_INT)
    for work_file in work_file_list:
        FileObj.objects.filter(id=work_file.src_file_id).update(del_flag=TRUE_INT)
        FileObj.objects.filter(id=work_file.des_file_id).update(del_flag=TRUE_INT)
        FileObj.objects.filter(id=work_file.img_file_id).update(del_flag=TRUE_INT)
    work_file_list.update(del_flag=TRUE_INT)

    # 将作品原始文件标记为删除
    FileObj.objects.filter(id=work_obj.rar_file_id).update(del_flag=TRUE_INT)
    FileObj.objects.filter(id=work_obj.img_file_id).update(del_flag=TRUE_INT)

    work_obj.img_file = None
    work_obj.rar_file = None
    work_obj.save()


def _check_delete_permission(user, activity, work):
    """
    校验删除权限:
        1. 活动创建者可以删除任意作品，但评审阶段之后不能删除
        2. 某级别的管理员可以删除自己上传的作品和下级上报上来的作品，且作品上报到父级后不能删除，且加入子级评审分组的作品也不能删除
        3. 作品上传者只能删除自己上传的作品，且作品提交后不能删除
    """
    if work.team:
        raise BusinessException(ERR_WORK_DEL_IN_JUDGE)
    if SubJudgeTeamWork.objects.filter(work=work).exists():
        raise BusinessException(ERR_WORK_DEL_IN_SUBJUDGE)
    if is_activity_owner(activity, user.account):
        return True
    user_with_perm = [r.user for r in work_current_place(work)['current_handler']]
    if user not in user_with_perm:
        raise BusinessException(ERR_WORK_DEL_NO_PERM)


# 删除作品导致用户提交作品数需要减少
def _delete_trigger_approve_limit_change(work):
    qs = WorkFlow.objects.filter(work=work).order_by('-update_time')
    for each in qs:
        skip = list()
        if not each:
            return
        if each.event == WORKFLOW_CREATE or each.event == WORKFLOW_UPLOAD:
            return
        if each.event == WORKFLOW_REJECT:
            skip.append(each.pre_flow)
        if each.event == WORKFLOW_APPROVE or each.event == WORKFLOW_SUBMIT:
            if each not in skip:
                each.trigger.approve_work = F('approve_work') - 1
                each.trigger.save()


@transaction.atomic
def delete_work(activity, account, user, work_list):
    work_can_delete = list()
    for each in work_list:
        try:
            _check_delete_permission(user, activity, each)
        except BusinessException as e:
            logger.warn(e.msg)
        else:
            work_can_delete.append(each)
    for w in work_can_delete:
        # 删除该作品对应的额外属性
        WorkAttrString.objects.filter(work_id=w.id).update(del_flag=TRUE_INT)
        # 将作品相关文件标记为删除
        _delete_work_file(w)
        # 删除作品记录
        w.del_flag=TRUE_INT
        w.save()

        activity.work_count = F('work_count') - 1
        activity.save()

    msg1 = (u'%d个作品删除成功' % len(work_can_delete)) if len(work_can_delete) > 0 else ''
    fail_count = len(work_list) - len(work_can_delete)
    msg2 = (u'%d个作品删除失败或不支持删除' % fail_count) if fail_count > 0 else ''
    return msg1 + msg2


# 1:add 2:subtract
def calculate_approve_count(role, many, action):
    if action == ADD:
        role.approve_work = F('approve_work') + int(many)
    elif action == SUBTRACT:
        role.approve_work = F('approve_work') - int(many)
    else:
        pass
    role.save()


@transaction.atomic()
def submit_work(activity, user, role, work_list, ):
    # 检查是否有权限提交这些作品
    if how_many_approve(role) != role.approve_work:
        logger.warn('approve count is conflict. method_count: %s, role_count: %s'
                        % (how_many_approve(role), role.approve_work))
    if role.max_work and (role.approve_work >= role.max_work):
        raise BusinessException(ERR_APPROVE_MAX_LIMIT)

    works_can_submit = list()
    for each in work_list:
        where = workflow(each)
        if each.status != WORK_STATUS_NOT_SUBMIT[0] \
                or user not in where.handler() \
                or each.activity != activity \
                or each.activity.stage != ACTIVITY_STAGE_UPLOAD:
            logger.warn('failed to submit work due to wrong status or no permission. work_id: %s' % each.id)
        else:
            works_can_submit.append(each)
            next_status, next_area = next_place(each, where.area, where.handler())
            WorkFlow.objects.create(work=each, pre_flow=where, trigger=role,
                        trigger_fullname=role.user.account.name, event=WORKFLOW_SUBMIT[0],
                        area=next_area, area_name=next_area.area_name, work_status=next_status)
            each.status = next_status
            each.save()

    calculate_approve_count(role, len(works_can_submit), ADD)

    msg1 = (u'%d个作品提交成功' % len(works_can_submit)) if len(works_can_submit) > 0 else ''
    fail_count = len(work_list) - len(works_can_submit)
    msg2 = (u'%d个作品提交失败或不支持提交' % fail_count) if fail_count > 0 else ''
    return msg1 + msg2


@transaction.atomic()
def approve_work(activity, user, work_list,):
    role = Role.objects.filter(user=user, activity=activity).first()
    if activity.stage != ACTIVITY_STAGE_UPLOAD:
        raise BusinessException(ERR_PAGE_NOT_ALLOW_APPROVE)
    work_can_approve = list()
    # 必须是当前用户有权限处理的作品
    for w in work_list:
        flow = workflow(w)
        current_handlers = flow.handler()
        if user not in current_handlers:
            continue
        next_status, next_area = next_place(w, flow.area, current_handlers)
        work_can_approve.append({'work': w, 'flow': flow, 'next_status': next_status, 'next_area': next_area})

    for each in work_can_approve:
        work = each['work']
        flow = each['flow']
        next_area = each['next_area']
        work.status = int(each['next_status'])
        work.save()
        WorkFlow.objects.create(
            work=work, pre_flow=flow, trigger=role, trigger_fullname=role.user.account.name, event=WORKFLOW_APPROVE[0],
            area=next_area, area_name=next_area.area_name, work_status=int(each['next_status']))
        calculate_approve_count(flow.trigger, 1, ADD)

    msg1 = (u'%d个作品审核成功' % len(work_can_approve)) if len(work_can_approve) > 0 else ''
    fail_count = len(work_list) - len(work_can_approve)
    msg2 = (u'%d个作品审核失败或不支持审核' % fail_count) if fail_count > 0 else ''
    return msg1 + msg2


@transaction.atomic()
def reject_work(activity, user, work_list,):
    role = Role.objects.filter(user=user, activity=activity).first()
    if activity.stage != ACTIVITY_STAGE_UPLOAD:
        raise BusinessException(ERR_PAGE_NOT_ALLOW_APPROVE)
    work_can_reject = list()
    # 必须是当前用户有权限处理的作品
    for w in work_list:
        flow = workflow(w)
        if user not in flow.handler():
            continue
        work_can_reject.append({'work': w, 'flow': flow})

    # 退回作品
    for each in work_can_reject:
        work = each['work']
        flow = each['flow']
        work.status = flow.pre_flow.work_status
        work.save()
        WorkFlow.objects.create(
                work=work, pre_flow=flow, trigger=role, trigger_fullname=role.user.account.name, event=WORKFLOW_REJECT[0],
                area=flow.pre_flow.area, area_name=flow.pre_flow.area_name, work_status=flow.pre_flow.work_status)
        calculate_approve_count(flow.trigger, 1, SUBTRACT)

    msg1 = (u'%d个作品退回成功' % len(work_can_reject)) if len(work_can_reject) > 0 else ''
    fail_count = len(work_list) - len(work_can_reject)
    msg2 = (u'%d个作品退回失败或不支持退回' % fail_count) if fail_count > 0 else ''
    return msg1 + msg2


@transaction.atomic()
def rank_work(activity, user, work_list, rank):
    done_count = len(work_list)
    for work in work_list:
        final_score = FinalScore.objects.filter(work=work).first()
        # 已有该作品的最终评定（组长已评定/平均分已生成）
        if final_score:
            final_score.rank = rank
            final_score.status = TRUE_INT
            final_score.save()
            work.ranks = rank
            work.save()
        # 无该作品的最终评定（组长已评定/平均分已生成）
        else:
            FinalScore.objects.create(work=work, rank=rank, status=TRUE_INT)
            work.ranks = rank
            work.save()

    msg1 = (u'%d个作品设奖成功' % done_count) if done_count > 0 else ''
    fail_count = len(work_list) - done_count
    msg2 = (u'%d个作品不支持设奖' % fail_count) if fail_count > 0 else ''
    return msg1 + msg2


def star_work(user, work_id):
    """
    给作品点赞，目前可以重复点赞，且点赞不可以取消。
    :param user:
    :param work_id:
    :return:
    """
    # 检查作品是否存在
    work_obj = Work.objects.filter(id=work_id, del_flag=FALSE_INT).first()
    if not work_obj:
        return {"c": ERR_WORK_ID_ERROR[0], "m": ERR_WORK_ID_ERROR[1], "d": []}

    work_obj.like = F('like') + 1
    work_obj.save()
    # 坑，返回work_obj.like会异常，只能再查一遍
    return {"c": SUCCESS[0], "m": SUCCESS[1], "d": Work.objects.filter(id=work_id, del_flag=FALSE_INT).first().like}


@transaction.atomic()
def vote_work(user, work_id, status):
    work_id = int(work_id)
    status = 1 if not status else int(status)

    # 检查作品是否存在
    work_obj = Work.objects.filter(id=work_id, del_flag=FALSE_INT).first()
    if not work_obj:
        return {"c": ERR_WORK_ID_ERROR[0], "m": ERR_WORK_ID_ERROR[1], "d": []}

    # 检查是否点赞过
    work_vote = WorkVote.objects.filter(work=work_obj, del_flag=FALSE_INT, account=user).first()
    if work_vote and status:
        raise BusinessException(ERR_WORK_IS_VOTE)
    if not work_vote and not status:
        raise BusinessException(ERR_WORK_IS_NO_VOTE)

    if status:
        work_obj.vote = F('vote') + 1
        new_star = WorkVote()
        new_star.work = work_obj
        new_star.account = user
        new_star.save()
    else:
        work_obj.vote = F('vote') - 1
        work_vote.del_flag = TRUE_INT
        work_vote.save()
    work_obj.save()

    return {"c": SUCCESS[0], "m": SUCCESS[1], "d": Work.objects.filter(id=work_id, del_flag=FALSE_INT).first().vote}


def download_work(activity, account, role, ea, work_list):
    work_can_download = work_list
    if not is_activity_owner(activity, account):
        work_can_download_1 = list()
        work_can_download_2 = list()
        if role:
            result, _ = works_manager_can_see(role.user, activity)
            result_work = [each['work'] for each in result]
            work_can_download_1 = [each for each in work_list if each in result_work]
        if ea:
            teams = TeamExpert.objects.filter(expert=ea.expert, team__activity=activity).values_list('team', flat=True)
            works_with_auth = Work.objects.filter(activity=activity, team__in=teams)
            work_can_download_2 = [each for each in work_list if each in works_with_auth]
        work_can_download = set(work_can_download_1 + work_can_download_2)
        work_without_perm = set(work_list) - work_can_download
        if len(work_without_perm) > 0:
            logger.warn('account %s wanna donwload work %s which is not perm for him' %
                            (account.id, ','.join([str(each.id) for each in work_without_perm])))

    return [each.rar_file.url for each in work_can_download if each.rar_file]


def match_work(activity, account, role, data):
    match_json = json.loads(data)
    for each in match_json:
        work_id = each['work_id']
        work = Work.objects.filter(id=int(work_id)).first()
        if not work:
            raise BusinessException(ERR_WORK_ID_ERROR)
        file_id = each['file_id']
        # TODO check max size
        #
        file_obj = FileObj.objects.filter(id=int(file_id), del_flag=FALSE_INT).first()
        if not file_obj:
            raise BusinessException(ERR_WORK_FILE_NOT_FOUND_ERR)
        work.rar_file = file_obj
        if work.status == WORK_STATUS_NOT_UPLOAD[0]:
            work.status = WORK_STATUS_NOT_SUBMIT[0]
        work.save()


WORK_PROPERTY_LIST = [u"作品名称*", u"学段*", u"项目*"]
WORK_PROPERTY_DICT = {u"作品名称": "name", u"学段": "phase", u"项目": "project", }


def download_work_template(activity_id, sub_activity):
    # check activity id
    if not activity_id:
        return {"c": ERR_ACTIVITY_ID_ERROR[0], "m": ERR_ACTIVITY_ID_ERROR[1], "d": []}
    activity_id = int(activity_id)
    activity_obj = Activity.objects.filter(id=activity_id, del_flag=FALSE_INT).first()
    if not activity_obj:
        return {"c": ERR_ACTIVITY_ID_ERROR[0], "m": ERR_ACTIVITY_ID_ERROR[1], "d": []}

    # 获取文件路径
    file_path = gen_path()
    wb = Workbook()
    ws = wb.active
    title1_font = Font(b=True, color="FF0000")
    title2_font = Font(b=True)

    # 设置学段-项目
    _set_project_data_validation(wb, ws, activity_obj.base_info_value, sub_activity, WORK_PROPERTY_LIST.index(u"学段*")+1,
                                  WORK_PROPERTY_LIST.index(u"项目*")+1)

    # 作品基本信息
    title2_row = 2
    title2_start_col = 1
    for work_property in WORK_PROPERTY_LIST:
        title2_cell = ws.cell(column=title2_start_col, row=title2_row, value=work_property)
        title2_cell.font = title2_font
        title2_start_col += 1

    work_attr_list = WorkAttr.objects.filter(activity_id=activity_id, category=ACTIVITY_CATEGORY_WORK_ATTR, del_flag=FALSE_INT).order_by("sn")
    for work_attr in work_attr_list:
        name = work_attr.name
        if work_attr.mandatory == TRUE_INT:
            name += "*"

        # 二级标题
        title2_cell = ws.cell(column=title2_start_col, row=title2_row, value=name)
        title2_cell.font = title2_font
        if work_attr.type == WORK_ATTR_TYPE_LIST:
            __set_data_validation(ws, work_attr.values, title2_start_col, title2_row+1)
        title2_start_col += 1

    # 一级标题，合并单元格
    title1_row = 1
    start_column = 1
    end_column = title2_start_col-1
    first_cell = ws.cell(column=start_column, row=title1_row, value=u'作品基本信息')
    ws.merge_cells(start_row=title1_row, start_column=start_column, end_row=title1_row, end_column=end_column)
    first_cell.alignment = Alignment(horizontal="center", vertical="center")
    first_cell.font = title1_font

    # 作者信息
    author_count = Activity.objects.filter(id=activity_id).values_list("author_count", flat=True)[0]
    for group_sn in xrange(author_count):
        group_sn += 1
        work_author_attr_list = []
        attr_list = WorkAttr.objects.filter(activity_id=activity_id, category=ACTIVITY_CATEGORY_AUTHOR_ATTR,
                                            group_sn=group_sn, del_flag=FALSE_INT).order_by("sn")
        if len(attr_list) == 0:
            continue
        for work_attr in attr_list:
            name = work_attr.name
            if work_attr.mandatory == TRUE_INT:
                name += "*"
            work_author_attr_list.append(name)

            # 二级标题
            title2_cell = ws.cell(column=title2_start_col, row=title2_row, value=name)
            title2_cell.font = title2_font
            if work_attr.type == WORK_ATTR_TYPE_LIST:
                __set_data_validation(ws, work_attr.values, title2_start_col, title2_row+1)
            title2_start_col += 1

        # 一级标题，合并单元格
        group_name = u'作者%d信息' % group_sn
        start_column = end_column + 1
        end_column = start_column + len(work_author_attr_list) - 1
        first_cell = ws.cell(column=start_column, row=title1_row, value=group_name)
        ws.merge_cells(start_row=title1_row, start_column=start_column, end_row=title1_row, end_column=end_column)
        first_cell.alignment = Alignment(horizontal="center", vertical="center")
        first_cell.font = title1_font

    # 指导教师信息
    tutor_count = Activity.objects.filter(id=activity_id).values_list("tutor_count", flat=True)[0]
    for group_sn in xrange(tutor_count):
        group_sn += 1
        work_author_attr_list = []
        attr_list = WorkAttr.objects.filter(activity_id=activity_id, category=ACTIVITY_CATEGORY_TUTOR_ATTR,
                                            group_sn=group_sn, del_flag=FALSE_INT).order_by("sn")
        if len(attr_list) == 0:
            continue
        for work_attr in attr_list:
            name = work_attr.name
            if work_attr.mandatory == TRUE_INT:
                name += "*"
            work_author_attr_list.append(name)

            title2_cell = ws.cell(column=title2_start_col, row=title2_row, value=name)
            title2_cell.font = title2_font
            if work_attr.type == WORK_ATTR_TYPE_LIST:
                __set_data_validation(ws, work_attr.values, title2_start_col, title2_row+1)
            title2_start_col += 1

        # 一级标题，合并单元格
        group_name = u'指导教师%d信息' % group_sn
        start_column = end_column + 1
        end_column = start_column + len(work_author_attr_list) - 1
        first_cell = ws.cell(column=start_column, row=title1_row, value=group_name)
        ws.merge_cells(start_row=title1_row, start_column=start_column, end_row=title1_row, end_column=end_column)
        first_cell.alignment = Alignment(horizontal="center", vertical="center")
        first_cell.font = title1_font

    wb.save(file_path)
    return file_path


def __set_data_validation(ws, json_str, column, start_row):
    # 构造备选值字符串
    # value_list = json.loads(json_str)
    value_list = json_str.split(";")  # 前台通过;分割
    values_str = ""
    for value in value_list:
        values_str += value
        values_str += ","
    if values_str[-1] == ",":
        values_str = values_str[:-1]
    values_str = '"' + values_str + '"'
    # 创建 DV
    dv = DataValidation(type="list", formula1=values_str, allow_blank=True)
    dv.error = u'请输入列表中的选项'
    dv.errorTitle = u'错误输入'
    ws.add_data_validation(dv)

    # 设置DV
    col_letter = get_column_letter(column)
    cell_range = "%s%d:%s%d" % (col_letter, start_row, col_letter, 1048576)
    dv.ranges.append(cell_range)


def _set_project_data_validation(wb, ws, json_str, sub_activity_name, period_data_col, project_data_col, project_data_start_row=3):
    # 构造备选值字符串
    activity_list = json.loads(json_str)
    period_col = 100
    project_col = period_col + 1

    for activity in activity_list:
        activity_name = activity["ac_type"]
        period_list = activity["period_list"]
        if activity_name != sub_activity_name:
            continue

        # 填充项目备选列表标题
        period_row = 1
        ws.cell(column=project_col, row=period_row, value=activity_name)
        period_row += 1
        period_start_row = period_row

        for period in period_list:
            period_name = period["period"]
            project_list = period["item_list"]

            # 填充学段备选列表
            ws.cell(column=period_col, row=period_row, value=period_name)
            period_row += 1

            # 填充项目备选列表标题
            project_row = 1
            ws.cell(column=project_col, row=project_row, value=period_name)
            project_row += 1
            project_start_row = project_row

            for project_info in project_list:
                # 填充项目备选列表数据
                ws.cell(column=project_col, row=project_row, value=project_info['name'])
                project_row += 1

            # 创建项目 name_range(EXCEL中“公式”菜单下“名称管理器”)
            project_col_letter = get_column_letter(project_col)
            cell_range = "%s%d:%s%d" % (project_col_letter, project_start_row, project_col_letter, project_row-1)

            wb.create_named_range(name=period_name, worksheet=ws, value=cell_range)

            project_col += 1

        # 创建学段的DV(EXCEL中“数据”菜单下“数据验证”)
        period_col_letter = get_column_letter(period_col)
        cell_range = "%s!$%s$%d:$%s$%d" % (ws.title, period_col_letter, period_start_row, period_col_letter, period_row-1)
        dv = DataValidation(type="list", formula1=cell_range.format(quote_sheetname("dv_sheet")) )
        dv.error = u'请输入列表中的选项'
        dv.errorTitle = u'错误输入'
        ws.add_data_validation(dv)

        # 设置 学段 DV
        period_data_col_letter = get_column_letter(period_data_col)
        cell_range = "%s%d:%s%d" % (period_data_col_letter, project_data_start_row, period_data_col_letter, 1048576)
        dv.ranges.append(cell_range)  # 要求openpyxl版本低于2.5.1

        project_col += 1

    # 创建 项目 DV
    for row in xrange(1000):
        if row < project_data_start_row:
            continue
        period_data_col_letter = get_column_letter(period_data_col)
        formula_str = "=INDIRECT($%s%d)" % (period_data_col_letter, row)
        dv = DataValidation(type="list", formula1=formula_str, allow_blank=True)
        dv.error = u'请输入列表中的选项'
        dv.errorTitle = u'错误输入'
        ws.add_data_validation(dv)

        # 设置 项目 DV
        project_data_col_letter = get_column_letter(project_data_col)
        cell_name = "%s%d" % (project_data_col_letter, row)
        cell = ws[cell_name]
        dv.add(cell)


def import_work(account, user, role, activity, sub_activity, file_obj):
    # 获取作品的WorkAttr
    work_attr_list = WorkAttr.objects.filter(activity=activity).values("id", "name", "category", "group_sn")

    ws = load_workbook(filename=BytesIO(file_obj.read())).active

    # 检查Excel模版
    title1_row = 1
    title2_row = 2
    for i in range(len(WORK_PROPERTY_LIST)):
        if WORK_PROPERTY_LIST[i] != ws.cell(column=i+1, row=title2_row).value:
            return {"c": ERR_FILE_TEMPLATE_ERROR[0], "m": ERR_FILE_TEMPLATE_ERROR[1],  "d": []}

    ret_data = []
    row_num = 0
    title2_list = []  # ( "name","phase","project", "subject", "attr_id")
    title1_list = []  # (category, group_sn)
    for row in ws.rows:
        work_info = {"id": "", "name": "", "sub_activity": sub_activity, "phase": "", "project": "", "subject": "", "additions": {}}
        # 一级标题栏
        row_num += 1
        if row_num == title1_row:
            none_count = 0
            title1_name = ""
            for cell in row:
                if none_count > MAX_WORK_ATTR_IN_GROUP:
                    break
                value = cell.value
                if not value:
                    none_count += 1
                    if not title1_name:
                        return {"c": ERR_WORK_IMPORT_TITLE1_ERR[0], "m": ERR_WORK_IMPORT_TITLE1_ERR[1], "d": []}
                    title1 = title1_name
                else:
                    none_count = 0
                    title1 = value.strip()
                    title1_name = title1
                title1 = _convert_title1_category(title1)
                title1_list.append(title1)

        # 二级标题列表
        col_num = 0
        if row_num == title2_row:
            for cell in row:
                col_num += 1
                value = cell.value
                if not value:
                    break
                property_name = value.strip()
                property_name = property_name.replace("*", "")
                if property_name in WORK_PROPERTY_DICT.keys():
                    title2_list.append(WORK_PROPERTY_DICT[property_name])
                    continue
                for work_attr in work_attr_list:
                    category, group_sn = title1_list[col_num-1]
                    if work_attr["name"] == property_name and work_attr["category"] == category and work_attr["group_sn"] == group_sn:
                        title2_list.append(work_attr["id"])
                        break
            # 检查二级标题数量
            if col_num-1 != len(title2_list):
                return {"c": ERR_WORK_IMPORT_TITLE2_ERR[0], "m": ERR_WORK_IMPORT_TITLE2_ERR[1], "d": []}

        # 解析数据
        if row_num > title2_row:
            for cell in row:
                col_num += 1
                if col_num >= len(title2_list):
                    break
                title2 = title2_list[col_num-1]

                # 检查单元格内数据类型，并进行转换。
                value = excel_cell2str(cell.value)

                # 附加属性
                if isinstance(title2, (int, long)):
                    work_info["additions"][title2] = value
                # 作品表属性
                else:
                    work_info[title2] = value

        # 添加作品
        if not work_info["name"]:
            continue
        work_info_str = json.dumps(work_info)
        err_code = create_work(user, role, activity, work_info_str)
        if err_code['c'] != SUCCESS[0]:
            err_msg = u"第%d行[%s]: " % (row_num, work_info["name"])
            err_msg += err_code['m']
            ret_data.append(err_msg)

    if ret_data:
        return {"c": ERR_IMPORT_DATA_ERROR[0], "m": ERR_IMPORT_DATA_ERROR[1], "d": ret_data}
    else:
        return {"c": SUCCESS[0], "m": SUCCESS[1], "d": []}


def _convert_title1_category(title1):
    category = ACTIVITY_CATEGORY_WORK_ATTR
    if u"作者" in title1:
        category = ACTIVITY_CATEGORY_AUTHOR_ATTR
    if u"教师" in title1:
        category = ACTIVITY_CATEGORY_TUTOR_ATTR
    if category == ACTIVITY_CATEGORY_WORK_ATTR:
        return category, 1
    group_sn = MAX_WORK_ATTR_IN_GROUP
    while group_sn > 0:
        if str(group_sn) in title1:
            return category, group_sn
        group_sn -= 1
    raise Exception(ERR_WORK_IMPORT_TITLE1_ERR[1])


# def __get_attr_value(work_addition_dict, work_id, attr_id):
#     ret_value = ""
#     if work_id not in work_addition_dict.keys():
#         return ret_value
#     attr_value_list = work_addition_dict[work_id]
#     if not attr_value_list:
#         return ret_value
#     for attr_value in attr_value_list:
#         if attr_id == attr_value["attr_id"]:
#             return attr_value["value"]
#     return ret_value


def list_no_file_work(user, activity_id, cur_user_id):
    cur_user = User.objects.get(id=cur_user_id, del_flag=FALSE_INT)

    ret_dict = list_file(user, cur_user, activity_id)

    if ret_dict["c"] != SUCCESS[0]:
        return ret_dict
    file_list = ret_dict["d"]
    work_list = Work.objects.filter(activity_id=activity_id, uploader=cur_user, status=WORK_STATUS_NOT_UPLOAD[0], del_flag=FALSE_INT).\
        values("id", "no", "name", "sub_activity", "phase", "project", "subject", "authors")
    work_list = list(work_list)
    for i, work in enumerate(work_list):
        author_name = work["authors"].split("/")[0]
        possible_file_list = _find_possible_file(file_list, [author_name, work["name"], work["no"]])
        work["possible_files"] = possible_file_list
        work["index"] = str(i)
    return {"c": SUCCESS[0], "m": SUCCESS[1], "d": work_list}


MAX_POSSIBLE_FILE_NUM = 3


def _find_possible_file(file_list, key_list):
    ret_file_list = []
    ret_file_id_list = []
    # 检查文件名是否以作者/作品名命名
    for file_info in file_list:
        file_name = file_info["name"]
        file_name = os.path.splitext(file_name)[0]
        if file_name in key_list:
            ret_file_list.append(file_info)
            ret_file_id_list.append(file_info["id"])
    # if ret_file_list:
    #     return ret_file_list
    # 检查文件名包含作者或作品名
    for file_info in file_list:
        file_name = file_info["name"]
        for key in key_list:
            if key and file_name and file_name.find(key) >= 0:
                if file_info["id"] not in ret_file_id_list:
                    ret_file_list.append(file_info)
                    ret_file_id_list.append(file_info["id"])
                break
        if len(ret_file_list) >= MAX_POSSIBLE_FILE_NUM:
            break

    # 检查作者或作品名中包含文件名
    for file_info in file_list:
        file_name = file_info["name"]
        file_name = os.path.splitext(file_name)[0]
        for key in key_list:
            if key and file_name and key.find(file_name) >= 0:
                if file_info["id"] not in ret_file_id_list:
                    ret_file_list.append(file_info)
                    ret_file_id_list.append(file_info["id"])
                break
        if len(ret_file_list) >= MAX_POSSIBLE_FILE_NUM:
            break
    return ret_file_list


def __check_period_project(in_sub_activity, in_phase, in_project, activity_id=0, json_str=""):
    base_info_value = ""
    if not json_str and activity_id > 0:
        activity_obj = Activity.objects.filter(id=activity_id).first()
        base_info_value = activity_obj.base_info_value
    elif json_str:
        base_info_value = json_str

    # 构造备选值字符串
    activity_list = json.loads(base_info_value)

    sub_activity_ok = False
    period_ok = False
    project_ok = False
    for activity in activity_list:
        activity_name = activity["ac_type"]
        period_list = activity["period_list"]
        if activity_name and activity_name != in_sub_activity:
            continue
        sub_activity_ok = True
        for period in period_list:
            period_name = period["period"]
            project_list = period["item_list"]

            if period_name != in_phase:
                continue
            period_ok = True

            for each in project_list:
                # NEW
                if isinstance(each, types.DictType):
                    if in_project == each['name']:
                        project_ok = True
                        break
                # OLD
                else:
                    if in_project == each:
                        project_ok = True
                        break

    if not sub_activity_ok:
        return ERR_WORK_SUB_ACTIVITY_ERROR
    elif not period_ok:
        return ERR_WORK_PERIOD_ERROR
    elif not project_ok:
        return ERR_WORK_PROJECT_ERROR
    else:
        return SUCCESS


def api_get_outsimipic(domain, pic_url, activity_id, file_obj_id):
    # 查询有没有相同相似的资源
    import urllib

    guessword = []
    same_in_site_list = []
    samelist_result = []
    similist_result = []

    USE_NEW_BAIDU_API = True  # 走新老百度接口的控制开关。
    # 站内查询有没有相同文件
    if not pic_url and not file_obj_id:
        raise BusinessException(ERR_REQUEST_PARAMETER_ERROR)

    try:
        # 内部文件直接取md5，外部文件下载文件后取md5
        md5 = ''
        if file_obj_id:
            file_obj = FileObj.objects.filter(id=file_obj_id).first()
            pic_url = get_image_url(file_obj.url)
            if not file_obj:
                raise BusinessException(ERR_WORK_FILE_NOT_FOUND_ERR)
            from applications.upload_resumable.files_s3 import get_file_md5
            md5 = file_obj.md5sum if file_obj.md5sum else get_file_md5(file_obj.url)

        else:
            from applications.upload_resumable.utils_data import get_nonce
            must_exist_folder(settings.TEMP_DIR)
            file_name_tmp = get_nonce(6)
            file_path_tmp = os.path.join(settings.TEMP_DIR, file_name_tmp)
            urllib.urlretrieve(pic_url, file_path_tmp)

            # 计算md5
            from utils.file_fun import get_file_md5
            md5 = get_file_md5(file_path_tmp)
            os.remove(file_path_tmp)

        # 去文件中找是否有md5相同的文件
        samemd5_file = FileObj.objects.filter(md5sum=md5, del_flag=FALSE_INT)
        if md5 in (MD5_COMPUTE_STATUS_GETFILEERR, MD5_COMPUTE_STATUS_COMPUTING):
            # 如果是前期没有计算出md5的文件，则只查询自己
            samemd5_file = samemd5_file.filter(id=file_obj_id)
            logger.info('activity_id=%s, file_obj_id=%s, md5=%s' % (activity_id, file_obj_id, md5))
            
        if file_obj_id:
            samemd5_file = samemd5_file.exclude(id=file_obj_id)
        same_in_site_list = list()
        for each_file in samemd5_file:
            # 不检查作品压缩包，只检查作品解压后的文件，因为压缩包最终还是会解压的。
            workfileobj = WorkFileObj.objects.filter(Q(src_file=each_file) | Q(des_file=each_file) | Q(img_file=each_file))
            if activity_id:
                workfileobj = workfileobj.exclude(work__activity_id=activity_id)
            for each_fileobj in workfileobj:
                if each_fileobj:
                    same_in_site_dict = {
                        "work_id": each_fileobj.work_id,
                        "work_no": each_fileobj.work.no,
                        "work_name": each_fileobj.work.name,
                        "work_authors": each_fileobj.work.authors,
                        "activity_id": each_fileobj.work.activity_id,
                        "activity_name": each_fileobj.work.activity.name,
                        "file_url": get_image_url(each_file.url),
                    }
                    same_in_site_list.append(same_in_site_dict)
                else:
                    continue

    except Exception as e:
        logger.error(traceback.format_exc())
        same_in_site_list = []

    try:
        # 只查询图片，其它文件类型的只在本地查
        file_ext = pic_url.split('.')[-1]
        if '.' + file_ext.lower() in FILE_TYPE_IMG[1] and not USE_NEW_BAIDU_API:

            # 到百度搜索外部相似图片或者相同图片，目前还缺少搜索站内相似图片功能。
            if pic_url.lower().split(':')[0] not in ('http', 'ftp', 'https'):
                pic_url = urljoin('http://' + domain, pic_url)

            pic_url_quote = urllib.quote(pic_url)
            url = 'http://image.baidu.com/n/pc_search?queryImageUrl=%s&uptype=urlsearch' % pic_url_quote
            print url
            # url = 'https://image.baidu.com/pcdutu?queryImageUrl=%s&fm=index&uptype=paste'  % pic_url
            headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/62.0.3202.89 Safari/537.36'}
            cook = {"Cookie": 'BAIDUID=FE0F97F1FC37C47792091A2523CD945F:FG=1; HMACCOUNT=CC6D0E280C842123'}
            res = requests.get(url, cookies=cook, headers=headers, timeout=10)
            logger.info('get res ok!')
            doc = etree.HTML(res.text)
            data = ''.join(doc.xpath('//script[contains(text(),"window.bd")]/text()'))
            data = data.replace('window.location.protocol === ', '')
            json_raw = re.search(r'({[\S\s]*});\s*bd.queryImageUrl([\S\s]*)', data).group(1)
            jsn = execjs.eval(json_raw)
            samelist = jsn.get('sameList', []) if jsn.get('sameList', []) else []
            similist = jsn.get('simiList', []) if jsn.get('simiList', []) else []
            guessword = jsn.get('guessWord', []) if jsn.get('guessWord', []) else []

            samelist_result = []
            similist_result = []

            for each_same in samelist:
                rowdict = {
                    "thumbURL": each_same['thumbURL'],
                    "fromURL": each_same['fromURL'],
                    "textHost": each_same['textHost'],
                    "fromPageTitle": each_same['fromPageTitle'],
                    "flowURL": each_same['flowURL'],
                }
                samelist_result.append(rowdict)

            for each_simi in similist:
                rowdict = {
                    "fromPageTitle": each_simi['fromPageTitle'],
                    "thumbURL": each_simi['MiddleThumbnailImageUrl'],
                    "fromURL": each_simi['fromURL'],
                    "flowURL": each_simi['objURL'],
                }
                similist_result.append(rowdict)

    except Exception as e:
        logger.error(traceback.format_exc())
        guessword = []
        samelist_result = []
        similist_result = []

    try:
        # 获取百度图片方式二
        # 只查询图片，其它文件类型的只在本地查
        file_ext = pic_url.split('.')[-1]
        if '.' + file_ext.lower() in FILE_TYPE_IMG[1] and USE_NEW_BAIDU_API:

            # 到百度搜索外部相似图片或者相同图片，目前还缺少搜索站内相似图片功能。
            if pic_url.lower().split(':')[0] not in ('http', 'ftp', 'https'):
                pic_url = urljoin('http://' + domain, pic_url)

            url = "https://graph.baidu.com/upload"

            payload = r"""------WebKitFormBoundaryF8xwZ1KVoapKxh73
Content-Disposition: form-data; name="image"

%s
------WebKitFormBoundaryF8xwZ1KVoapKxh73
Content-Disposition: form-data; name="tn"

pc
------WebKitFormBoundaryF8xwZ1KVoapKxh73
Content-Disposition: form-data; name="from"

pc
------WebKitFormBoundaryF8xwZ1KVoapKxh73
Content-Disposition: form-data; name="image_source"

PC_UPLOAD_SEARCH_URL
------WebKitFormBoundaryF8xwZ1KVoapKxh73
Content-Disposition: form-data; name="range"

{"page_from": "searchIndex"}
------WebKitFormBoundaryF8xwZ1KVoapKxh73--
""" % pic_url
            headers = {
                'content-type': "multipart/form-data; boundary=----WebKitFormBoundaryF8xwZ1KVoapKxh73",
                'Host': "graph.baidu.com",
                'Connection': "keep-alive",
                'Accept': "*/*",
                'Origin': "https://www.baidu.com",
                'User-Agent': "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/69.0.3497.100 Safari/537.36",
                'Content-Type': "multipart/form-data; boundary=----WebKitFormBoundaryF8xwZ1KVoapKxh73",
                'Referer': "https://www.baidu.com/",
                'Accept-Encoding': "gzip, deflate, br",
                'Accept-Language': "zh-CN,zh;q=0.9",
                'Cache-Control': "no-cache",
            }

            res = requests.request("POST", url, data=payload, headers=headers)

            step2_url = json.loads(res.text)['data']['url']

            headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/62.0.3202.89 Safari/537.36'}
            cook = {"Cookie": 'BAIDUID=FE0F97F1FC37C47792091A2523CD945F:FG=1; HMACCOUNT=CC6D0E280C842123'}
            res = requests.get(step2_url, cookies=cook, headers=headers, timeout=10)

            doc = etree.HTML(res.text)
            data = ''.join(doc.xpath('//script[contains(text(),"window.cardData")]/text()'))
            json_raw = re.search(r'window.cardData = ([\S\s]*)window.commonData([\S\s]*)', data).group(1)
            jsn = execjs.eval(json_raw.strip(';'))
            # print jsn[3]['cardName']

            for each_obj in jsn:
                if each_obj['cardName'] == 'simipic':
                    # 相似图片
                    for each_pic in each_obj['tplData']['list']:
                        rowdict = {
                            "fromPageTitle": each_pic['resultSrc'],  # 本字段实际为IMAGE，百度就没给title
                            "thumbURL": each_pic['thumbUrl'],
                            "fromURL": each_pic['fromUrl'],
                            "flowURL": each_pic['hoverUrl'],
                        }
                        similist_result.append(rowdict)

                elif each_obj['cardName'] == 'same':
                    # 相同图片
                    samelist_index = 0
                    pics = each_obj['extData']['showInfo']
                    for each_pic in pics['titles']:
                        rowdict = {
                            "thumbURL": pics['imgs_src'][samelist_index],
                            "fromURL": pics['snippets_url'][samelist_index],
                            "textHost": pics['snippets'][samelist_index],
                            "fromPageTitle": pics['titles'][samelist_index],
                            "flowURL": pics['imgs_url'][samelist_index],
                        }
                        samelist_result.append(rowdict)
                        samelist_index += 1
                elif each_obj['cardName'] == 'cardHeader':
                    # 猜测可能是什么
                    for each_subtitle in each_obj['tplData']['pano']['list']:
                        guessword.append(each_subtitle['subTitle'])

    except Exception as e:
        logger.error(traceback.format_exc())
        guessword = []
        samelist_result = []
        similist_result = []

    result = {
        "guessword": guessword,
        "samelist": samelist_result,
        "similist": similist_result,
        "same_in_site": same_in_site_list,
    }
    return {"c": SUCCESS[0], "m": SUCCESS[1], "d": result}


def api_public_work(account, work_id_list, public_status):
    # 修改作品公示状态
    work_id_list = json.loads(work_id_list)

    # 检查作品是否存在
    public_status = int(public_status)
    cur_works = Work.objects.filter(id__in=work_id_list, del_flag=FALSE_INT)
    if not cur_works or len(work_id_list) != len(cur_works):
        raise BusinessException(ERR_WORK_ID_ERROR)

    # 检查我是否是该活动的创建者，只有创建者可以修改作品公示状态
    for each_work in cur_works:
        if each_work.activity.user.account != account:
            raise BusinessException(AUTH_WRONG_TYPE)

    is_public = 1 if public_status else 0
    cur_works.update(is_public=is_public)

    return {"c": SUCCESS[0], "m": SUCCESS[1], "d": []}


def api_list_publicwork(user, activity_id, order, rows, page, last_id):
    result = dict()
    # 检查活动是否存在
    cur_activity = Activity.objects.filter(id=activity_id, del_flag=FALSE_INT).first()
    if not cur_activity:
        raise BusinessException(ERR_ACTIVITY_ID_ERROR)

    works = Work.objects.filter(is_public=TRUE_INT, activity_id=activity_id, del_flag=FALSE_INT)
    if order == '2':
        works = works.order_by('-like')
    else:
        works = works.order_by('-pv')

    # 分页
    cnt = len(works)
    num_pages = 1
    if rows and page:
        num_pages, cur_start, cur_end = get_pages(cnt, page, rows)
        works = works[cur_start:cur_end]
    elif rows:
        cur_start = get_lastid_index(works, last_id)
        works = works[cur_start:cur_start + int(rows)]
        pass

    result["max_page"] = num_pages
    result["total"] = cnt
    result["page"] = int(page) if page else 1
    work_list = list()
    for each_work in works:
        work_dict = {
            "work_id": each_work.id,
            "work_name": each_work.name,
            "work_authors": each_work.authors,
            "like": each_work.like,
            "i_like": '',  # 本字段废弃，后台不记录哪些人点赞了。只记录投票人数。点赞和投票功能分开。
            "vote": each_work.vote,
            "i_vote": 1 if is_vote(user.id, each_work.id) else 0,
            "pv": each_work.pv,
            "img_url": get_image_url(each_work.img_file.url) if each_work.img_file else '',
        }
        work_list.append(work_dict)
    result["work_list"] = work_list

    # 统计总作品数量及获奖作品数量
    total_works = Work.objects.filter(activity_id=activity_id, del_flag=FALSE_INT).count()
    rank_works = Work.objects.filter(ranks_id__isnull=False, activity_id=activity_id, del_flag=FALSE_INT).count()
    result["total_works"] = total_works
    result["rank_works"] = rank_works
    return {"c": SUCCESS[0], "m": SUCCESS[1], "d": result}


def is_vote(account_id, work_id):
    workvote = WorkVote.objects.filter(del_flag=FALSE_INT, account_id=account_id, work_id=work_id)
    return True if workvote else False


def debug_show_workflow(work):
    result = list()
    flows = WorkFlow.objects.filter(work=work).order_by('-create_time')
    for each in flows:
        if each.work_status == WORK_STATUS_HAVE_EXAMINED[0]:
            show_status = u'已结束审批，进入评审'
        else:
            show_status = const_main_status().dictionary()[each.work_status]
        result.append({
            'id': str(each.id),
            'trigger_id': str(each.trigger.id),
            'trigger_fullname': each.trigger_fullname,
            'event': each.event,
            'area_id': str(each.area.id),
            'area_name': each.area_name,
            'work_status': str(each.work_status),
            'work_status_desc': show_status
        })
    return result


def debug_approve_count(role):
    result = dict()
    result['from_workflow'] = how_many_approve(role)
    result['from_role_table'] = role.approve_work
    return result
