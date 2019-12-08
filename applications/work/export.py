# coding=utf-8
import logging
import json
import copy
import os
import re
import traceback
import types
from collections import namedtuple
from urlparse import urljoin

from django.conf import settings
from django.http import HttpResponse
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import quote_sheetname

from applications.activity.models import Rule
from applications.common.services import get_areadetail_by_id
from applications.score.models import Score, FinalScore
from applications.subjudge.models import SubJudgeRule, SubJudgeScore, SubJudgeTeamExpert, SubJudgeTeamWork
from applications.team.models import TeamExpert
from applications.work.models import WorkAttr, WorkAttrString
from utils.const_err import SUCCESS
from utils.net_helper import response200
from utils.public_fun import load_clazz, const_main_status, suuid

logger = logging.getLogger('task_trace')


Excel = namedtuple('Excel', [
    'BASIC', 'STATUS', 'AUTHOR', 'TUTOR', 'MY_SCORE', 'SCORE', 'FINAL_SCORE', 'SUBJG_MY_SCORE', 'SUBJG_SCORE', 'SUBJG_FINAL_SCORE'
])
EXCEL = Excel(
    'basic', 'status', 'author', 'tutor', 'my_score', 'score', 'final_score', 'subjg_my_score', 'subjg_score', 'subjg_final_score',
)

JUDGE_DESC = {
    'score': u'评分',
    'rank': u'等级',
    'comments': u'评语',
    'comment': u'评语',
}

ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
FONT_RED_BOLD = Font(b=True, color="FF0000")


class WorkExcelExport:
    data_choosed = None  # 选择导出哪些模组
    activity = None  # 活动
    score_dimentions = None  # 评价维度（分、奖项、评语）
    expert_count = None  # 专家数量
    st = None  # EXCEL
    work_list = None  # 导出的作品列表
    expert = None  # 当前专家
    subjudge = None
    col_index_def = None  # 每个导出模组的列序号
    task = None
    progress_all = 1
    progess_block = 0
    progress_finegrain = 0

    fix_col = [u"编号", u"名称", u"子活动", u"学段", u"项目", u"省或直辖市", u"市州", u"区县", u"学校或机构"]

    def __init__(self, activity, work_list, data_choosed, expert=None, subjudge=None, task=None):
        self.data_choosed = data_choosed
        self.activity = activity
        rule = Rule.objects.filter(activity=activity).first()
        dimention_count, dimentions = rule.parse_rule().dimention()
        self.score_dimentions = dimentions
        self.expert_count = rule.parse_rule().expert_count()
        self.expert = expert
        self.task = task

        self.subjudge = subjudge
        if subjudge:
            subjg_rule = SubJudgeRule.objects.filter(subjudge=subjudge).first()
            subjg_dimention_count, subjg_dimentions = subjg_rule.parse_rule().dimention()
            self.subjudge_score_dimentions = subjg_dimentions
            self.subjudge_expert_count = subjg_rule.parse_rule().expert_count()

        self.work_list = work_list
        self.wb = Workbook()
        self.st = self.wb.active

    def write(self):
        self.col_index_def = dict()
        accumulate = 0
        for each in self.data_choosed:
            func = getattr(self, 'calc_'+each, None)
            if isinstance(func, types.MethodType):
                col_num = func()
                self.col_index_def[each] = (accumulate, col_num)
                accumulate += col_num
            self.progress_all += len(self.work_list)

        for each in self.data_choosed:
            func = getattr(self, 'write_' + each, None)
            if isinstance(func, types.MethodType):
                func(*self.col_index_def[each])
        return self

    def update_progress(self, step=1):
        if self.task:
            self.progress_finegrain += step
            new_block = self.progress_finegrain * 100 / self.progress_all
            if new_block > self.progess_block:
                self.task.progress = str(new_block)
                self.task.save()
                self.progess_block = new_block

    def dump_stream(self):
        fname = u'%s.xlsx' % suuid()
        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename=' + fname.encode('utf-8')
        self.wb.save(response)
        return response

    def dump_file(self):
        path = os.path.join(settings.BASE_DIR, 'media/%s.xlsx' % suuid())
        self.wb.save(path)
        return {'c': SUCCESS[0], 'm': SUCCESS[1], 'd': path}

    def calc_basic(self):
        return WorkAttr.objects.filter(activity=self.activity, category=1).count() + len(self.fix_col)

    def write_basic(self, start_col, col_num):
        attr_1 = WorkAttr.objects.filter(activity=self.activity, category=1).values('id', 'name').order_by('create_time')
        # 浮动列
        float_col = [each['name'] for each in attr_1]
        float_map = {}
        for i, each in enumerate(attr_1):
            float_map[start_col+1+len(self.fix_col)+i] = each['id']
        # head
        for i, each in enumerate(self.fix_col + float_col):
            self.st.cell(column=start_col+1+i, row=2, value=each)
        # head merge
        cell = self.st.cell(column=start_col+1, row=1, value=u'基本信息')
        self.st.merge_cells(start_row=1, start_column=start_col+1, end_row=1, end_column=start_col+1+col_num-1)
        cell.alignment = ALIGN_CENTER
        cell.font = FONT_RED_BOLD
        # content
        for i, each in enumerate(self.work_list):
            area_detail = get_areadetail_by_id(each.area.id)
            value_map = {
                start_col + 1 + 0: each.no,
                start_col + 1 + 1: each.name,
                start_col + 1 + 2: each.sub_activity or '',
                start_col + 1 + 3: each.phase or '',
                start_col + 1 + 4: each.project or '',
                start_col + 1 + 5: area_detail['province'] or '',
                start_col + 1 + 6: area_detail['city'] or '',
                start_col + 1 + 7: area_detail['country'] or '',
                start_col + 1 + 8: area_detail['institution'] or '',
            }
            attr_str_all = WorkAttrString.objects.filter(work=each).values('attr__id', 'value')
            attr_str_dict = {each['attr__id'] : each['value'] for each in attr_str_all}
            for j, f in enumerate(float_col):
                c = start_col + 1 + len(self.fix_col) + j
                value_map[c] = attr_str_dict[float_map[c]] if float_map[c] in attr_str_dict else ''
            for k, _ in enumerate(self.fix_col + float_col):
                c = start_col + 1 + k
                self.st.cell(column=c, row=3+i, value=value_map[c])
            logger.info('<export> handled basic info for work %s, progress: %s / %s' %
                        (each.id, i+1, len(self.work_list)))
            self.update_progress()
        logger.info('<export> finish basic info handling')

    def calc_status(self):
        return 1

    def write_status(self, start_col, col_num):
        fix_col = [u"所处阶段", ]
        # head
        self.st.cell(column=start_col + 1, row=2, value=fix_col[0])
        # head merge
        cell = self.st.cell(column=start_col+1, row=1, value=u'所处阶段')
        self.st.merge_cells(start_row=1, start_column=start_col+1, end_row=1, end_column=start_col+1+col_num-1)
        cell.alignment = ALIGN_CENTER
        cell.font = FONT_RED_BOLD
        # content
        for i, each in enumerate(self.work_list):
            status = const_main_status().dictionary()[each.status]
            self.st.cell(column=start_col + 1, row=3 + i, value=status)
            logger.info('<export> handled status info for work %s, progress: %s / %s' %
                        (each.id, i+1, len(self.work_list)))
            self.update_progress()
        logger.info('<export> finish status info handling')

    def calc_author(self):
        count = WorkAttr.objects.filter(activity=self.activity, category=2, group_sn=1).count()
        return self.activity.author_count * count

    def write_author(self, start_col, col_num):
        author_attr_def = WorkAttr.objects.filter(activity=self.activity, category=2).order_by('group_sn', 'sn')
        col_def = {start_col + 1 + i : each.id for i, each in enumerate(author_attr_def)}
        # head
        for i, each in enumerate(author_attr_def):
            self.st.cell(column=start_col + 1 + i, row=2, value=each.name)
        # head merge
        author_num = self.activity.author_count
        attr_count = WorkAttr.objects.filter(activity=self.activity, category=2, group_sn=1).count()
        author_head_starter = start_col+1
        for i in xrange(author_num):
            cell = self.st.cell(column=author_head_starter, row=1, value=u'作者%s'%(i+1))
            self.st.merge_cells(start_row=1, start_column=author_head_starter, end_row=1, end_column=author_head_starter+attr_count-1)
            cell.alignment = ALIGN_CENTER
            cell.font = FONT_RED_BOLD
            author_head_starter += attr_count
        # content
        for i, each in enumerate(self.work_list):
            work_attr_str = WorkAttrString.objects.filter(work=each).all()
            work_attr_dict = {was.attr.id : was.value for was in work_attr_str}
            for j in xrange(len(author_attr_def)):
                col_index = start_col + 1 + j
                self.st.cell(column=col_index, row=3+i,
                             value=work_attr_dict[col_def[col_index]] if col_def[col_index] in work_attr_dict else '')
            logger.info('<export> handled author info for work %s, progress: %s / %s' %
                        (each.id, i+1, len(self.work_list)))
            self.update_progress()
        logger.info('<export> finish author info handling')

    def calc_tutor(self):
        count = WorkAttr.objects.filter(activity=self.activity, category=3, group_sn=1).count()
        return self.activity.tutor_count * count

    def write_tutor(self, start_col, col_num):
        author_attr_def = WorkAttr.objects.filter(activity=self.activity, category=3).order_by('group_sn', 'sn')
        col_def = {start_col + 1 + i : each.id for i, each in enumerate(author_attr_def)}
        # head
        for i, each in enumerate(author_attr_def):
            self.st.cell(column=start_col + 1 + i, row=2, value=each.name)
        # head merge
        tutor_num = self.activity.tutor_count
        attr_count = WorkAttr.objects.filter(activity=self.activity, category=3, group_sn=1).count()
        tutor_head_starter = start_col+1
        for i in xrange(tutor_num):
            cell = self.st.cell(column=tutor_head_starter, row=1, value=u'指导老师%s'%(i+1))
            self.st.merge_cells(start_row=1, start_column=tutor_head_starter, end_row=1, end_column=tutor_head_starter+attr_count-1)
            cell.alignment = ALIGN_CENTER
            cell.font = FONT_RED_BOLD
            tutor_head_starter += attr_count
        # content
        for i, each in enumerate(self.work_list):
            work_attr_str = WorkAttrString.objects.filter(work=each).all()
            work_attr_dict = {was.attr.id : was.value for was in work_attr_str}
            for j in xrange(len(author_attr_def)):
                col_index = start_col + 1 + j
                self.st.cell(column=col_index, row=3+i,
                             value=work_attr_dict[col_def[col_index]] if col_def[col_index] in work_attr_dict else '')
            logger.info('<export> handled tutor info for work %s, progress: %s / %s' %
                        (each.id, i+1, len(self.work_list)))
            self.update_progress()
        logger.info('<export> finish tutor info handling')

    def calc_my_score(self):
        return len(self.score_dimentions)

    def _score_data_by_dimention(self, score, dmtn):
        value = ''
        if not score:
            return value
        if dmtn == 'score':
            value = str(score.score)
        if dmtn == 'rank':
            value = score.rank.name if score.rank else ''
        if dmtn == 'comments':
            value = score.comments
        return value

    def _subjudge_score_data_by_dimention(self, subjudge_score, dmtn):
        value = ''
        if not subjudge_score:
            return value
        if dmtn == 'score':
            value = str(subjudge_score.score)
        if dmtn == 'rank':
            value = subjudge_score.rank.name if subjudge_score.rank else ''
        if dmtn == 'comment':
            value = subjudge_score.comment
        return value

    def write_my_score(self, start_col, col_num):
        # head
        for i, each in enumerate(self.score_dimentions):
            self.st.cell(column=start_col + 1 + i, row=2, value=JUDGE_DESC[each])
        # head merge
        cell = self.st.cell(column=start_col+1, row=1, value=u'我的评分')
        self.st.merge_cells(start_row=1, start_column=start_col+1, end_row=1, end_column=start_col+1+col_num-1)
        cell.alignment = ALIGN_CENTER
        cell.font = FONT_RED_BOLD
        # content
        for i, each in enumerate(self.work_list):
            my_score = Score.objects.filter(work=each, team_expert__expert=self.expert).first()
            for j, dmtn in enumerate(self.score_dimentions):
                value = ''
                if self.expert and my_score:
                    value = self._score_data_by_dimention(my_score, dmtn)
                self.st.cell(column=start_col + 1 + j, row=3 + i, value=value)
            logger.info('<export> handled my_score info for work %s, progress: %s / %s' %
                        (each.id, i+1, len(self.work_list)))
            self.update_progress()
        logger.info('<export> finish my_score info handling')

    def calc_score(self):
        return self.expert_count * len(self.score_dimentions)

    def _score_data(self, work, sn, dmtn):
        te = TeamExpert.objects.filter(team=work.team, sn=int(sn)).first()
        if not te:
            return ''
        score = Score.objects.filter(work=work, team_expert=te).first()
        value = self._score_data_by_dimention(score, dmtn)
        return value

    def write_score(self, start_col, col_num):
        cell_data_def = dict()
        # head
        for i in xrange(self.expert_count):
            for j, each in enumerate(self.score_dimentions):
                key = start_col + 1 + i*len(self.score_dimentions) + j
                self.st.cell(column=key, row=2, value=JUDGE_DESC[each])
                cell_data_def[key] = '%s,%s' % (i+1, each)
        # head merge
        attr_count = len(self.score_dimentions)
        expert_head_starter = start_col+1
        for i in xrange(self.expert_count):
            cell = self.st.cell(column=expert_head_starter, row=1, value=u'专家%s评分'%(i+1))
            self.st.merge_cells(start_row=1, start_column=expert_head_starter, end_row=1, end_column=expert_head_starter+attr_count-1)
            cell.alignment = ALIGN_CENTER
            cell.font = FONT_RED_BOLD
            expert_head_starter += attr_count
        # content
        for i, each in enumerate(self.work_list):
            for j in xrange(col_num):
                sn = cell_data_def[start_col+1+j].split(',')[0]
                dmtn = cell_data_def[start_col+1+j].split(',')[1]
                self.st.cell(column=start_col + 1 + j, row=3 + i, value=self._score_data(each, sn, dmtn))
            logger.info('<export> handled score info for work %s, progress: %s / %s' %
                        (each.id, i+1, len(self.work_list)))
            self.update_progress()
        logger.info('<export> finish score info handling')

    def calc_final_score(self):
        return 3

    def write_final_score(self, start_col, col_num):
        # head
        self.st.cell(column=start_col + 1 + 0, row=2, value=JUDGE_DESC['score'])
        self.st.cell(column=start_col + 1 + 1, row=2, value=JUDGE_DESC['rank'])
        self.st.cell(column=start_col + 1 + 2, row=2, value=JUDGE_DESC['comments'])
        # head merge
        cell = self.st.cell(column=start_col+1, row=1, value=u'作品最终得分')
        self.st.merge_cells(start_row=1, start_column=start_col+1, end_row=1, end_column=start_col+1+col_num-1)
        cell.alignment = ALIGN_CENTER
        cell.font = FONT_RED_BOLD
        # content
        for i, each in enumerate(self.work_list):
            final_score = FinalScore.objects.filter(work=each).first()
            if final_score:
                rk = final_score.rank.name if final_score.rank else ''
                self.st.cell(column=start_col + 1 + 0, row=3 + i, value=str(final_score.score))
                self.st.cell(column=start_col + 1 + 1, row=3 + i, value=rk)
                self.st.cell(column=start_col + 1 + 2, row=3 + i, value=final_score.comments)
            logger.info('<export> handled final_score info for work %s, progress: %s / %s' %
                        (each.id, i+1, len(self.work_list)))
            self.update_progress()
        logger.info('<export> finish final_score info handling')

    def calc_subjg_my_score(self):
        return len(self.subjudge_score_dimentions)

    def write_subjg_my_score(self, start_col, col_num):
        # head
        for i, each in enumerate(self.subjudge_score_dimentions):
            self.st.cell(column=start_col + 1 + i, row=2, value=JUDGE_DESC[each])
        # head merge
        cell = self.st.cell(column=start_col+1, row=1, value=u'我的评分(子级评审)')
        self.st.merge_cells(start_row=1, start_column=start_col+1, end_row=1, end_column=start_col+1+col_num-1)
        cell.alignment = ALIGN_CENTER
        cell.font = FONT_RED_BOLD
        # content
        for i, each in enumerate(self.work_list):
            my_score = SubJudgeScore.objects.filter(
                subjudge_team__subjudge=self.subjudge, subjudge_team_work__work=each, subjudge_team_expert__expert=self.expert).first()
            for j, dmtn in enumerate(self.score_dimentions):
                value = ''
                if self.expert and my_score:
                    value = self._subjudge_score_data_by_dimention(my_score, dmtn)
                self.st.cell(column=start_col + 1 + j, row=3 + i, value=value)
            logger.info('<export> handled subjudge_my_score info for work %s, progress: %s / %s' %
                        (each.id, i+1, len(self.work_list)))
            self.update_progress()
        logger.info('<export> finish subjudge_my_score info handling')

    def calc_subjg_score(self):
        return self.subjudge_expert_count * len(self.subjudge_score_dimentions)

    def _subjudge_score_data(self, work, sn, dmtn):
        subjg_team_work = SubJudgeTeamWork.objects.filter(subjudge=self.subjudge, work=work).first()
        if not subjg_team_work:
            return ''
        subjg_te = SubJudgeTeamExpert.objects.filter(subjudge_team=subjg_team_work.subjudge_team, sn=int(sn)).first()
        if not subjg_te:
            return ''
        score = SubJudgeScore.objects.filter(subjudge_team_work=subjg_team_work, subjudge_team_expert=subjg_te).first()
        value = self._subjudge_score_data_by_dimention(score, dmtn)
        return value

    def write_subjg_score(self, start_col, col_num):
        cell_data_def = dict()
        # head
        for i in xrange(self.subjudge_expert_count):
            for j, each in enumerate(self.subjudge_score_dimentions):
                key = start_col + 1 + i*len(self.subjudge_score_dimentions) + j
                self.st.cell(column=key, row=2, value=JUDGE_DESC[each])
                cell_data_def[key] = '%s,%s' % (i+1, each)
        # head merge
        attr_count = len(self.subjudge_score_dimentions)
        expert_head_starter = start_col+1
        for i in xrange(self.expert_count):
            cell = self.st.cell(column=expert_head_starter, row=1, value=u'专家%s评分(子级评审)'%(i+1))
            self.st.merge_cells(start_row=1, start_column=expert_head_starter, end_row=1, end_column=expert_head_starter+attr_count-1)
            cell.alignment = ALIGN_CENTER
            cell.font = FONT_RED_BOLD
            expert_head_starter += attr_count
        # content
        for i, each in enumerate(self.work_list):
            for j in xrange(col_num):
                sn = cell_data_def[start_col+1+j].split(',')[0]
                dmtn = cell_data_def[start_col+1+j].split(',')[1]
                self.st.cell(column=start_col + 1 + j, row=3 + i, value=self._subjudge_score_data(each, sn, dmtn))
            logger.info('<export> handled subjudge_score info for work %s, progress: %s / %s' %
                        (each.id, i+1, len(self.work_list)))
            self.update_progress()
        logger.info('<export> finish subjudge_score info handling')

    def calc_subjg_final_score(self):
        return 0

    def write_subjg_final_score(self, start_col, col_num):
        pass



